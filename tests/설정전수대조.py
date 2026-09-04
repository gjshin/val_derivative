#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""설정을 하나씩 바꿔가며 수식 조서가 엔진을 따라오는지 전수 확인한다.

`조서대조.py` 가 대표 시나리오 몇 개를 깊게 본다면, 이쪽은 **모든 설정**을
한 번씩 흔들어 놓친 배선이 없는지 넓게 본다. 설정을 바꿨는데 조서가 안 따라오면
그 설정을 참조하지 않는 시트가 어딘가 있다는 뜻이다.

두 가지를 함께 본다.

    1. 조서 값 == 엔진 값        — 배선이 맞는가
    2. 조서 값이 기본값과 다른가  — 그 설정이 실제로 결과를 움직이는가

2번이 없으면 "우연히 안 쓰이는 설정"을 통과시켜 버린다. 결과가 안 움직이는
설정은 목록에 이유를 적어 둔다.

    pip install formulas
    python3 tests/설정전수대조.py
"""
import sys, os, re, types, warnings, tempfile
warnings.filterwarnings("ignore")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "tests"))

# (라벨, 덮어쓸 값, 결과가 안 움직여도 되는 이유 — None 이면 움직여야 한다)
CASES = [
    ("기본", {}, "기준선"),
    # 주가·전환가액
    ("주가 +40%", dict(S0=1194.2), None),
    ("전환가액 +20%", dict(K0=1023.6), None),
    # 일자·격자
    ("중간평가", dict(d_base="2026-03-31"), None),
    ("노드를 촘촘히", dict(_gap=3.), None),
    # 이자
    ("표면 3%", dict(cpn=.03), None),
    ("이자 지급 반기", dict(cpn=.03, ipay=6.), None),
    ("만기보장 7%", dict(ytm=.07), None),
    ("보장 복리 연 1회", dict(ytm=.07, ytm_cmp=1), None),
    # 전환
    ("전환 시작 24개월", dict(cv_s=24., _gap=3.), None),
    ("전환 종료 48개월", dict(cv_e=48.), None),
    # 리픽싱 — 성긴 격자에서는 주기가 모두 1스텝으로 뭉개져 _gap 을 준다
    ("리픽싱 없음", dict(rfx_mode=0), None),
    # 주기 7개월·분기 노드에서는 상향 허용 여부가 결과를 안 바꾼다(엔진도 같은 값).
    # 12개월로 늘리면 1.49 만큼 갈린다.
    ("하향만 리픽싱", dict(rfx_mode=1, rfx_cyc=12., _gap=3.), None),
    ("리픽싱 주기 3개월", dict(rfx_cyc=3., _gap=3.), None),
    ("최저 조정가액 700", dict(floor=700.), None),
    ("액면가 750", dict(floor=700., par=750.), None),
    # 조정 주기가 2스텝이면 사이 칸이 하나뿐이라 세 근사가 수학적으로 같아진다.
    # 12개월로 늘려 4스텝을 만들어야 갈라진다.
    ("조정일 처리 · 경로가중", dict(rfx_cyc=12., carry=1, _gap=3.), None),
    ("조정일 처리 · 확률가중", dict(rfx_cyc=12., carry=2, _gap=3.), None),
    ("조정일 처리 · 특정노드", dict(rfx_cyc=12., carry=3, _gap=3.), None),
    # 조기상환
    ("조기상환 시작 12개월", dict(p_s=12.), None),
    ("조기상환 종료 36개월", dict(p_e=36., _gap=3.), None),
    ("조기상환 주기 9개월", dict(p_f=9., _gap=3.), None),
    ("조기상환 주기 6개월", dict(p_f=6., _gap=3.),
     "행사금액이 고정 100 이면 행사일이 드물어도 값이 거의 같다. 각 행사일에서 "
     "MAX(계속보유, 100) 이라 값이 100 아래로 못 내려가고, 첫 행사일이 하한을 "
     "정한 뒤로는 빈도가 거의 영향을 주지 않는다. 9개월로 늘리면 움직인다."),
    ("조기상환 행사금액 105", dict(p_rate=105.), None),
    ("조기상환 보장 5%", dict(p_mode="accrue", p_yield=.05), None),
    ("보장 복리 연 2회", dict(p_mode="accrue", p_yield=.05, p_cmp=2), None),
    # 매도청구
    ("매도청구 시작 6개월", dict(k_s=6., k_lock=0., _gap=3.), None),
    ("매도청구 종료 36개월", dict(k_e=36., k_lock=0., _gap=3.), None),
    # 의무보유가 매도청구 종료보다 늦으면 첫 행사일에 콜이 확정되어 주기가
    # 결과를 못 움직인다. 주기 배선을 보려면 의무보유를 풀어야 한다.
    ("매도청구 주기 9개월 · 의무보유 없음", dict(k_f=9., k_lock=0., _gap=3.), None),
    ("매도청구 프리미엄 6%", dict(k_prem=.06), None),
    ("매도청구 복리 연 1회", dict(k_prem=.06, k_cmp=1), None),
    ("매도청구 한도 60%", dict(k_w=.6), None),
    ("의무보유 40개월", dict(k_lock=40.),
     "매도청구 종료가 24개월인데 의무보유 25개월(기본)이 이미 콜 기간 전체를 "
     "덮는다. 그 이상 늘려도 콜 기간에는 어차피 전환할 수 없어 값이 같다."),
    ("평가방법 1 혼합할인율", dict(k_method=1), None),
    ("평가방법 2 지분·부채", dict(k_method=2), None),
    # 모형·분류
    ("변동성 25%", dict(sig=.25), None),
    ("GS", dict(model="GS"), None),
    ("조기상환권 미분리", dict(p_sep=0),
     "트랜치·주계약·부채요소·매도청구권은 그대로다. 바뀌는 것은 배분표와 분개이고 "
     "그 둘은 매 케이스마다 따로 확인한다."),
    ("조기상환권 미분리 · 방법1", dict(p_sep=0, k_method=1),
     "위와 같다. 배분표만 갈린다."),
    ("콜 내재파생 포함", dict(k_sep=0),
     "트랜치·주계약·부채요소·매도청구권은 그대로다. 바뀌는 것은 배분표와 분개이고 "
     "그 둘은 매 케이스마다 따로 확인한다."),
    ("콜 내재파생 포함 · 부채", dict(k_sep=0, conv_class="liability"),
     "위와 같다. 전환권 분류까지 바꿔도 트랜치 다섯 값은 움직이지 않는다."),
    ("전환권 부채", dict(conv_class="liability"),
     "트랜치·주계약·부채요소·매도청구권은 그대로다. 바뀌는 것은 배분표와 분개이고 "
     "그 둘은 매 케이스마다 따로 확인한다."),
    # 곡선
    ("무위험 +2%p", dict(_rf=.02), None),
    ("신용 +3%p", dict(_cr=.03), None),
    ("무위험 복리 연 4회", dict(cmp_rf=4), None),
    ("신용 복리 연 2회", dict(cmp_cr=2), None),
    ("현물 곡선 입력", dict(y_type="spot"), None),
    # 매도청구 주기가 조기상환 주기와 다른 경우 — 예전에 한쪽을 잘못 참조했다
    ("매도청구 주기 ≠ 조기상환", dict(k_f=9., p_f=6., k_e=36., k_lock=0., _gap=3.), None),
    ("중간평가 · 분기 노드", dict(d_base="2026-03-31", _gap=3.), None),
    ("중간평가 · 리픽싱 12개월", dict(d_base="2026-03-31", rfx_cyc=12., _gap=3.), None),
    # 결과를 안 움직이는 것이 정상인 설정
    ("전자등록총액 500억", dict(face_total=5e10),
     "100 기준 결과는 그대로다. 전액 기준 환산 열만 바뀐다."),
]


def load_app():
    stub = types.ModuleType("streamlit"); stub.cache_data = lambda **k: (lambda f: f)
    sys.modules["streamlit"] = stub
    m = types.ModuleType("cbapp"); sys.modules["cbapp"] = m
    src = open(os.path.join(ROOT, "cb_app.py"), encoding="utf-8").read()
    exec(compile(src.split("st.set_page_config")[0], "cb_app.py", "exec"), m.__dict__)
    return m.__dict__


def terms(G, over):
    t = G["Terms"]()
    rf, cr = over.get("_rf", 0.), over.get("_cr", 0.)
    t.rf_curve = [(1, .0226+rf), (3, .0240+rf), (5, .0252+rf)]
    t.cr_curve = [(1, .1409+cr), (3, .1740+cr), (5, .1905+cr)]
    for k, v in over.items():
        if not k.startswith("_"): setattr(t, k, v)
    # 계산 시간을 줄인다. 구조는 같다. gap_m 은 반드시 _gap 으로만 준다.
    assert "gap_m" not in over, "격자는 _gap 으로 지정하십시오"
    t.gap_m = over.get("_gap", 6.0)
    if "carry" not in over: t.carry = 1     # 상태확장은 조서로 못 옮긴다
    G["derive"](t)
    return t


def _combin(f):
    """formulas 는 COMBIN 을 구현하지 않는다. 값으로 바꿔 우회한다.

    조서 자체는 정상이다. 엑셀은 COMBIN 을 계산한다. 검사 도구의 한계라
    여기서만 우회한다.

    도달확률은 스텝(4행)과 하락 횟수(B열)를 참조하는 동적 형태다.
    COMBIN(<열>$4, $B<행>) 의 열·행에서 스텝과 r 을 되짚는다. 열 C 가 스텝 0,
    5행이 r=0 이다.
    """
    import math
    from openpyxl.utils import column_index_from_string as ci

    f = re.sub(r"COMBIN\(([A-Z]+)\$4,\$B(\d+)\)",
               lambda m: repr(math.comb(ci(m.group(1)) - 3, int(m.group(2)) - 5)), f)
    return re.sub(r"COMBIN\((\d+),(\d+)\)",
                  lambda m: repr(math.comb(int(m.group(1)), int(m.group(2)))), f)


def build(G, over, path):
    t = terms(G, over)
    full, b0, b1, b2, ca, conv = G["decompose"](t)
    b3 = G["pick"](G["engine"](t, conv=True, put=True, call=True,
                               conv_start=max(t.cv_s, t.k_lock)), t.model)
    open(path, "wb").write(
        G["build_xlsx_formula"](t, full, b0, b1, b2, ca, conv,
                                G["eir_table"](t, G["acc_host"](t, full, b0, b1, b2, ca))))
    import openpyxl
    wb = openpyxl.load_workbook(path)
    mp = {nm: f"S{i:02d}" for i, nm in enumerate(wb.sheetnames)}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    f = c.value
                    for o, nn in sorted(mp.items(), key=lambda x: -len(x[0])):
                        f = f.replace(f"'{o}'!", f"{nn}!").replace(f"{o}!", f"{nn}!")
                    c.value = _combin(f)
    for o, nn in mp.items(): wb[o].title = nn
    wb.save(path)
    # 분개 차변은 배분표의 음수 항목이 넘어온 것이다. 콜을 내재파생에 넣으면
    # 자산 줄이 사라져 100 이 되고, 파생 순액이 자산이면 그 줄이 대신 차변으로 간다.
    al = G["allocate"](t, full, b0, b1, b2, ca)[0]
    return dict(b2=b2, b3=b3, b0=b0, b1=b1, ca=ca, conv=conv,
                dr_want=100 + sum(-v for _, v in al[:-1] if v < 0),
                eq=(t.conv_class == "equity")), mp["결과"], mp["회계처리"]


def solve(path, sheets):
    import formulas
    sol = formulas.ExcelModel().loads(path).finish().calculate()
    base = os.path.basename(path).upper()
    out = {nm: {} for nm in sheets}
    for k, v in sol.items():
        ku = k.upper()
        for nm in sheets:
            if ku.startswith(f"'[{base}]{nm}'!"):
                try: out[nm][ku.split("!")[-1]] = float(v.value[0, 0])
                except Exception: pass
    return out


# 조서에는 앱에서 고른 방법만 들어간다. 30% 트랜치는 유무가치비교법에만 있다.
ROWS = [("콜 없는 트랜치", "C10", "b2"),
        ("주계약", "C16", "b0"), ("부채요소", "C17", "b1"),
        ("매도청구권 적용값", "C22", "ca")]


def run_one(idx):
    """케이스 하나를 재고 결과를 JSON 한 줄로 뱉는다. 별도 프로세스에서 돈다.

    formulas 는 한 번 풀 때마다 메모리를 크게 물고 놓지 않아, 한 프로세스에서
    수십 케이스를 이어 돌리면 죽는다. 그래서 케이스마다 프로세스를 새로 띄운다.
    """
    import json
    G = load_app()
    lbl, over, why = CASES[idx]
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "wb.xlsx")
        eng, res, acc = build(G, over, path)
        got = solve(path, (res, acc))
    out = {"lbl": lbl, "why": why, "ca": eng["dr_want"],
           "vals": [got[res].get(c) for _, c, _ in ROWS],
           "want": [eng[k] for _, _, k in ROWS],
           "alloc": got[acc].get("C13"),
           "dr": got[acc].get("C25"), "cr": got[acc].get("D25"),
           "gap": over.get("_gap", 6.0)}
    print("@@" + json.dumps(out), flush=True)
    return 0


def main():
    import json, subprocess
    bad, dead, base = [], [], {}      # 기준선은 격자별로 따로 잡는다
    # 열은 ROWS 를 따라간다. 조서에 실리는 항목이 바뀌면 여기도 같이 바뀐다.
    _hdr = "%-22s" + " %10s"*len(ROWS) + "  %s"
    print(_hdr % ("설정", *[nm for nm, _, _ in ROWS], "판정"))
    for idx in range(len(CASES)):
        r = subprocess.run([sys.executable, os.path.abspath(__file__), str(idx)],
                           capture_output=True, text=True)
        line = [x for x in r.stdout.splitlines() if x.startswith("@@")]
        if not line:
            lbl = CASES[idx][0]
            bad.append(f"{lbl}: 계산 실패 — {r.stderr.strip().splitlines()[-1:] }")
            print("%-22s %s" % (lbl, "★ 계산 실패"))
            continue
        o = json.loads(line[0][2:])
        lbl, why, eng_ca = o["lbl"], o["why"], o["ca"]
        vals, ok = o["vals"], True
        for (nm, _, _), have, want in zip(ROWS, vals, o["want"]):
            if have is None or abs(have - want) > 1e-4:
                ok = False
                bad.append(f"{lbl} · {nm}: 조서 {have} · 엔진 {want:.4f}")
        # 배분 합계와 분개 대차도 매번 본다
        if o["alloc"] is None or abs(o["alloc"] - 100.0) > 1e-4:
            ok = False; bad.append(f"{lbl} · 배분 합계: {o['alloc']}")
        if (o["dr"] is None or o["cr"] is None
                or abs(o["dr"] - o["cr"]) > 1e-4
                or abs(o["dr"] - eng_ca) > 1e-4):
            ok = False; bad.append(f"{lbl} · 분개 대차: {o['dr']} / {o['cr']}")
        g = o["gap"]
        if g not in base:
            base[g] = vals                     # 그 격자의 첫 케이스가 기준선이다
            moved = True
        else:
            moved = any(v is not None and b is not None and abs(v - b) > 1e-6
                        for v, b in zip(vals, base[g]))
        mark = "" if ok else "★불일치"
        if ok and not moved and why is None:
            dead.append(lbl); mark = "★안 움직임"
        elif ok and not moved:
            mark = "동일 (정상)"
        print(_hdr % (lbl, *[f"{v:.4f}" if v is not None else "없음" for v in vals], mark))
    print()
    if bad:
        print("불일치 %d건" % len(bad))
        for b in bad: print("   ★ " + b)
    if dead:
        print("결과가 움직이지 않은 설정 %d건 — 조서가 참조하지 않을 수 있다" % len(dead))
        for d in dead: print("   ★ " + d)
    if not bad and not dead: print("전 설정 일치 · 모든 설정이 결과를 움직인다")
    return 0 if not bad and not dead else 1


if __name__ == "__main__":
    sys.exit(run_one(int(sys.argv[1])) if len(sys.argv) > 1 else main())
