#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""값 조서와 수식 조서가 같은 답을 내는지 확인한다.

수식 조서를 실제로 계산해 엔진 값과 맞춰본다. 트리 수식을 손으로 고치면
엔진과 어긋나기 쉬워서, 고칠 때마다 이 스크립트를 돌려야 한다.

    pip install formulas
    python3 tests/조서대조.py

시트 이름에 %와 한글이 있으면 formulas 가 깨져서 ASCII 로 바꾼 사본을 만들어 푼다.
노드를 10개로 줄여 계산 시간을 줄인다. 구조가 같으므로 검증에는 충분하다.
"""
import sys, os, re, types, json, warnings, tempfile
warnings.filterwarnings("ignore")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

CASES = [
    ("기본 (티사이언티픽)", {}),
    ("리픽싱 없음", dict(rfx_mode=0)),
    ("하향만 리픽싱", dict(rfx_mode=1)),
    ("표면 3% · 보장 7%",
     dict(cpn=.03, ytm=.07, ipay=6., ytm_cmp=2, p_mode="accrue", p_yield=.07)),
    ("표면 8% · 매도청구 5%",
     dict(cpn=.08, ytm=.08, ipay=6., ytm_cmp=2, p_mode="accrue", p_yield=.08, k_prem=.05)),
    ("중간평가 · 한도 100%", dict(d_base="2025-12-31", k_w=1.0)),
    ("GS", dict(model="GS")),
    ("GS · 전환권 부채", dict(model="GS", conv_class="liability")),
    ("GS · 옵션차익혼합할인법", dict(model="GS", k_method=1)),
    ("전환권 부채", dict(conv_class="liability")),
    ("방법2 지분·부채 분리", dict(k_method=2)),
    ("방법2 · 전환권 부채", dict(k_method=2, conv_class="liability")),
    ("조기상환권 미분리", dict(p_sep=0)),
    ("조기상환권 미분리 · 방법1", dict(p_sep=0, k_method=1)),
    ("조기상환권 미분리 · 콜 내재파생 (스위치 꺼짐)", dict(p_sep=0, k_sep=0)),
    ("조기상환권 미분리 · 전환권 부채 (스위치 꺼짐)",
     dict(p_sep=0, conv_class="liability")),
    ("콜 내재파생 포함 · 자본", dict(k_sep=0)),
    ("콜 내재파생 포함 · 부채", dict(k_sep=0, conv_class="liability")),
    # 아래 세 건은 리픽싱 주기가 2 스텝 이상이라야 살아나는 경로다.
    # 다른 케이스는 노드가 성겨 주기가 1 스텝으로 뭉개지고, 그러면
    # 이월 계산도 조정일 동점도 생기지 않아 오류를 통째로 비껴간다.
    ("월 노드 · 리픽싱 7스텝", dict(_gap=1., d_issue="2025-05-23", d_base="2025-06-30",
                              d_mat="2027-08-23", cv_s=12., cv_e=26., p_s=12., p_e=24.,
                              k_s=6., k_e=18., k_lock=20., rfx_cyc=7.)),
    ("월 노드 · 리픽싱 7스텝 · 방법1 · 부채",
     dict(_gap=1., k_method=1, conv_class="liability", d_issue="2025-05-23",
          d_base="2025-06-30", d_mat="2027-08-23", cv_s=12., cv_e=26., p_s=12.,
          p_e=24., k_s=6., k_e=18., k_lock=20., rfx_cyc=7.)),
    # 조기상환권을 BDT 금리격자로 잴 때. 자본·TF 에서만 열린다.
    ("BDT σ=20% · 위험곡선", dict(put_bdt=1, bdt_sig=.20)),
    ("BDT σ=20% · 무위험+스프레드", dict(put_bdt=1, bdt_sig=.20, bdt_base=1)),
    ("BDT σ=0 · 확정 격자와 같아야", dict(put_bdt=1, bdt_sig=0.0)),
    ("BDT 등가격 (보장=할인)",
     dict(put_bdt=1, bdt_sig=.20, S0=450., ytm=.08, ytm_cmp=4,
          p_mode="accrue", p_yield=.08, p_cmp=4,
          _cr=[(0.25, .08), (1, .08), (3, .08), (5, .08)])),
    ("BDT · 표면3% 보장7%",
     dict(put_bdt=1, bdt_sig=.20, cpn=.03, ytm=.07, ipay=6., ytm_cmp=2,
          p_mode="accrue", p_yield=.07)),
    ("월 노드 · 리픽싱 7스텝 · GS",
     dict(_gap=1., model="GS", d_issue="2025-05-23", d_base="2025-06-30",
          d_mat="2027-08-23", cv_s=12., cv_e=26., p_s=12., p_e=24.,
          k_s=6., k_e=18., k_lock=20., rfx_cyc=7.)),
]


def load_app():
    stub = types.ModuleType("streamlit"); stub.cache_data = lambda **k: (lambda f: f)
    sys.modules["streamlit"] = stub
    m = types.ModuleType("cbapp"); sys.modules["cbapp"] = m
    src = open(os.path.join(ROOT, "cb_app.py"), encoding="utf-8").read()
    exec(compile(src.split("st.set_page_config")[0], "cb_app.py", "exec"), m.__dict__)
    return m.__dict__


def _combin(f):
    """formulas 는 COMBIN 을 구현하지 않는다. 값으로 바꿔 우회한다.

    조서 자체는 정상이다. 엑셀은 COMBIN 을 계산한다. 검사 도구의 한계라
    여기서만 우회한다.

    도달확률은 스텝(4행)과 하락 횟수(B열)를 참조하는 동적 형태다.
    COMBIN(<열>$4, $B<행>) 의 열·행에서 스텝과 r 을 되짚어 상수로 만든다.
    열 C 가 스텝 0, 5행이 r=0 이다.
    """
    import math
    from openpyxl.utils import column_index_from_string as ci

    def dyn(m):
        return repr(math.comb(ci(m.group(1)) - 3, int(m.group(2)) - 5))

    f = re.sub(r"COMBIN\(([A-Z]+)\$4,\$B(\d+)\)", dyn, f)
    return re.sub(r"COMBIN\((\d+),(\d+)\)",
                  lambda m: repr(math.comb(int(m.group(1)), int(m.group(2)))), f)


def build(G, over, path):
    T, derive, decompose = G["Terms"], G["derive"], G["decompose"]
    t = T(); t.rf_curve = [(1, .0226), (3, .0240), (5, .0252)]
    t.cr_curve = [(1, .1409), (3, .1740), (5, .1905)]
    # _gap 은 노드 간격만 바꾸는 검사용 키다. 리픽싱 주기가 살아나는
    # 촘촘한 격자를 만들 때 쓴다.
    t.carry = 1; t.gap_m = over.get("_gap", 6.0)
    if "_cr" in over: t.cr_curve = over["_cr"]
    for k, v in over.items():
        if k not in ("_gap", "_cr"): setattr(t, k, v)
    derive(t)
    full, b0, b1, b2, ca, conv = decompose(t)
    b3 = G["pick"](G["engine"](t, conv=True, put=True, call=True,
                               conv_start=max(t.cv_s, t.k_lock)), t.model)
    ctp1 = G["call_third_party"](t, full, 1)
    ctp2 = G["call_third_party"](t, full, 2)
    open(path, "wb").write(
        G["build_xlsx_formula"](t, full, b0, b1, b2, ca, conv,
                                G["eir_table"](t, G["acc_host"](t, full, b0, b1, b2, ca))))

    import openpyxl
    wb = openpyxl.load_workbook(path)
    mp = {nm: f"S{i:02d}" for i, nm in enumerate(wb.sheetnames)}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    f = c.value
                    for o, nn in sorted(mp.items(), key=lambda x: -len(x[0])):
                        f = f.replace(f"'{o}'!", f"{nn}!").replace(f"{o}!", f"{nn}!")
                    c.value = _combin(f)
    for o, nn in mp.items(): wb[o].title = nn
    wb.save(path)
    al, _ = G["allocate"](t, full, b0, b1, b2, ca)
    return dict(b0=b0, b1=b1, b2=b2, gs=full["GS"], b3=b3, ca=ca, conv=conv,
                ctp1=t.k_w*ctp1, ctp2=t.k_w*ctp2,
                al=al, eq=(t.conv_class == "equity"), sep=(t.k_sep != 0),
                nosep=(t.conv_class == "equity" and t.k_sep != 0
                       and int(t.p_sep) == 0)), \
        mp["결과"], mp["회계처리"]


def solve(path, sheets):
    import formulas
    xl = formulas.ExcelModel().loads(path).finish()
    sol = xl.calculate()
    base = os.path.basename(path).upper()
    out = {nm: {} for nm in sheets}
    for k, v in sol.items():
        ku = k.upper()
        for nm in sheets:
            if ku.startswith(f"'[{base}]{nm}'!"):
                cell = ku.split("!")[-1]
                try: out[nm][cell] = float(v.value[0, 0])
                except Exception: pass
    return out


def main():
    G = load_app()
    # C10·C11 이 적용 모형의 트랜치다. 엔진의 b2·b3 는 이미 모형을 반영한 값이다.
    # 조서에는 앱에서 고른 방법만 들어간다. 그래서 적용값과 공통 항목만 본다.
    BASE = [("적용 70% 트랜치", "C10", "b2"), ("주계약", "C16", "b0"),
            ("부채요소", "C17", "b1"), ("조기상환청구권", "C18", None),
            ("매도청구권 적용값", "C22", "ca")]
    bad = 0
    for lbl, over in CASES:
        # 전환권대가는 자본으로 분류할 때만 나온다. 부채면 조서가 빈칸이 맞다.
        ROWS = BASE + ([("전환권대가", "C23", "conv")]
                       if over.get("conv_class", "equity") == "equity" else [])
        # 유무가치비교법일 때만 30% 트랜치가 조서에 있다
        if over.get("k_method", 0) == 0:
            ROWS = ROWS + [("적용 30% 트랜치", "C11", "b3")]
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "wb.xlsx")
            eng, res, acc = build(G, over, path)
            got = solve(path, (res, acc))
        print(f"\n[{lbl}]")
        for nm, cell, key in ROWS:
            want = eng[key] if key else eng["b1"] - eng["b0"]
            have = got[res].get(cell)
            ok = have is not None and abs(have - want) < 1e-4
            if not ok: bad += 1
            print("   %-22s 조서 %11s · 엔진 %11.4f  %s"
                  % (nm, f"{have:.4f}" if have is not None else "없음", want,
                     "" if ok else "★"))
        # 회계처리 — 배분표 C7:C11 과 분개 차·대변 합계
        A = got[acc]
        want_al = {k.split(" · ")[0].replace(" (잔여)", "").replace(
                       " (옵션 없는 사채)", "").replace(
                       " (사채 + 조기상환권)", ""): v for k, v in eng["al"][:-1]}
        have_al = {}
        for r in range(7, 13):
            v = A.get(f"C{r}")
            if v is not None: have_al[r] = v
        # 부호까지 포함해 합이 100 이 되어야 한다
        tot = A.get("C13")
        okt = tot is not None and abs(tot - 100.0) < 1e-4
        if not okt: bad += 1
        print("   %-22s 조서 %11s · 기준 %11.4f  %s"
              % ("배분 합계", f"{tot:.4f}" if tot is not None else "없음", 100.0,
                 "" if okt else "★"))
        drr, crr = A.get("C25"), A.get("D25")
        # 분개는 배분표를 뒤집은 것이다. 음수 항목만 차변으로 간다.
        want_dr = 100 + sum(-v for _, v in eng["al"][:-1] if v < 0)
        okj = (drr is not None and crr is not None
               and abs(drr - crr) < 1e-4 and abs(drr - want_dr) < 1e-4)
        if not okj: bad += 1
        print("   %-22s 차변 %11s · 대변 %11s · 기준 %10.4f  %s"
              % ("분개 대차", f"{drr:.4f}" if drr is not None else "없음",
                 f"{crr:.4f}" if crr is not None else "없음", want_dr,
                 "" if okj else "★"))
        # 배분 각 줄이 allocate() 와 같은가
        # 7 주계약 · 8 부채요소 · 9 조기상환권 · 10 복합내재파생 · 11 매도청구 · 12 전환권대가
        if eng["eq"] and eng["nosep"]:
            rows_ord = [("부채요소", 8), ("매도청구권", 11), ("전환권대가", 12)]
        elif eng["eq"]:
            rows_ord = [("주계약", 7), ("조기상환청구권", 9), ("매도청구권", 11),
                        ("전환권대가", 12)]
        else:
            rows_ord = [("주계약", 7), ("복합내재파생상품", 10), ("매도청구권", 11)]
        if not eng["sep"]:
            # 콜을 내재파생에 넣으면 자산 줄이 비고 파생 줄이 순액이 된다
            rows_ord = [(nm, r) for nm, r in rows_ord if nm != "매도청구권"]
            if eng["eq"]: rows_ord[0] = ("주계약", 7); rows_ord[1] = ("복합내재파생상품", 10)
        for nm, r in rows_ord:
            want = want_al.get(nm)
            if nm == "매도청구권": want = -eng["ca"]
            have = have_al.get(r)
            ok = have is not None and want is not None and abs(have - want) < 1e-4
            if not ok: bad += 1
            print("   %-22s 조서 %11s · 배분표 %9s  %s"
                  % ("배분 · " + nm, f"{have:.4f}" if have is not None else "없음",
                     f"{want:.4f}" if want is not None else "없음", "" if ok else "★"))
    print("\n" + ("모든 항목 일치" if bad == 0 else f"★ {bad}건 불일치"))
    return 0 if bad == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
