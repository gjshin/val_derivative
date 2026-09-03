#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""값 조서가 엔진과 같은 숫자를 싣는지 확인한다.

값 조서는 계산이 끝난 숫자를 받아 적는 스냅샷이다. 그래도 배분표·분개·결과 시트는
`build_xlsx` 안에서 다시 조립하므로 엔진과 어긋날 수 있다. 실제로 배분표와 분개가
같은 시트 안에서 13.5 만큼 어긋난 적이 있다.

수식 조서 쪽은 `조서대조.py` 가, 머리 행 배선은 `배선대조.py` 가 본다.
엑셀을 풀지 않으므로 몇 초면 끝난다.

    python3 tests/값조서대조.py
"""
import sys, os, io, types, warnings
warnings.filterwarnings("ignore")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def load_app():
    stub = types.ModuleType("streamlit"); stub.cache_data = lambda **k: (lambda f: f)
    sys.modules["streamlit"] = stub
    m = types.ModuleType("cbapp"); sys.modules["cbapp"] = m
    src = open(os.path.join(ROOT, "cb_app.py"), encoding="utf-8").read()
    exec(compile(src.split("st.set_page_config")[0], "cb_app.py", "exec"), m.__dict__)
    return m.__dict__


# 마지막 칸은 BDT 변동성이다. None 이면 조기상환권을 금리 고정 격자로 잰다.
CASES = [(cls, km, md, ks, None)
         for cls in ("equity", "liability")
         for km in (0, 1, 2)
         for md in ("TF", "GS")
         for ks in (1, 0)]
# 조기상환권을 BDT 로 잴 때. 자본·TF 에서만 열리므로 그 조합만 더한다.
# σ=0 은 확정 격자와 같은 값이 나와야 하고, σ>0 은 BDT 시트가 결과와 맞아야 한다.
CASES += [("equity", km, "TF", 1, sg) for km in (0, 1) for sg in (0.0, 0.20)]


def main():
    G = load_app()
    import openpyxl
    bad = []

    def chk(tag, a, b, tol=1e-6):
        ok = a is not None and abs(a - b) < tol
        print("   %-42s %12s %12.4f  %s"
              % (tag, f"{a:.4f}" if a is not None else "없음", b,
                 "OK" if ok else "★"))
        if not ok: bad.append(tag)

    for cls, km, md, ks, bsg in CASES:
        t = G["Terms"]()
        t.rf_curve = [(1, .0226), (3, .0240), (5, .0252)]
        t.cr_curve = [(1, .1409), (3, .1740), (5, .1905)]
        t.conv_class, t.k_method, t.model, t.k_sep = cls, km, md, ks
        t.put_bdt = 1 if bsg is not None else 0
        if bsg is not None: t.bdt_sig = bsg
        t.face_total = 9_000_000_000
        G["derive"](t)
        full, b0, b1, b2, ca, conv = G["decompose"](t)
        rows, _ = G["allocate"](t, full, b0, b1, b2, ca)
        # 앱과 같은 경로로 만든다. 상각표는 이론적 주계약이 아니라 인식액에서 출발한다.
        host = G["acc_host"](t, full, b0, b1, b2, ca)
        wb = openpyxl.load_workbook(io.BytesIO(
            G["build_xlsx"](t, full, b0, b1, b2, ca, conv, G["eir_table"](t, host))),
            data_only=True)
        E, R = wb["회계처리"], wb["결과"]
        print(f"\n[{cls} · 방법{km} · {md} · {'별도' if ks else '내재파생'}"
              + (f" · BDT σ{bsg:.0%}" if bsg is not None else "") + "]")
        # BDT 를 켰으면 그 시트가 결과와 맞아야 한다. 옵션 없는 사채는 ⑩ 주계약과
        # 같아야 하고 — 곡선을 정확히 되돌린다는 뜻이다.
        if bsg is not None:
            nn = int(t.n)
            chk("BDT 부채요소 = B1", wb["BDT 부채요소"].cell(13+nn+2, 3).value, b1, 1e-5)
            chk("BDT 주계약 = B0", wb["BDT 주계약"].cell(13+nn+2, 3).value, b0, 1e-5)
            if bsg == 0:
                g = G["pick"](G["engine"](t, conv=False, put=True, call=False), t.model)
                chk("σ=0 이면 확정 격자와 같다", b1, g, 1e-9)
        # 결과 시트 · 순차 차감
        for i, (nm, v) in enumerate([("B0 주계약", b0), ("B1 부채요소", b1),
                                     ("B2 전체", b2), ("B3 매도청구 반영", b2-ca)]):
            chk("결과 " + nm, R.cell(6+i, 3).value, v)
        # 배분표 각 줄과 합계
        for i, (k, v) in enumerate(rows[:-1]):
            chk("배분표 " + k.split(" · ")[0], E.cell(7+i, 3).value, v)
        chk("배분표 합계", E.cell(7+len(rows)-1, 3).value, 100.0, 0.01)
        # 전액 기준 환산
        chk("전액 기준 · 주계약", E.cell(7, 4).value,
            rows[0][1]/100*t.face_total, 1.0)
        # 분개 대차
        dr = cr = None
        for r in range(13, E.max_row+1):
            # 정확히 100 이면 openpyxl 이 int 로 돌려준다. float 만 보면 놓친다.
            if (E.cell(r, 2).value == "합계"
                    and isinstance(E.cell(r, 3).value, (int, float))):
                dr, cr = E.cell(r, 3).value, E.cell(r, 4).value
        # 분개는 배분표를 뒤집은 것이다. 음수 항목이 차변으로 간다.
        # 매도청구권을 내재파생에 넣으면 자산 줄이 사라지고, 파생 순액이 자산 쪽이면
        # 그 줄이 대신 차변으로 간다. 어느 쪽이든 대차는 맞아야 한다.
        want_dr = 100 + sum(-v for _, v in rows[:-1] if v < 0)
        chk("분개 차변", dr, want_dr, 0.01)
        chk("분개 대변", cr, want_dr, 0.01)
        # 상각표 마지막 줄 기말 = 만기상환금액
        M = wb["상각표"]
        r_eir, rows_eir, redm, nper = G["eir_table"](t, host)
        # 지급일 열이 생겨 기말은 H(8열)다.
        chk("상각표 기말 = 만기상환금액", M.cell(12+len(rows_eir), 8).value, redm, 0.01)
        chk("상각표 유효이자율", M.cell(9, 3).value, r_eir)
        chk("상각표 출발 = 배분된 주계약", M.cell(5, 3).value, rows[0][1], 0.01)

    print("\n" + ("모든 항목 일치" if not bad else "★ %d건 불일치" % len(bad)))
    for b in bad: print("   ★ " + b)
    return 0 if not bad else 1


if __name__ == "__main__":
    sys.exit(main())
