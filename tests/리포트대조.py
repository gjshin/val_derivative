#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""부속 리포트(변동성·이자율)가 엔진과 같은 답을 내는지 확인한다.

리포트는 계산을 전부 수식으로 담는다. 손으로 수식을 고치면 엔진과 어긋나기
쉬워서, 고칠 때마다 이 스크립트를 돌려야 한다.

    pip install formulas
    python3 tests/리포트대조.py

시트 이름에 한글이 있으면 formulas 가 깨져서 ASCII 로 바꾼 사본을 만들어 푼다.
"""
import sys, os, types, warnings, tempfile, math, random, datetime as dt
warnings.filterwarnings("ignore")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TOL = 1e-9
bad = 0


def load_app():
    stub = types.ModuleType("streamlit"); stub.cache_data = lambda **k: (lambda f: f)
    sys.modules["streamlit"] = stub
    m = types.ModuleType("cbapp"); sys.modules["cbapp"] = m
    src = open(os.path.join(ROOT, "cb_app.py"), encoding="utf-8").read()
    exec(compile(src.split("st.set_page_config")[0], "cb_app.py", "exec"), m.__dict__)
    return m.__dict__


def solve(data, want):
    """리포트를 실제로 계산한다. 시트 이름을 ASCII 로 바꿔 넣는다."""
    import openpyxl, formulas
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "r.xlsx")
        open(p, "wb").write(data)
        wb = openpyxl.load_workbook(p)
        mp = {nm: f"S{i:02d}" for i, nm in enumerate(wb.sheetnames)}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        f = c.value
                        for o, nn in sorted(mp.items(), key=lambda x: -len(x[0])):
                            f = f.replace(f"'{o}'!", f"{nn}!").replace(f"{o}!", f"{nn}!")
                        c.value = f
        for o, nn in mp.items(): wb[o].title = nn
        wb.save(p)
        sol = formulas.ExcelModel().loads(p).finish().calculate()
        base = os.path.basename(p).upper()
    out = {nm: {} for nm in want}
    for k, v in sol.items():
        ku = k.upper()
        for nm in want:
            if ku.startswith(f"'[{base}]{mp[nm]}'!"):
                try: out[nm][ku.split("!")[-1]] = float(v.value[0, 0])
                except Exception: pass
    return out, mp


def chk(lbl, got, exp, tol=TOL):
    global bad
    ok = got is not None and abs(got-exp) <= tol*max(1.0, abs(exp))
    bad += 0 if ok else 1
    print(f"   {lbl:26s} 리포트 {('—' if got is None else f'{got:12.8f}')} · "
          f"엔진 {exp:12.8f}  {'OK' if ok else '★'}")


def gen(n, s0, sig, seed, spike=None):
    random.seed(seed)
    d, px, p = dt.date(2024, 9, 2), [], s0
    for i in range(n):
        while d.weekday() >= 5: d += dt.timedelta(days=1)
        px.append((d.isoformat(), round(p, 0)))
        d += dt.timedelta(days=1)
        p *= math.exp(random.gauss(0, sig/math.sqrt(250))
                      + (0.25 if spike is not None and i == spike else 0))
    return px


def vol_cases(G):
    build, vol_from = G["build_xlsx_vol"], G["vol_from"]
    series = [("대상", gen(120, 16370, .67, 3, spike=60)),
              ("피어A", gen(120, 8200, .52, 5)),
              ("피어B", gen(120, 33000, .41, 9))]
    for lbl, kw in [("단일 · 이상치 제거", dict(one=True, drop=True)),
                    ("단일 · 제거 없음", dict(one=True, drop=False)),
                    ("피어 3개 · 중앙값", dict(pick="median")),
                    ("피어 3개 · 평균", dict(pick="mean")),
                    ("거래일수 252", dict(tdays=252))]:
        one = kw.pop("one", False)
        ser = series[:1] if one else series
        tdays = kw.get("tdays", 250); drop = kw.get("drop", True)
        pick = kw.get("pick", "median")
        data = build(ser, tdays=tdays, drop=drop, pick=pick)
        want = ([f"{i:02d} {nm}" for i, (nm, _) in enumerate(ser, 1)]
                + (["종합"] if len(ser) > 1 else []))
        got, _ = solve(data, want)
        print(f"\n[변동성 · {lbl}]")
        anns = []
        for i, (nm, px) in enumerate(ser, 1):
            v = vol_from(px, tdays, drop); anns.append(v["annual"])
            g = got[f"{i:02d} {nm}"]
            chk(f"{nm} 일 변동성", g.get("C15"), v["daily"])
            chk(f"{nm} 연 변동성", g.get("C16"), v["annual"])
            chk(f"{nm} 제외 개수", g.get("C14"), float(v["removed"]))
        if len(ser) > 1:
            a = sorted(anns)
            exp = ({"median": (a[len(a)//2] if len(a) % 2
                               else (a[len(a)//2-1]+a[len(a)//2])/2),
                    "mean": sum(a)/len(a)})[pick]
            chk("종합", got["종합"].get("C6"), exp)


def rate_cases(G):
    T, derive, curves, fwd = (G[k] for k in
                              ("Terms", "derive", "curves", "forward_rate"))
    build = G["build_xlsx_rate"]
    for lbl, kw in [("YTM 입력", {}), ("현물 입력", dict(y_type="spot")),
                    ("반기 복리", dict(cmp_rf=2, cmp_cr=2)),
                    ("등급 보간", dict(rate_mode="rating",
                                    cr_curve_b=[(0.25, .16), (1, .17), (3, .20), (5, .22)])),
                    ("월 노드", dict(gap_m=1.))]:
        t = T()
        t.rf_curve = [(0.25, .0250), (1, .0226), (3, .0240), (5, .0252)]
        t.cr_curve = [(0.25, .1300), (1, .1409), (3, .1740), (5, .1905)]
        t.gap_m = 6.
        for k, v in kw.items(): setattr(t, k, v)
        derive(t)
        got, _ = solve(build(t), ["선도이자율"])
        F = got["선도이자율"]
        RF, CR = curves(t)
        n, dt_ = int(t.n), t.T/int(t.n)
        u = math.exp(t.sig*math.sqrt(dt_)); d = 1/u
        miss = 0
        for i in range(n):
            rr = 10+i
            for col, exp in (("G", fwd(RF, i*dt_, (i+1)*dt_)),
                             ("J", fwd(CR, i*dt_, (i+1)*dt_)),
                             ("L", (math.exp(fwd(RF, i*dt_, (i+1)*dt_)*dt_)-d)/(u-d))):
                x = F.get(f"{col}{rr}")
                if x is None or abs(x-exp) > TOL*max(1.0, abs(exp)):
                    if miss < 3:
                        print(f"   ★ 스텝 {i} {col}: 리포트 {x} · 엔진 {exp}")
                    miss += 1
        global bad
        bad += miss
        print(f"\n[이자율 · {lbl}] n={n} · 어긋난 셀 {miss}건  "
              f"{'OK' if miss == 0 else '★'}")


def rate_series(n, r0, sig, seed):
    """로그정규 랜덤워크로 만든 금리 시계열."""
    random.seed(seed)
    d, out, r = dt.date(2024, 1, 2), [], r0
    for _ in range(n):
        while d.weekday() >= 5: d += dt.timedelta(days=1)
        out.append((d.isoformat(), round(r, 4)))
        d += dt.timedelta(days=1)
        r *= math.exp(random.gauss(0, sig/math.sqrt(250)))
    return out


def ratevol_cases(G):
    """금리변동성 리포트 — 상대·절대 둘 다 엔진과 맞는지."""
    build, rate_vol = G["build_xlsx_vol"], G["rate_vol"]
    for lbl, kw in [("금리 · 이상치 제거", dict(drop=True)),
                    ("금리 · 제거 없음", dict(drop=False)),
                    ("금리 · 거래일수 252", dict(tdays=252))]:
        tdays = kw.get("tdays", 250); drop = kw.get("drop", True)
        ser = [("A0·A- → BBB+ 고정만기", rate_series(300, 4.10, .20, 4))]
        data = build(ser, tdays=tdays, drop=drop, kind="rate",
                     how="검사용 합성 시계열")
        got, _ = solve(data, ["01 A0·A- → BBB+ 고정만기"])
        g = got["01 A0·A- → BBB+ 고정만기"]
        v = rate_vol(ser[0][1], tdays, drop)
        print(f"\n[금리변동성 · {lbl}]")
        chk("상대 일 변동성", g.get("C15"), v["daily"])
        chk("상대 연 변동성", g.get("C16"), v["annual"])
        chk("절대 일 변동성", g.get("C17"), v["abs_daily"])
        chk("절대 연 변동성", g.get("C18"), v["abs_annual"])
        chk("평균 금리", g.get("C19"), v["mean"])
        chk("제외 개수", g.get("C14"), float(v["removed"]))


def main():
    G = load_app()
    vol_cases(G)
    ratevol_cases(G)
    rate_cases(G)
    print("\n" + ("모든 항목 일치" if bad == 0 else f"★ {bad}건 불일치"))
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
