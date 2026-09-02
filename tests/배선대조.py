#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""수식 조서의 머리 행이 가정 시트의 **올바른 항목**을 참조하는지 본다.

값이 맞는지는 `조서대조.py` 와 `설정전수대조.py` 가 본다. 이쪽은 배선만 본다.
기본값이 우연히 같으면(매도청구 주기 3개월 = 조기상환 주기 3개월) 값 대조는
통과하면서도 엉뚱한 항목을 참조하고 있을 수 있다. 그런 것을 잡는다.

각 트리 시트 머리 행의 수식에서 `가정!$C$n` 참조를 뽑아 가정 시트의 항목
이름으로 되돌린 뒤, 그 행이 참조해야 할 항목 집합과 맞춰본다.

또 값 조서·수식 조서·엔진 셋이 같은 행사 시점을 쓰는지 경과기간을 바꿔가며 본다.
엑셀을 풀지 않으므로 몇 초면 끝난다.

    python3 tests/배선대조.py
"""
import sys, os, re, io, types, warnings, datetime as dt
warnings.filterwarnings("ignore")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 머리 행 -> 참조해야 하는 가정 항목 (키 이름)
WANT = {
    3:  {"cvs", "cve", "cv30"},                # Flag(전환) — ⑮ 는 cv30
    4:  {"pst", "pen", "frq"},                 # Flag(조기상환)
    5:  {"kst", "ken", "kfrq"},                # Flag(매도청구)
    6:  {"roff", "cyc"},                       # Flag(리픽싱)
    7:  {"pyld", "cpn", "pcmp", "dt", "elm", "prate"},   # 조기상환금액
    8:  {"prem", "cpn", "dt", "elm"},          # 매도청구금액
    9:  {"ipay", "cpn", "ipaym"},              # 쿠폰
    10: {"n", "red"},                          # 만기상환
}
ROWNAME = {3: "Flag(전환)", 4: "Flag(조기상환)", 5: "Flag(매도청구)",
           6: "Flag(리픽싱)", 7: "조기상환금액", 8: "매도청구금액",
           9: "쿠폰", 10: "만기상환"}


def load_app():
    stub = types.ModuleType("streamlit"); stub.cache_data = lambda **k: (lambda f: f)
    sys.modules["streamlit"] = stub
    m = types.ModuleType("cbapp"); sys.modules["cbapp"] = m
    src = open(os.path.join(ROOT, "cb_app.py"), encoding="utf-8").read()
    exec(compile(src.split("st.set_page_config")[0], "cb_app.py", "exec"), m.__dict__)
    return m.__dict__


def keymap(G):
    """가정 시트 행 번호 -> 키 이름.  cb_app 의 spec 순서를 그대로 읽는다."""
    src = open(os.path.join(ROOT, "cb_app.py"), encoding="utf-8").read()
    i = src.index('        ("평가기준일 주가", "S0"')
    j = src.index("    ROWN = {key:")
    keys = re.findall(r'\(\s*"[^"]*"\s*,\s*"(\w+)"', src[i:j])
    # spec 은 발행일부터 시작한다. 그 앞 세 줄을 찾아 붙인다.
    head = src[:i]
    pre = re.findall(r'\(\s*"[^"]*"\s*,\s*"(\w+)"', head[head.rindex("    spec = ["):])
    keys = pre + keys
    return {3+n: k for n, k in enumerate(keys)}


def terms(G, **kw):
    t = G["Terms"]()
    t.rf_curve = [(1, .0226), (3, .0240), (5, .0252)]
    t.cr_curve = [(1, .1409), (3, .1740), (5, .1905)]
    for k, v in kw.items(): setattr(t, k, v)
    G["derive"](t)
    return t


def main():
    G = load_app()
    import openpyxl
    bad = 0

    print("1. 머리 행이 가정 시트의 올바른 항목을 참조하는가")
    # 기본값이 서로 다르도록 일부러 어긋나게 준다
    t = terms(G, p_f=6., k_f=1., k_e=36., ipay=6., cpn=.03, ytm=.07, rfx_cyc=5.,
              gap_m=6.)
    full, b0, b1, b2, ca, conv = G["decompose"](t)
    data = G["build_xlsx_formula"](t, full, b0, b1, b2, ca, conv, G["eir_table"](t, b0))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    km = keymap(G)
    # ⑯ 부채요소는 머리 행 구성이 다르므로 뺀다. ⑮ 30% 트랜치의 전환 Flag 는
    # 의무보유를 반영한 cv30 을 쓰는 것이 맞다.
    trees = [s for s in wb.sheetnames
             if re.match(r"^\d\d ", s) and not s.startswith("16 ")]
    for row, want in WANT.items():
        seen = set()
        for nm in trees:
            ws = wb[nm]
            for c in range(3, 3+t.n+1):
                v = ws.cell(row, c).value
                if isinstance(v, str) and v.startswith("="):
                    for mm in re.finditer(r"가정!\$C\$(\d+)", v):
                        seen.add(km.get(int(mm.group(1)), "?%s" % mm.group(1)))
        extra, miss = seen - want, want - seen
        ok = not extra and not miss
        if not ok: bad += 1
        print("   %-14s %-34s %s" % (ROWNAME[row], " · ".join(sorted(seen)),
                                     "OK" if ok else
                                     "★ 잘못 참조 %s / 빠짐 %s" % (sorted(extra), sorted(miss))))

    print("\n2. 엔진 · 값 조서 · 수식 조서가 같은 시점을 쓰는가")
    d0 = dt.date.fromisoformat(G["Terms"]().d_issue)
    for el in (0, 3, 6, 12, 18, 24, 36):
        t = terms(G, d_base=(d0+dt.timedelta(days=int(el*30.4375))).isoformat())
        n = t.n; mper = n/(t.T*12)
        stp = lambda mth: int(round((mth-t.elapsed_m)*mper))
        per = max(1, int(round(t.rfx_cyc*mper)))
        off = int(round((t.rfx_cyc - t.elapsed_m % t.rfx_cyc)*mper)) if t.rfx_cyc > 0 else 1
        eng = {
            "리픽싱": {i for i in range(1, n+1) if i >= off and (i-off) % per == 0},
            "조기상환": {i for i in range(0, n+1)
                     if max(stp(t.p_s), 0) <= i <= stp(t.p_e)
                     and (i-stp(t.p_s)) % max(1, int(round(t.p_f*mper))) == 0},
            "매도청구": {i for i in range(0, n+1)
                     if max(stp(t.k_s), 0) <= i <= stp(t.k_e)
                     and (i-stp(t.k_s)) % max(1, int(round(t.k_f*mper))) == 0},
        }
        full, b0, b1, b2, ca, conv = G["decompose"](t)
        wv = openpyxl.load_workbook(io.BytesIO(
            G["build_xlsx"](t, full, b0, b1, b2, ca, conv, G["eir_table"](t, b0))))["01 주가"]
        val = {"리픽싱": {i for i in range(1, n+1) if wv.cell(6, 3+i).value == 1},
               "조기상환": {i for i in range(0, n+1) if wv.cell(4, 3+i).value == 1},
               "매도청구": {i for i in range(0, n+1) if wv.cell(5, 3+i).value == 1}}
        # 수식 조서는 가정 값으로 같은 식을 다시 세워 확인한다
        wf = openpyxl.load_workbook(io.BytesIO(
            G["build_xlsx_formula"](t, full, b0, b1, b2, ca, conv, G["eir_table"](t, b0))))["가정"]
        A = {}
        for r in range(3, 90):
            nm2 = wf.cell(r, 2).value
            if nm2: A[nm2] = wf.cell(r, 3).value
        fm = {
            "리픽싱": {i for i in range(1, n+1)
                    if i >= A["첫 조정 스텝"] and (i-A["첫 조정 스텝"]) % A["리픽싱 주기 (스텝)"] == 0},
            "조기상환": {i for i in range(0, n+1)
                     if A["조기상환 시작 (스텝)"] <= i <= A["조기상환 종료 (스텝)"]
                     and (i-A["조기상환 시작 (스텝)"]) % A["조기상환 주기 (스텝)"] == 0},
            "매도청구": {i for i in range(0, n+1)
                     if A["매도청구 시작 (스텝)"] <= i <= A["매도청구 종료 (스텝)"]
                     and (i-A["매도청구 시작 (스텝)"]) % A["매도청구 주기 (스텝)"] == 0},
        }
        marks = []
        for k in ("리픽싱", "조기상환", "매도청구"):
            ok = eng[k] == val[k] == fm[k]
            if not ok: bad += 1
            marks.append("%s %s" % (k, "OK" if ok else "★"))
        print("   경과 %2.0f개월  n=%2d  %s" % (t.elapsed_m, n, " · ".join(marks)))

    print("\n3. 값 조서의 금액 행이 엔진 산식과 같은가")
    for lbl, kw in (("무이표", {}),
                    ("표면 2% · 보장 7% 복리", dict(cpn=.02, ytm=.07,
                                                 p_mode="accrue", p_yield=.07)),
                    ("중간평가 12개월", dict(d_base="2026-03-31", cpn=.02,
                                        p_mode="accrue", p_yield=.07)),
                    ("중간평가 18개월 · 반기", dict(d_base="2026-09-30", cpn=.03,
                                             ipay=6., ytm=.07, p_mode="accrue",
                                             p_yield=.07, k_prem=.05))):
        t = terms(G, **kw)
        n = t.n; mper = n/(t.T*12); dt_ = t.T/n; ey = t.elapsed_m/12
        stp = lambda mth: int(round((mth-t.elapsed_m)*mper))
        per = lambda mth: max(1, int(round(mth*mper)))
        ar = G["accrue_rate"]
        inset = lambda i, a, b, fr: stp(a) <= i <= stp(b) and (i-stp(a)) % per(fr) == 0
        full, b0, b1, b2, ca, conv = G["decompose"](t)
        W = openpyxl.load_workbook(io.BytesIO(
            G["build_xlsx"](t, full, b0, b1, b2, ca, conv,
                            G["eir_table"](t, b0))))["01 주가"]
        worst = [("", 0.0)]
        for i in range(n+1):
            # 조기상환금액
            want = (0.0 if not inset(i, t.p_s, t.p_e, t.p_f) else
                    (100*(1+ar(i*dt_+ey, t.p_yield, t.cpn, t.p_cmp))
                     if t.p_mode == "accrue" else t.p_rate))
            got = W.cell(7, 3+i).value or 0.0
            if abs(got-want) > worst[0][1]: worst = [("조기상환금액", abs(got-want))]
            # 매도청구금액
            want = (999999 if not inset(i, t.k_s, t.k_e, t.k_f) else
                    100*(1+ar(i*dt_+ey, t.k_prem, t.cpn, 1)))
            got = W.cell(8, 3+i).value or 0.0
            if abs(got-want) > worst[0][1]: worst = [("매도청구금액", abs(got-want))]
            # 쿠폰
            want = 100*t.cpn*t.ipay/12 if (t.cpn > 0 and i > 0
                                           and i % per(t.ipay) == 0) else 0.0
            got = W.cell(9, 3+i).value or 0.0
            if abs(got-want) > worst[0][1]: worst = [("쿠폰", abs(got-want))]
            # 만기상환
            want = 100*(1+ar(t.T+ey, t.ytm, t.cpn, t.ytm_cmp)) if i == n else 0.0
            got = W.cell(10, 3+i).value or 0.0
            if abs(got-want) > worst[0][1]: worst = [("만기상환", abs(got-want))]
        nm2, gap = worst[0]
        ok = gap < 1e-3
        if not ok: bad += 1
        print("   %-22s n=%2d  최대 차이 %s %.6f  %s"
              % (lbl, n, nm2 or "—", gap, "OK" if ok else "★"))

    print("\n4. 상태확장을 골랐을 때 수식 조서가 그 사실을 밝히는가")
    # 상태확장 격자는 재결합하지 않아 엑셀 트리로 못 옮긴다. 조서는 경로가중치로
    # 다시 계산하므로 앱 값과 다르다. 그 차이를 조서가 스스로 적어야 한다.
    for carry, lbl in ((0, "상태확장"), (1, "경로가중치"), (2, "확률가중"), (3, "특정노드")):
        t = terms(G, carry=carry, rfx_mode=2, rfx_cyc=7., gap_m=3.)
        full, b0, b1, b2, ca, conv = G["decompose"](t)
        R = openpyxl.load_workbook(io.BytesIO(
            G["build_xlsx_formula"](t, full, b0, b1, b2, ca, conv,
                                    G["eir_table"](t, b0))))["결과"]
        stamped, warned = R["C32"].value, bool(R["B33"].value)
        if carry == 0:
            ok = (stamped is not None and abs(stamped - b2) < 1e-9 and warned)
            note = "앱 값 %.4f 기재 · 경고 %s" % (stamped or -1, "있음" if warned else "없음")
        else:
            ok = (stamped in (None, "") and not warned)
            note = "해당 없음 (조서와 같은 방법)"
        if not ok: bad += 1
        print("   %-8s %-40s %s" % (lbl, note, "OK" if ok else "★"))

    print("\n" + ("배선 이상 없음" if bad == 0 else "★ %d건" % bad))
    return 0 if bad == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
