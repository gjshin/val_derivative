#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""전환사채 평가 — Streamlit

실행
    pip install streamlit numpy pandas matplotlib openpyxl yfinance
    streamlit run cb_app.py

금액은 전자등록금액 100 기준이다.
"""
from __future__ import annotations
import math, json, io, re, calendar, datetime as dt
from dataclasses import dataclass, asdict, field

import numpy as np
import pandas as pd
import streamlit as st

# ══════════════════════════════════════════════════════════
# 1. 인풋
# ══════════════════════════════════════════════════════════
@dataclass
class Terms:
    S0: float = 853.0
    K0: float = 853.0
    d_issue: str = "2025-03-31"     # 발행일
    d_base: str = "2025-03-31"      # 평가기준일
    d_mat: str = "2030-03-31"       # 만기일
    gap_m: float = 1.0              # 노드 간격 (개월)
    T: float = 5.0                  # 잔존기간 — derive() 가 채운다
    n: int = 60                     # 노드 수 — derive() 가 채운다
    elapsed_m: float = 0.0          # 발행일 → 평가기준일 경과 개월
    cpn: float = 0.0            # 표면이자율
    ipay: float = 3.0           # 이자 지급주기 (개월)
    ytm: float = 0.0            # 만기보장수익률
    ytm_cmp: int = 4            # 만기보장수익률 복리 횟수 (분기)
    cv_s: float = 12.0          # 전환 시작 (개월)
    cv_e: float = 59.0
    rfx_mode: int = 2           # 0 없음 / 1 하향만 / 2 하향+상향
    rfx_cyc: float = 7.0
    floor: float = 598.0
    par: float = 500.0
    carry: int = 0              # 0 상태확장 / 1 경로가중치 / 2 확률가중평균 / 3 특정노드선택
    p_s: float = 24.0
    p_e: float = 57.0
    p_f: float = 3.0
    p_rate: float = 100.0
    k_s: float = 12.0
    k_e: float = 24.0
    k_f: float = 3.0
    k_prem: float = 0.01
    k_cmp: int = 4
    k_w: float = 0.30
    k_lock: float = 25.0
    k_third: int = 1              # 매도청구권을 제3자에게 지정할 수 있는가
    k_transfer: int = 0           # 매도청구권을 사채와 독립적으로 양도할 수 있는가
    p_lost_int: int = 0           # 조기상환 행사금액이 상실이자 보상 수준인가
    fvpl_whole: int = 0           # 복합계약 전체를 당기손익-공정가치로 지정했는가
    k_method: int = 0
    k_sep: int = 1                  # 1 별도 금융상품 / 0 복합내재파생에 포함            # 0 유무가치비교 / 1 혼합할인율 / 2 지분·부채 분리
    sig: float = 0.4130
    model: str = "TF"
    conv_class: str = "equity"   # equity 전환권 자본 / liability 전환권 파생상품부채
    cmp_rf: int = 2              # 무위험 복리 횟수 (국고채 반기)
    cmp_cr: int = 4              # 위험 복리 횟수 (회사채 분기)
    p_mode: str = "fixed"        # fixed 고정률 / accrue 보장수익률 복리
    p_yield: float = 0.0         # 조기상환 보장수익률
    p_cmp: int = 4               # 조기상환 보장수익률 복리 횟수
    face_total: float = 25_000_000_000.0   # 전자등록총액 (원)
    rate_mode: str = "direct"      # direct 직접 · pick 등급 하나 · rating 두 등급 보간
    cr_src: str = ""               # 위험 곡선을 어디서 가져왔는지 (조서에 적는다)     # direct 곡선 직접 / rating 등급 보간
    rt_a: str = "BBB+"            # 인풋 곡선 A 등급
    rt_b: str = "BBB-"            # 인풋 곡선 B 등급
    rt_tgt: str = "BBB0"          # 평가대상 등급
    cr_curve_b: list = field(default_factory=list)
    put_bdt: int = 0              # 조기상환권 0 격자(확정) / 1 BDT 금리격자
    bdt_sig: float = 0.20         # BDT 단기이자율 변동성 (로그정규, 연)
    bdt_base: int = 0             # 0 위험 곡선 직접 / 1 무위험 + 확정 스프레드
    y_type: str = "par"           # par 만기수익률 / spot 현물이자율
    rf_curve: list = field(default_factory=list)   # [(만기, 연이율)]
    cr_curve: list = field(default_factory=list)


def accrue_rate(t_year: float, g: float, c: float, m: int) -> float:
    """상환할증금률. 미지급 보장수익률을 매 회차 적립해 굴린 연금의 미래가치다.

        Σ_{k=1..mt} ((g−c)/m)·(1+g/m)^(mt−k)  =  (g−c)/g · ((1+g/m)^(mt) − 1)

    표면이자율 c 를 빼는 것은 그만큼 이미 현금으로 지급했기 때문이다.
    c 가 0 이면 (1+g/m)^(mt) − 1 로 줄어 종전 산식과 같아진다.

    보장수익률이 표면이자율보다 낮으면 산식이 음수가 된다. 그러나 할증금은
    수익률을 채워 주려고 **더** 얹는 돈이라 음수가 될 수 없다 — 이미 지급한
    이자를 만기에 되돌려 받는 계약은 없다. 그래서 0 에서 끊는다. 그런
    입력은 애초에 잘못이므로 validate() 가 따로 경고한다.
    """
    if t_year <= 0: return 0.0
    m = max(1, int(m))
    if g <= 1e-12: return max(0.0, (g - c)*t_year)   # g → 0 극한
    return max(0.0, (g - c)/g * ((1 + g/m)**(m*t_year) - 1))


def step_mapper(tm: "Terms", n: int, dt_: float):
    """계약상 월(발행일 기준)을 노드 번호로 바꾸는 두 함수를 만든다.

    스텝을 반올림으로 잡으면 계약일 **전**의 노드에서 행사가 열려 옵션이
    과대평가된다. 그래서 노드의 실제 날짜를 계약일과 직접 견준다.

        lo(m)  계약일 **이후** 첫 노드   — 행사기간 시작에 쓴다
        hi(m)  계약일 **이전** 마지막 노드 — 행사기간 종료에 쓴다

    노드가 하나도 조건을 만족하지 않으면 lo 는 n+1, hi 는 −1 을 돌려주어
    그 구간이 비어 있음을 알린다.
    """
    di = dt.date.fromisoformat(tm.d_issue)
    db = dt.date.fromisoformat(tm.d_base)
    day = dt_*365                                   # 한 스텝의 일수
    nd = [db + dt.timedelta(days=round(i*day)) for i in range(n+1)]
    # 노드는 한 달을 30.4일로 잡아 놓으므로 달력 기준일과 하루이틀 어긋난다.
    # 그만큼은 같은 날로 본다. 노드 하나를 통째로 앞당길 만큼은 못 된다.
    tol = dt.timedelta(days=min(5, max(1, int(day//4))))

    def cd(m):                                      # 발행일 + m 개월
        k = int(math.floor(m)); fr = m - k
        d = _add_months(di, k)
        return d + dt.timedelta(days=round(fr*30.4375)) if fr else d

    def lo(m):
        c = cd(m) - tol
        return next((i for i, x in enumerate(nd) if x >= c), n+1)

    def hi(m):
        c = cd(m) + tol
        return next((i for i in range(n, -1, -1) if nd[i] <= c), -1)
    return lo, hi


def _add_months(d: dt.date, k: int) -> dt.date:
    """d 에서 k 개월 뒤 같은 날. 그 달에 그 날이 없으면 말일로 맞춘다."""
    y, mo = d.year + (d.month - 1 + k)//12, (d.month - 1 + k) % 12 + 1
    return dt.date(y, mo, min(d.day, calendar.monthrange(y, mo)[1]))


def months_between(d1: dt.date, d2: dt.date) -> float:
    """개월 단위 경과기간.

    꽉 찬 개월을 세고, 남는 일수는 **그 구간의 실제 한 달 길이**로 나눈다.
    말일까지 남은 날수로 나누면 안 된다 — 그러면 12월 23일에서 31일까지 8일이
    0.89개월로 부풀어 결산일 평가가 통째로 밀린다.
    """
    if d2 <= d1: return 0.0
    m = (d2.year - d1.year)*12 + (d2.month - d1.month)
    if d2.day < d1.day: m -= 1
    same, nxt = _add_months(d1, m), _add_months(d1, m + 1)
    return m + (d2 - same).days/max(1, (nxt - same).days)


def derive(tm: Terms) -> Terms:
    """날짜에서 경과기간·잔존기간·노드 수를 계산해 채운다."""
    di = dt.date.fromisoformat(tm.d_issue)
    db = dt.date.fromisoformat(tm.d_base)
    dm = dt.date.fromisoformat(tm.d_mat)
    tm.elapsed_m = max(0.0, months_between(di, db))
    tm.T = max(1e-6, (dm-db).days/365)
    gap = max(0.25, tm.gap_m)
    tm.n = max(4, int(round(tm.T*12/gap)))
    return tm


# ══════════════════════════════════════════════════════════
# 2. 이자율 곡선
# ══════════════════════════════════════════════════════════
def _lin(pts, t):
    if not pts: return None
    if t <= pts[0][0]: return pts[0][1]
    if t >= pts[-1][0]: return pts[-1][1]
    for i in range(1, len(pts)):
        if t <= pts[i][0]:
            (x0, y0), (x1, y1) = pts[i-1], pts[i]
            return y0 + (y1-y0)*(t-x0)/(x1-x0)


def bootstrap_df(par_pts, Tmax, m=1):
    """만기수익률 곡선 → 할인계수.

    par_pts 는 [(만기, 연 만기수익률)] 이고 m 은 연간 이표 횟수다.
    선형보간으로 이표 시점마다 수익률을 만든 뒤 앞에서부터 순차로 푼다.
        1 = c·(DF1 + … + DFk) + DFk        c = 해당 만기 수익률 ÷ m
    """
    N = max(1, int(math.ceil(Tmax*m)))
    out, acc = [(0.0, 1.0)], 0.0
    for k in range(1, N+1):
        t = k/m
        c = _lin(par_pts, t)/m
        df = (1 - c*acc)/(1 + c)
        acc += df
        out.append((t, df))
    return out


def make_curve(par_pts, Tmax, m=1):
    """할인계수에서 연속복리 현물이자율 함수를 만든다."""
    dfs = bootstrap_df(par_pts, Tmax, m)
    spot = [(t, -math.log(df)/t) for t, df in dfs if t > 0]
    if not spot: return lambda t: 0.0
    spot = [(1e-6, spot[0][1])] + spot
    return lambda t: spot[0][1] if t <= 0 else _lin(spot, t)


def spot_from_zero(pts, m=1):
    """이미 현물이자율(이산)로 받은 경우 — 연속복리로 바꾼다.

    m 은 그 현물이자율의 복리 횟수다. 국고채는 연복리로 고시되는 경우가 많지만
    회사채 제로커브는 분기복리인 경우가 있다. 복리 횟수를 무시하고 ln(1+r) 로만
    환산하면 할인계수와 위험중립확률이 어긋난다 (책 3.7.4.4 오류 사례).

        연속 = m · ln(1 + r/m)        m=1 이면 ln(1+r) 로 되돌아온다
    """
    m = max(1, int(m))
    cont = [(t, m*math.log(1 + r/m)) for t, r in pts]
    return lambda t: cont[0][1] if t <= 0 else _lin(cont, t)


def forward_rate(F, t0, t1):
    """구간 선도이자율.  f = [r(t1)·t1 − r(t0)·t0] ÷ (t1 − t0)"""
    return (F(t1)*t1 - F(t0)*t0)/(t1-t0)


RATINGS = ["AAA", "AA+", "AA0", "AA-", "A+", "A0", "A-",
           "BBB+", "BBB0", "BBB-", "BB+", "BB0", "BB-",
           "B+", "B0", "B-", "CCC+", "CCC0", "CCC-", "CC", "C", "D"]


def rating_idx(r: str) -> int:
    """등급을 노치 번호로. AAA=0 에서 한 단계씩 내려간다."""
    r = (r or "").strip().upper().replace(" ", "")
    if r in RATINGS: return RATINGS.index(r)
    for alt in (r+"0", r.replace("0", "")):
        if alt in RATINGS: return RATINGS.index(alt)
    return -1


# 긴 등급부터 맞춰야 'BBB-' 를 'BB' 로 잘못 집지 않는다. 앞뒤로 다른 알파벳이
# 붙으면 등급이 아니다 — 그래야 'CD(91일)' 을 등급 C 로 집지 않는다.
_RATING_PAT = re.compile(
    r"(?<![A-Z])(?:" + "|".join(sorted((re.escape(r) for r in RATINGS),
                                       key=len, reverse=True)) + r")(?![A-Z])")


def rating_in(text) -> str:
    """문자열 안에서 신용등급을 찾는다. 못 찾으면 None.

    고시표의 줄 이름은 '회사채 I(공모사채) / 무보증 / BBB0' 처럼 등급이 뒤에
    붙는다. 그 줄이 어느 등급인지 알아야 화면에서 **표에 실제로 있는 등급만**
    고르게 할 수 있다.
    """
    t = (text or "").upper().replace(" ", "")
    m = _RATING_PAT.search(t)
    if m: return m.group(0)
    # 'AA' · 'BBB' 처럼 0 을 안 붙인 표기. 뒤에 숫자가 붙으면 'CP(A2)' 처럼
    # 다른 척도의 기호이므로 등급으로 보지 않는다.
    m = re.search(r"(?<![A-Z])(AAA|AA|A|BBB|BB|B|CCC|CC|C|D)(?![A-Z0-9+\-])", t)
    if m:
        c = m.group(1)
        return c if c in RATINGS else (c+"0" if c+"0" in RATINGS else None)
    return None


def blend_curves(pts_a, pts_b, ra, rb, rt):
    """두 등급 곡선을 노치 거리로 선형보간해 대상 등급 곡선을 만든다.

    대상 등급이 두 등급 사이면 내삽, 밖이면 같은 기울기로 외삽한다.
    """
    ia, ib, it = rating_idx(ra), rating_idx(rb), rating_idx(rt)
    if not pts_a: return pts_b or []
    if not pts_b or ia < 0 or ib < 0 or it < 0 or ia == ib: return pts_a
    w = (it-ia)/(ib-ia)
    ts = sorted({t for t, _ in pts_a} | {t for t, _ in pts_b})
    out = []
    for t in ts:
        ya, yb = _lin(pts_a, t), _lin(pts_b, t)
        if ya is None or yb is None: continue
        out.append((t, ya + (yb-ya)*w))
    return out


def to_cont(r, m):
    """이산복리(연 m회) → 연속복리.  국고채 반기(2), 회사채 분기(4)가 관행이다."""
    return m*math.log(1 + r/m) if (m and m > 0) else r


def credit_curve(tm: Terms):
    """위험 곡선. 등급 보간 방식이면 두 등급 곡선을 섞는다."""
    if tm.rate_mode == "rating" and len(tm.cr_curve) >= 2 and len(tm.cr_curve_b) >= 2:
        return blend_curves(tm.cr_curve, tm.cr_curve_b, tm.rt_a, tm.rt_b, tm.rt_tgt)
    return tm.cr_curve


def curves(tm: Terms):
    """무위험·위험 모두 만기수익률 곡선을 부트스트래핑해 쓴다."""
    cc = credit_curve(tm)
    if len(tm.rf_curve) >= 2 and len(cc) >= 2:
        if tm.y_type == "spot":
            # 현물이자율은 고시된 복리 횟수로 연속환산한다 (책 3.7.4.4)
            return (spot_from_zero(tm.rf_curve, tm.cmp_rf),
                    spot_from_zero(cc, tm.cmp_cr))
        return (make_curve(tm.rf_curve, tm.T, tm.cmp_rf),
                make_curve(cc, tm.T, tm.cmp_cr))
    # 곡선이 아직 없을 때의 임시값 — 화면이 경고를 띄운다
    return (lambda t: to_cont(0.028, tm.cmp_rf)), (lambda t: to_cont(0.07, tm.cmp_cr))


# ══════════════════════════════════════════════════════════
# 3. 격자 엔진
# ══════════════════════════════════════════════════════════
TOL = 1e-9      # 동점 판정 허용오차. 값이 100 근처라 1e-9 은 잡음보다 크고 실질 차이보다 작다


def engine(tm: Terms, conv=True, put=True, call=False, conv_start=None):
    RF, CR = curves(tm)
    n, T = int(tm.n), tm.T
    dt_ = T/n
    mper = n/(T*12)
    el = tm.elapsed_m                       # 발행일 → 평가기준일 경과 개월
    # 계약상 개월 → 노드 번호. 시작은 계약일 이후 첫 노드, 종료는 이전 마지막 노드.
    st_lo, st_hi = step_mapper(tm, n, dt_)
    ey = el/12                               # 경과 연수
    u = math.exp(tm.sig*math.sqrt(dt_)); d = 1/u
    fwd = lambda F, i: (F((i+1)*dt_)*(i+1)*dt_ - F(i*dt_)*i*dt_)/dt_
    qi = lambda i: (math.exp(fwd(RF, i)*dt_) - d)/(u - d)
    cs = tm.cv_s if conv_start is None else conv_start

    def in_set(i, a, b, fr):
        lo, hi = st_lo(a), st_hi(b)
        if i < max(lo, 0) or i > hi: return False
        per = max(1, int(round(fr*mper)))
        return (i-lo) % per == 0
    rfx_per = max(1, int(round(tm.rfx_cyc*mper)))
    # 다음 조정일은 발행일 + (지난 회차 + 1) × 주기 다. 그 날 이후 첫 노드가 시작이다.
    rfx_off = (st_lo(tm.rfx_cyc*(math.floor(el/tm.rfx_cyc) + 1))
               if tm.rfx_cyc > 0 else 1)
    is_rfx = lambda i: (tm.rfx_mode > 0 and i > 0 and i >= rfx_off
                        and (i-rfx_off) % rfx_per == 0)
    pay_per = max(1, int(round(tm.ipay*mper)))
    is_pay = lambda i: tm.cpn > 0 and i > 0 and i % pay_per == 0
    cpn_amt = 100*tm.cpn*tm.ipay/12
    red = 100*(1 + accrue_rate(T + ey, tm.ytm, tm.cpn, tm.ytm_cmp))
    S = lambda i, j: tm.S0 * u**j * d**(i-j)
    clip = lambda s: min(max(s, tm.floor, tm.par), tm.K0)
    def put_amt(i):
        """행사금액은 발행일부터 붙는다. 경과분을 더해 계산한다."""
        if tm.p_mode == "accrue":
            return 100*(1 + accrue_rate(i*dt_ + ey, tm.p_yield, tm.cpn, tm.p_cmp))
        return tm.p_rate
    put_a = lambda i: put_amt(i) if (put and in_set(i, tm.p_s, tm.p_e, tm.p_f)) else 0.0
    # kstrike 는 콜 스위치와 무관한 행사금액이다. 행사기간이 아니면 None.
    # call_a 는 call=False 면 항상 inf 라 제3자 콜옵션 평가에 쓸 수 없다.
    kstrike = lambda i: (100*(1 + accrue_rate(i*dt_ + ey, tm.k_prem, tm.cpn,
                                             tm.k_cmp))
                         if in_set(i, tm.k_s, tm.k_e, tm.k_f) else None)
    call_a = lambda i: (kstrike(i) if (call and in_set(i, tm.k_s, tm.k_e, tm.k_f))
                        else math.inf)
    conv_ok = lambda i: conv and st_lo(cs) <= i <= st_hi(tm.cv_e)
    exact = (tm.rfx_mode > 0 and tm.carry == 0)

    Kg = None
    if not exact:
        Kg = [[tm.K0]]
        for i in range(1, n+1):
            q = qi(i-1)
            P = lambda a, b: math.comb(a, b) * q**b * (1-q)**(a-b)
            row = []
            for j in range(i+1):
                if tm.rfx_mode == 0:
                    row.append(tm.K0); continue
                if is_rfx(i):
                    base = S(i, j) if tm.rfx_mode == 2 else min(Kg[i-1][min(j, i-1)], S(i, j))
                    row.append(clip(base)); continue
                up = Kg[i-1][j-1] if j-1 >= 0 else None
                dn = Kg[i-1][j] if j <= i-1 else None
                if up is None: row.append(dn); continue
                if dn is None: row.append(up); continue
                if tm.carry == 3: row.append(dn)
                elif tm.carry == 2: row.append(up*q + dn*(1-q))
                else:
                    wu, wd = P(i-1, j-1)*q, P(i-1, j)*(1-q)
                    row.append((up*wu + dn*wd)/(wu+wd))
            Kg.append(row)

    memo = {}
    def rec(i, j, K):
        key = (i, j, round(K, 6)) if exact else (i, j)
        if key in memo: return memo[key]
        if i == n:
            KK = K if exact else Kg[n][j]
            cv = 100*S(n, j)/KK if conv_ok(n) else 0.0
            pv = put_a(n)
            # 만기에도 이자 지급일이면 이자를 함께 받는다 (책 5-7 만기 현금흐름).
            # 전환을 택하면 주식을 받으므로 중간 노드와 같이 이자는 사라진다.
            cm = cpn_amt if is_pay(n) else 0.0
            cash = max(pv, red) + cm
            # 리픽싱이 주가로 재설정되는 날에는 전환가치가 정확히 100 이 되어
            # 상환금액과 동점이 된다. 부동소수 잡음으로 갈리지 않게 전환은
            # TOL 만큼 앞설 때만 이긴다. 동점이면 현금(부채)이다.
            if cv >= max(cash, 0.0) + TOL and cv > 0:
                o = dict(E=cv, B=0.0, V=cv, P=1.0, kind="conv", hold=red, cv=cv, K=KK)
            else:
                o = dict(E=0.0, B=cash, V=cash, P=0.0, kind="mat",
                         hold=red, cv=cv, K=KK)
        else:
            def nk(s):
                if tm.rfx_mode == 0: return tm.K0
                if not is_rfx(i+1): return K
                return clip(s) if tm.rfx_mode == 2 else clip(min(K, s))
            KU = nk(S(i, j)*u) if exact else Kg[i+1][j+1]
            KD = nk(S(i, j)*d) if exact else Kg[i+1][j]
            ku = (i+1, j+1, round(KU, 6)) if exact else (i+1, j+1)
            kd = (i+1, j,   round(KD, 6)) if exact else (i+1, j)
            a, b = rec(i+1, j+1, KU), rec(i+1, j, KD)
            q = qi(i); c = cpn_amt if is_pay(i) else 0.0
            fr, fc = fwd(RF, i), fwd(CR, i)
            E = (q*a["E"] + (1-q)*b["E"]) * math.exp(-fr*dt_)
            B = (q*a["B"] + (1-q)*b["B"]) * math.exp(-fc*dt_) + c
            # GS — 자식 노드의 전환확률로 각각 할인한다 (순환참조가 생기지 않는다)
            ya = a["P"]*fr + (1-a["P"])*fc
            yb = b["P"]*fr + (1-b["P"])*fc
            Vc = q*a["V"]*math.exp(-ya*dt_) + (1-q)*b["V"]*math.exp(-yb*dt_) + c
            pr = q*a["P"] + (1-q)*b["P"]
            KK = K if exact else Kg[i][j]
            cv = 100*S(i, j)/KK if conv_ok(i) else 0.0
            pv, kv = put_a(i), call_a(i)
            hold = E + B; inner = min(hold, kv)
            Vg = max(cv, pv, min(Vc, kv))
            # GS 의 전환확률은 GS 자신의 판단을 따른다. TF 와 다른 갈래를 고를 수 있다.
            # GS 도 같은 순서다. 현금이 동점이면 전환확률 0 이다.
            if abs(Vg - pv) < TOL or (kv < math.inf and abs(Vg - kv) < TOL): Pg = 0.0
            elif cv > 0 and abs(Vg - cv) < TOL:                              Pg = 1.0
            else:                                                            Pg = pr
            # up·dn 은 자식 노드 키다. 만기 노드에는 없어 자식 없음의 표시가 된다.
            ex = dict(hold=hold, cv=cv, K=KK, pv=pv, kv=kv, Vc=Vc, up=ku, dn=kd)
            # 동점 처리는 위 만기 노드와 같다. 전환은 TOL 만큼 앞설 때만 이긴다.
            if i == 0:                       o = dict(E=E, B=B, V=Vg, P=Pg, kind="hold", **ex)
            elif cv >= max(pv, inner) + TOL: o = dict(E=cv, B=0.0, V=Vg, P=Pg, kind="conv", **ex)
            elif pv >= inner - TOL:          o = dict(E=0.0, B=pv, V=Vg, P=Pg, kind="put", **ex)
            elif hold <= kv + TOL:           o = dict(E=E, B=B, V=Vg, P=Pg, kind="hold", **ex)
            else:                            o = dict(E=0.0, B=kv, V=Vg, P=Pg, kind="call", **ex)
        memo[key] = o
        return o

    r0 = rec(0, 0, tm.K0)

    # 정산 유형 분포
    dist = dict(conv=0.0, put=0.0, call=0.0, mat=0.0, tc=0.0, tp=0.0, tk=0.0)
    layer = {((0, 0, round(tm.K0, 6)) if exact else (0, 0)): (1.0, tm.K0, 0)}
    for i in range(n):
        nxt, q = {}, qi(i)
        for key, (p_, K, j) in layer.items():
            o = memo.get(key)
            if o is None: continue
            if i > 0 and o["kind"] != "hold":
                if o["kind"] == "conv": dist["conv"] += p_; dist["tc"] += p_*i
                elif o["kind"] == "put": dist["put"] += p_; dist["tp"] += p_*i
                elif o["kind"] == "call": dist["call"] += p_; dist["tk"] += p_*i
                continue
            def nk2(s):
                if tm.rfx_mode == 0: return tm.K0
                if not is_rfx(i+1): return K
                return clip(s) if tm.rfx_mode == 2 else clip(min(K, s))
            KU = nk2(S(i, j)*u) if exact else Kg[i+1][j+1]
            KD = nk2(S(i, j)*d) if exact else Kg[i+1][j]
            for kk, pp, jj, KK in (
                ((i+1, j+1, round(KU, 6)) if exact else (i+1, j+1), p_*q, j+1, KU),
                ((i+1, j, round(KD, 6)) if exact else (i+1, j), p_*(1-q), j, KD)):
                if kk in nxt:
                    a0, b0, c0 = nxt[kk]; nxt[kk] = (a0+pp, b0, c0)
                else:
                    nxt[kk] = (pp, KK, jj)
        layer = nxt
    dist["mat"] = sum(v[0] for v in layer.values())

    root = (0, 0, round(tm.K0, 6)) if exact else (0, 0)
    return dict(TF=r0["E"]+r0["B"], E=r0["E"], B=r0["B"], GS=r0["V"], P=r0["P"],
                q=qi(0), u=u, d=d, dt=dt_, mper=mper, n=n, memo=memo, Kg=Kg,
                exact=exact, S=S, host=100*math.exp(-CR(T)*T), dist=dist,
                root=root, qi=qi, fwdRF=lambda i: fwd(RF, i),
                fwdCR=lambda i: fwd(CR, i), kstrike=kstrike)


def pick(res, model): return res["GS"] if model == "GS" else res["TF"]

def call_third_party(tm: Terms, full, method: int) -> float:
    """제3자 지정 가능 콜옵션 — 옵션차익혼합할인법.

    한국공인회계사회 『K-IFRS 실무사례와 해설 11 복합금융상품』 4.4.3 과
    부속예제 [사례 4-4] 의 산식이다. 발행자가 지정한 제3자에게 넘어갈 수 있는
    콜옵션은 기준서 1109 문단 4.3.1 상 별도의 금융상품이고, 기초자산이
    전환사채인 복합옵션(an option on an option)이므로 격자 안에서
    MIN(계속보유, 콜금액) 으로 누르는 발행자 콜옵션과 다르게 평가한다.

    기초자산은 ``콜과 그 부속조항(의무보유 등)을 포함하지 않은 전환사채`` 다.
    ``decompose`` 가 넘기는 ``full = engine(tm, call=False)`` 이 정확히 그것이라
    새 격자를 만들지 않고 그 memo 를 한 번 더 역진한다.

    method 1  혼합할인율      — 값 하나를 자식의 구성비율로 섞은 할인율로 할인
    method 2  지분·부채 분리   — 페이오프를 구성비율로 쪼개 각각 Rf·Rd 로 할인
    """
    memo, dt_ = full["memo"], full["dt"]
    qi, fRF, fCR = full["qi"], full["fwdRF"], full["fwdCR"]
    kstrike = full["kstrike"]
    cache = {}

    def w(o):
        """구성비율 — 노드 가치 중 지분 몫."""
        v = o["E"] + o["B"]
        return o["E"]/v if v > 1e-12 else 0.0

    def rec(key, i):
        if key in cache: return cache[key]
        o = memo[key]
        K = kstrike(i)
        pay = max(o["E"] + o["B"] - K, 0.0) if K is not None else 0.0
        if "up" not in o:                       # 만기 — 자식이 없다
            ww = w(o)
            r = (pay, pay*ww, pay*(1-ww))
        else:
            q, ou, od = qi(i), memo[o["up"]], memo[o["dn"]]
            cu, eu, bu = rec(o["up"], i+1)
            cd, ed, bd = rec(o["dn"], i+1)
            if method == 1:
                yu = w(ou)*fRF(i) + (1-w(ou))*fCR(i)
                yd = w(od)*fRF(i) + (1-w(od))*fCR(i)
                cont = q*cu*math.exp(-yu*dt_) + (1-q)*cd*math.exp(-yd*dt_)
                r = (max(pay, cont), 0.0, 0.0)
            else:
                he = (q*eu + (1-q)*ed) * math.exp(-fRF(i)*dt_)
                hb = (q*bu + (1-q)*bd) * math.exp(-fCR(i)*dt_)
                if pay >= he + hb:
                    ww = w(o); r = (pay, pay*ww, pay*(1-ww))
                else:
                    r = (he + hb, he, hb)
        cache[key] = r
        return r

    return rec(full["root"], 0)[0]


K_METHODS = {0: "유무가치비교법", 1: "옵션차익혼합할인법 · 혼합할인율",
             2: "옵션차익혼합할인법 · 지분·부채 분리"}


# ══════════════════════════════════════════════════════════
# 3-1. BDT 금리격자 — 조기상환권 전용
# ══════════════════════════════════════════════════════════
# 전환을 끄면 격자가 주가와 무관해져 스텝마다 값이 하나뿐이다. 즉 지금
# 조기상환권은 불확실성이 없는 확정 계산이고 옵션의 시간가치가 없다.
# 금리를 확률변수로 두면 그 시간가치가 생긴다. 조기상환권은 주가와 무관한
# 순수 금리·신용 상품이라, 주가 격자를 건드리지 않고 여기서만 따로 잰다.


def bdt_tree(spot, T: float, n: int, sig: float):
    """현물이자율 곡선에 맞춘 BDT 단기이자율 격자.

        r(i, j) = a_i · exp(2·σ·j·√Δt)          j 는 상승 횟수 (0..i)

    로그정규라 이자율이 음수가 되지 않는다. a_i 는 (i+1)Δt 만기 무이표채를
    정확히 재현하도록 역산한다 — 그래서 옵션이 없는 사채는 곡선을 그대로
    되돌려 준다. 위험중립확률은 BDT 관행대로 0.5 다.

    σ 가 0 이면 a_i 가 구간 선도이자율이 되어 확정 격자와 완전히 같아진다.
    이 성질을 검사에서 쓴다.
    """
    dt_ = T/max(1, n)
    sq = math.sqrt(dt_)
    P = [math.exp(-spot(k*dt_)*k*dt_) for k in range(n+1)]   # 시장 할인계수
    r, base, Q = [], [], [[1.0]]                             # Q 는 도달가격
    for i in range(n):
        mul = [math.exp(2*sig*j*sq) for j in range(i+1)]
        def price(a):
            return sum(Q[i][j]*math.exp(-a*mul[j]*dt_) for j in range(i+1))
        lo, hi = 1e-10, 5.0                    # 연 500% 까지 잡으면 넉넉하다
        for _ in range(200):                   # price 는 a 에 대해 감소한다
            mid = (lo+hi)/2
            if price(mid) > P[i+1]: lo = mid
            else: hi = mid
        a = (lo+hi)/2
        base.append(a)
        r.append([a*x for x in mul])
        nq = [0.0]*(i+2)
        for j in range(i+1):
            d = 0.5*Q[i][j]*math.exp(-r[i][j]*dt_)
            nq[j+1] += d; nq[j] += d
        Q.append(nq)
    return r, base


def bdt_parts(tm: Terms):
    """BDT 격자에서 쓸 재료를 한곳에서 만든다. 조서도 이것을 그대로 쓴다.

    기준 곡선은 두 가지로 고를 수 있다.

    * 0 위험 곡선 직접 — 단기이자율이 곧 위험이자율이다. σ 가 위험이자율
      전체의 변동성이라 신용스프레드 변동성까지 안고 간다. 옵션 없는 사채가
      격자의 주계약과 정확히 같아져 검산이 쉽다.
    * 1 무위험 + 확정 스프레드 — 국고채에 σ 를 태우고 구간 선도 스프레드를
      확정으로 얹는다. σ 를 국고채에서 관측한 값으로 쓸 수 있지만,
      스프레드가 금리와 무관하다고 본 것이므로 그 한계를 조서에 적어야 한다.
    """
    derive(tm)
    RF, CR = curves(tm)
    n, T = int(tm.n), tm.T
    dt_ = T/n
    ey = tm.elapsed_m/12
    mper = n/(T*12)
    st_lo, st_hi = step_mapper(tm, n, dt_)
    if tm.bdt_base == 0:
        rt, ab = bdt_tree(CR, T, n, tm.bdt_sig)
        add = [0.0]*n
    else:
        rt, ab = bdt_tree(RF, T, n, tm.bdt_sig)
        add = [forward_rate(CR, i*dt_, (i+1)*dt_) - forward_rate(RF, i*dt_, (i+1)*dt_)
               for i in range(n)]
    red = 100*(1 + accrue_rate(T + ey, tm.ytm, tm.cpn, tm.ytm_cmp))
    cpn_amt = 100*tm.cpn*tm.ipay/12
    pay_per = max(1, int(round(tm.ipay*mper)))
    is_pay = lambda i: tm.cpn > 0 and i > 0 and i % pay_per == 0
    p_lo, p_hi = st_lo(tm.p_s), st_hi(tm.p_e)
    p_per = max(1, int(round(tm.p_f*mper)))
    in_put = lambda i: (max(p_lo, 0) <= i <= p_hi and (i-p_lo) % p_per == 0)
    put_a = lambda i: ((100*(1 + accrue_rate(i*dt_ + ey, tm.p_yield, tm.cpn, tm.p_cmp))
                        if tm.p_mode == "accrue" else tm.p_rate)
                       if in_put(i) else 0.0)
    # 캘리브레이션 검산 재료 — 도달가격 Q 와 시장 할인계수.
    # Σ_j Q(k,j) 가 시장 할인계수와 같아야 한다. 이것이 무차익거래 조건이고,
    # 기준금리 a 를 그 조건에 맞춰 역산한 것이다. 조서에서 눈으로 확인하도록
    # 격자와 함께 내보낸다.
    base = CR if tm.bdt_base == 0 else RF
    mkt = [math.exp(-base(k*dt_)*k*dt_) for k in range(n+1)]
    Q = [[1.0]]
    for i in range(n):
        nq = [0.0]*(i+2)
        for j in range(i+1):
            d = 0.5*Q[i][j]*math.exp(-rt[i][j]*dt_)
            nq[j+1] += d; nq[j] += d
        Q.append(nq)
    return dict(r=rt, a=ab, add=add, n=n, T=T, dt=dt_, red=red, cpn=cpn_amt,
                is_pay=is_pay, in_put=in_put, put_a=put_a, Q=Q, mkt=mkt,
                base_nm=("위험 곡선" if tm.bdt_base == 0 else "무위험 곡선"))


def bdt_grid(tm: Terms, put: bool):
    """BDT 격자에서 사채를 역진하고 전 노드 값을 돌려준다.

    만기 노드와 중간 노드의 판정을 격자 엔진과 같은 순서로 맞춘다 —
    만기는 MAX(조기상환금액, 만기상환금액) + 쿠폰, 중간은 MAX(계속보유,
    조기상환금액) 이다. 조서가 표를 그릴 때 이 격자를 그대로 쓴다.
    """
    B = bdt_parts(tm)
    n, dt_ = B["n"], B["dt"]
    cm = B["cpn"] if B["is_pay"](n) else 0.0
    V = [[0.0]*(i+1) for i in range(n+1)]
    for j in range(n+1):
        V[n][j] = max(B["put_a"](n) if put else 0.0, B["red"]) + cm
    for i in range(n-1, -1, -1):
        c = B["cpn"] if B["is_pay"](i) else 0.0
        for j in range(i+1):
            d = math.exp(-(B["r"][i][j] + B["add"][i])*dt_)
            h = (0.5*V[i+1][j+1] + 0.5*V[i+1][j])*d + c
            if put and B["in_put"](i): h = max(h, B["put_a"](i))
            V[i][j] = h
    return B, V


def bond_bdt(tm: Terms, put: bool) -> float:
    """BDT 격자에서 잰 사채 가치 (t=0)."""
    return bdt_grid(tm, put)[1][0][0]


def put_bdt_on(tm: Terms) -> bool:
    """BDT 를 실제로 쓸 조건인가.

    전환권을 자본으로 두고 TF 를 쓸 때만 연다. 자본이면 전환권대가가 잔여라
    부채요소만 바꿔도 배분이 그대로 성립하지만, 부채로 두면 복합내재파생을
    전체로서 재야 해서 전체 가치(주가 격자)까지 같이 손봐야 하기 때문이다.
    """
    return bool(tm.put_bdt) and tm.conv_class == "equity" and tm.model == "TF" \
        and tm.p_s <= tm.p_e


def decompose(tm: Terms):
    derive(tm)
    full = engine(tm, call=False)
    b0 = pick(engine(tm, conv=False, put=False, call=False), tm.model)
    b1 = pick(engine(tm, conv=False, put=True, call=False), tm.model)
    # 조기상환권을 BDT 로 재면 부채요소만 갈아 끼운다. 전환권이 자본이면
    # 전환권대가가 잔여라 배분이 그대로 성립하고 합계도 100 을 지킨다.
    # 전체 가치(b2)는 주가 격자 그대로 두므로, 그 차이는 잔여가 흡수한다.
    if put_bdt_on(tm):
        b1 = bond_bdt(tm, True)
    b2 = pick(full, tm.model)
    # 행사 가능한 시점이 하나도 없으면 매도청구권은 없다.
    ks = full["kstrike"]
    has_call = tm.k_w > 0 and any(ks(i) is not None for i in range(full["n"]+1))
    if not has_call:
        ca = 0.0
    elif tm.k_method:
        # 제3자 지정 가능 콜옵션 — 기초자산은 콜·의무보유를 뺀 full 그대로다
        ca = tm.k_w*call_third_party(tm, full, tm.k_method)
    else:
        # 의무보유는 전환을 늦추기만 한다. 전환 시작보다 이르면 아무 제약이 아니다.
        cs = max(tm.cv_s, tm.k_lock)
        b3 = pick(engine(tm, conv=True, put=True, call=True, conv_start=cs), tm.model)
        ca = tm.k_w*(b2-b3)
    resid = 100 - b1 + ca          # 전환권이 자본일 때의 잔여 (전환권대가)
    return full, b0, b1, b2, ca, resid

# 분리 판단에서 "행사가격이 상각후원가와 거의 같다" 로 볼 문턱.
# 기준서는 "거의 같다" 라고만 하고 수치를 주지 않는다. 실무에서 널리 쓰는
# 10% 를 기본으로 두되, 문턱 근처면 그 사실 자체를 알린다.
SPLIT_TOL = 0.10


def _close_test(strike, amort):
    """행사가격과 상각후원가가 '거의 같은가' — 문단 B4.3.5(5)(가)."""
    gap = abs(strike - amort)/max(abs(amort), 1e-9)
    return gap, gap <= SPLIT_TOL


def split_test(tm: Terms, full, b0, b1, b2, ca, rows_eir):
    """조기상환권·매도청구권을 주계약과 분리해야 하는지 계약 조항으로 판단한다.

    판단 순서가 정해져 있다. 문단 B4.3.5 말미가 못박는다 — "기업회계기준서
    제1032호에 따라 전환채무상품의 자본요소를 분리하기 **전에** 내재된
    콜옵션이나 풋옵션이 주채무계약과 밀접하게 관련되어 있는지를 판단한다."

    돌려주는 것은 옵션마다 결론·근거·평가방법·지표를 담은 사전이다. 화면과
    조서가 같은 것을 쓰므로 둘이 어긋날 수 없다.
    """
    n, dt_ = int(tm.n), tm.T/int(tm.n)
    ey = tm.elapsed_m/12
    liab = tm.conv_class != "equity"

    def amort_at(t_year):
        """그 시점 주계약의 상각후원가 — 실제로 인식한 배분액에서 상각한 값."""
        return next((en for _, tt_, _b, _i, _c, en in rows_eir
                     if tt_ >= t_year - 1e-9), b0)

    out = {}

    # ── 조기상환청구권 ────────────────────────────────────────
    if tm.p_s > tm.p_e or tm.T <= 0:
        out["put"] = dict(있음=False, 결론="해당 없음",
                          이유=["계약에 조기상환청구권이 없습니다."],
                          근거=[], 평가="—", 지표={})
    else:
        pv = (100*(1 + accrue_rate(tm.p_s/12, tm.p_yield, tm.cpn, tm.p_cmp))
              if tm.p_mode == "accrue" else tm.p_rate)
        bv = amort_at(max(0.0, (tm.p_s - tm.elapsed_m)/12))
        gap, close = _close_test(pv, bv)
        why, cite = [], []
        if tm.fvpl_whole:
            res = "분리하지 않음"
            why.append("복합계약 전체를 당기손익-공정가치로 측정하므로 분리 요건이 "
                       "성립하지 않습니다.")
            cite.append("1109 문단 4.3.3(3)")
        elif liab:
            res = "묶어서 분리"
            why.append("전환권이 파생상품부채이므로 조기상환권을 따로 떼지 않고 "
                       "전환권과 하나의 복합내재파생상품으로 묶어 전체로서 "
                       "측정합니다. 전환하거나 상환받거나 둘 중 하나라 서로 "
                       "배타적이어서, 따로 재어 더하면 총액이 부풀려집니다.")
            cite.append("1109 문단 B4.3.4")
        elif tm.p_lost_int:
            res = "분리하지 않음"
            why.append("행사가격이 잔여기간 상실이자의 현재가치를 보상하는 "
                       "수준이므로 주계약과 밀접하게 관련되어 있습니다.")
            cite.append("1109 문단 B4.3.5(5)(나)")
        elif close:
            res = "분리하지 않을 여지"
            why.append(f"첫 조기상환일 행사금액 {pv:,.2f} 와 같은 시점 주계약 "
                       f"상각후원가 {bv:,.2f} 의 차이가 {gap*100:.1f}% 로 "
                       "거의 같습니다. 밀접하게 관련되어 있다고 볼 여지가 "
                       "있습니다.")
            cite.append("1109 문단 B4.3.5(5)(가)")
        else:
            res = "분리"
            why.append(f"첫 조기상환일 행사금액 {pv:,.2f} 와 같은 시점 주계약 "
                       f"상각후원가 {bv:,.2f} 의 차이가 {gap*100:.1f}% 로 "
                       "거의 같지 않습니다.")
            why.append("같은 조건의 별도 금융상품이 파생상품의 정의를 충족하고, "
                       "복합계약 전체를 당기손익-공정가치로 측정하지 않습니다.")
            cite += ["1109 문단 B4.3.5(5)", "문단 4.3.3"]
        val = ("순차 차감 — 조기상환권만 얹은 값에서 옵션 없는 사채를 뺍니다 "
               f"(B1 − B0 = {b1-b0:,.4f}). 전환권과 대체 관계라 따로 재어 "
               "더하면 총액이 부풀려집니다."
               if res == "분리" else
               "묶음 전체를 공정가치로 측정합니다 (B2 − B0)." if res == "묶어서 분리"
               else "분리하지 않으므로 주계약에 포함해 상각후원가로 측정합니다.")
        out["put"] = dict(있음=True, 결론=res, 이유=why, 근거=cite, 평가=val,
                          지표={"첫 조기상환일 행사금액": pv,
                                "같은 시점 상각후원가": bv,
                                "차이": gap})

    # ── 매도청구권 ────────────────────────────────────────────
    ks = full["kstrike"]
    first_k = next((i for i in range(n+1) if ks(i) is not None), None)
    if tm.k_w <= 0 or first_k is None:
        out["call"] = dict(있음=False, 결론="해당 없음",
                           이유=["계약에 매도청구권이 없거나 행사 가능한 시점이 "
                                 "없습니다."], 근거=[], 평가="—", 지표={})
    else:
        kv = ks(first_k)
        kb = amort_at(first_k*dt_)
        kgap, kclose = _close_test(kv, kb)
        why, cite = [], []
        if tm.k_third or tm.k_transfer:
            res = "별도의 금융상품"
            why.append("계약상 " + ("발행회사가 제3자를 지정할 수 있어 거래상대방이 "
                                  "달라질 수 있습니다. " if tm.k_third else "")
                       + ("사채와 독립적으로 양도할 수 있습니다. "
                          if tm.k_transfer else "")
                       + "내재파생상품이 아니라 별도의 금융상품이므로 분리 요건을 "
                         "따질 것 없이 처음부터 별개의 파생상품으로 인식합니다.")
            cite.append("1109 문단 4.3.1 마지막 문장")
        elif tm.fvpl_whole:
            res = "분리하지 않음"
            why.append("복합계약 전체를 당기손익-공정가치로 측정하므로 분리 요건이 "
                       "성립하지 않습니다.")
            cite.append("1109 문단 4.3.3(3)")
        elif liab:
            res = "묶어서 분리"
            why.append("발행회사만 행사할 수 있어 내재파생상품이고, 전환권이 "
                       "파생상품부채이므로 전환권·조기상환권과 하나의 "
                       "복합내재파생상품으로 묶어 전체로서 측정합니다.")
            cite.append("1109 문단 B4.3.4")
        elif kclose:
            res = "분리하지 않을 여지"
            why.append(f"첫 매도청구일 매매대금 {kv:,.2f} 와 같은 시점 주계약 "
                       f"상각후원가 {kb:,.2f} 의 차이가 {kgap*100:.1f}% 로 "
                       "거의 같습니다.")
            cite.append("1109 문단 B4.3.5(5)(가)")
        else:
            res = "분리"
            why.append(f"발행회사만 행사할 수 있어 내재파생상품이고, 첫 매도청구일 "
                       f"매매대금 {kv:,.2f} 와 같은 시점 주계약 상각후원가 "
                       f"{kb:,.2f} 의 차이가 {kgap*100:.1f}% 로 거의 같지 "
                       "않습니다.")
            cite += ["1109 문단 4.3.1", "문단 B4.3.5(5)"]
        if res == "별도의 금융상품":
            val = ("기초자산이 전환사채 전체인 미국형 복합옵션이므로 "
                   "**옵션차익혼합할인법**이 개념적으로 정합합니다. 다만 계약에 "
                   "의무보유 조건이 있으면 그 효과를 값에 넣으려고 "
                   "유무가치비교법을 쓰기도 합니다 — 재는 대상이 다르므로 "
                   "고른 방법을 조서에 밝히십시오."
                   if tm.k_lock > tm.cv_s else
                   "기초자산이 전환사채 전체인 미국형 복합옵션이므로 "
                   "**옵션차익혼합할인법**이 개념적으로 정합합니다.")
        elif res in ("분리", "묶어서 분리"):
            val = ("순차 차감 — 콜을 넣고 뺀 차액입니다 "
                   f"(적용값 {ca:,.4f}). 의무보유로 잃는 전환권 가치까지 값에 "
                   "들어가므로, 계약에 의무보유가 없으면 그만큼 과대해집니다."
                   if tm.k_lock > tm.cv_s else
                   f"순차 차감 — 콜을 넣고 뺀 차액입니다 (적용값 {ca:,.4f}).")
        else:
            val = "분리하지 않으므로 주계약에 포함해 상각후원가로 측정합니다."
        out["call"] = dict(있음=True, 결론=res, 이유=why, 근거=cite, 평가=val,
                           지표={"첫 매도청구일 매매대금": kv,
                                 "같은 시점 상각후원가": kb,
                                 "차이": kgap,
                                 "제3자 지정 가능": bool(tm.k_third),
                                 "독립 양도 가능": bool(tm.k_transfer)})

    # 화면에서 고른 회계 처리와 판정이 어긋나면 알린다
    want = 1 if out["call"]["결론"] == "별도의 금융상품" else 0
    out["call"]["설정일치"] = (not out["call"]["있음"]) or (tm.k_sep == want)
    return out


def split_memo(sp) -> str:
    """분리 판단을 조서에 옮길 글로 편다. 화면과 조서가 같은 문안을 쓴다."""
    out = []
    for key, nm in (("put", "조기상환청구권"), ("call", "매도청구권")):
        d = sp[key]
        if not d["있음"]:
            out.append(f"[{nm}] {d['이유'][0]}"); continue
        out.append(f"[{nm}] 결론 — {d['결론']}\n"
                   + "\n".join("  · " + x for x in d["이유"])
                   + (f"\n  근거 — {' · '.join(d['근거'])}" if d["근거"] else "")
                   + "\n  평가방법 — " + d["평가"].replace("**", ""))
    return "\n\n".join(out)


def allocate(tm: Terms, full, b0, b1, b2, ca):
    """최초 인식 배분.  전환권 분류에 따라 무엇을 잔여로 두는지가 뒤바뀐다.

    매도청구권을 어떻게 볼지는 ``k_sep`` 이 정한다.

    * 1 별도 금융상품 — 제3자 지정이 가능하면 거래상대방이 달라지므로 내재파생이
      아니라 별도의 금융상품이다 (기준서 1109 문단 4.3.1 마지막 문장).
      파생상품자산으로 따로 세운다.
    * 0 복합내재파생에 포함 — 발행회사만 행사할 수 있으면 거래상대방이 그대로라
      내재파생상품이다. 전환권·조기상환권과 하나의 복합내재파생상품으로 묶는다
      (문단 B4.3.4). 콜은 발행자에게 유리하므로 묶음을 그만큼 줄인다.

    어느 쪽이든 주계약과 전환권대가는 같다. 파생을 총액으로 볼지 순액으로 볼지가
    다를 뿐이다.
    """
    sep = tm.k_sep != 0
    if tm.conv_class == "liability":
        # 전환권이 파생상품부채 — 내재파생을 공정가치로 두고 주계약을 잔여로.
        # 전환권과 조기상환권은 상호의존적이라 하나의 복합내재파생상품으로 묶어
        # 전체로서(as a whole) 측정한다 (문단 B4.3.4).
        deriv = (b2 - b0) if sep else (b2 - ca - b0)
        host_acc = (100 + ca) - (b2 - b0)      # 어느 쪽이든 같다
        rows = [("주계약 (잔여)", host_acc),
                ("복합내재파생상품 · 파생상품부채", deriv)]
        if sep: rows.append(("매도청구권 · 파생상품자산", -ca))
        note = ("전환권이 파생상품부채이므로 전환권과 조기상환권을 하나의 "
                "복합내재파생상품으로 묶어 공정가치로 측정하고 주계약을 잔여로 둡니다 "
                "(기준서 1109 문단 B4.3.4). "
                + ("매도청구권은 제3자 지정이 가능해 별도의 금융상품이므로 "
                   "이 묶음에 넣지 않습니다 (문단 4.3.1). 전환사채에 배분된 금액은 "
                   f"{100+ca:,.2f} 입니다."
                   if sep else
                   "매도청구권은 발행회사만 행사할 수 있어 거래상대방이 그대로이므로 "
                   "내재파생상품이고, 같은 묶음에 넣어 순액으로 측정합니다 (문단 4.3.1).")
                + f" 이론적 주계약가치는 {b0:,.2f} 입니다.")
    else:
        rows = [("주계약 (옵션 없는 사채)", b0),
                ("조기상환청구권 · 파생상품부채", (b1-b0) if sep else (b1-b0-ca))]
        if sep: rows.append(("매도청구권 · 파생상품자산", -ca))
        if not sep:
            rows[1] = ("복합내재파생상품 · 파생상품부채", b1-b0-ca)
        rows.append(("전환권대가 · 자본", 100-b1+ca))
        note = ("기업회계기준서 제1032호 문단 31 — 부채요소를 먼저 정하고 나머지를 자본에 배분합니다. "
                "최초 인식에는 손익이 생기지 않습니다."
                + ("" if sep else
                   " 매도청구권은 발행회사만 행사할 수 있어 내재파생상품이므로 "
                   "조기상환권과 하나로 묶어 순액으로 봅니다 (문단 4.3.1 · B4.3.4)."))
    rows.append(("합계", sum(v for _, v in rows)))
    if tm.elapsed_m > 0.01:
        note += ("  ※ 이 배분은 **최초 인식**용입니다. 평가기준일이 발행일보다 뒤이므로 "
                 "결산 회계처리에는 그대로 쓰지 마십시오. 결산일에 필요한 것은 파생상품의 "
                 "공정가치뿐이고, 주계약은 발행일 배분액을 유효이자율로 상각한 장부금액입니다.")
    return rows, note


def acc_host(tm: Terms, full, b0, b1, b2, ca):
    """상각표가 출발해야 하는 금액 — 실제로 인식한 주계약이다.

    자본 분류면 이론적 주계약(b0)이 그대로 인식되지만, 부채 분류면 잔여로
    떨어진 금액이 인식된다. 상각후원가는 최초 인식액에서 출발해야 하므로
    이론값이 아니라 배분액으로 유효이자율을 역산한다.
    """
    return allocate(tm, full, b0, b1, b2, ca)[0][0][1]


def allocate_full(tm: Terms, rows):
    """100 기준 배분에 전액 기준 금액을 붙인다."""
    f = tm.face_total
    return [(k, v, v/100*f) for k, v in rows]


def pay_index(tm: Terms, t_year: float) -> int:
    """상각 회차의 연수를 계약상 지급 회차 번호로 되짚는다.

    t 는 평가기준일 기준이므로 경과분을 더해 발행일 기준으로 옮긴 뒤
    지급주기로 나눈다. 지급일이 발행일 + m×주기 이므로 m 이 나온다.
    """
    per = max(1e-6, tm.ipay/12)
    return max(1, int(round((t_year + tm.elapsed_m/12)/per)))


def eir_table(tm: Terms, host):
    c = 100*tm.cpn*tm.ipay/12
    per = max(1e-6, tm.ipay/12)
    red = 100*(1 + accrue_rate(tm.T + tm.elapsed_m/12, tm.ytm, tm.cpn, tm.ytm_cmp))
    # 지급일은 계약상 일정이므로 **발행일**부터 센다. 평가기준일이 발행일보다
    # 뒤이면 첫 회차만 짧아지고 나머지는 온전한 한 주기다. 평가기준일에서
    # 세면 지급일이 계약과 어긋나 이자비용이 회차마다 밀린다.
    # 마지막은 만기다. 남는 조각이 주기의 10% 미만이면 앞 회차에 붙여
    # 하루짜리 회차를 만들지 않는다.
    ey_ = tm.elapsed_m/12
    ts, k = [], 1
    while k*per - ey_ < tm.T - per*0.1:
        t_ = k*per - ey_
        if t_ > per*0.1: ts.append(t_)
        k += 1
    ts.append(tm.T)
    nper = len(ts)
    def pv(r):
        return (sum(c*(1+r)**(-t) for t in ts[:-1])
                + (c + red)*(1+r)**(-tm.T))
    lo, hi = -0.5, 5.0
    for _ in range(200):
        m = (lo+hi)/2
        if pv(m) > host: lo = m
        else: hi = m
    r = (lo+hi)/2
    rows, bv, prev = [], host, 0.0
    for k, t in enumerate(ts, 1):
        it = bv*((1+r)**(t-prev) - 1); end = bv + it - c
        rows.append((k, t, bv, it, c, end)); bv, prev = end, t
    return r, rows, red, nper


def validate(tm: Terms):
    derive(tm)
    w = []
    if tm.cv_s >= tm.cv_e: w.append("전환 시작이 종료보다 늦거나 같습니다.")
    horizon = tm.T*12 + tm.elapsed_m + 0.5      # 발행일 기준 총 개월
    if tm.cv_e > horizon: w.append(f"전환 종료({tm.cv_e:.0f}개월)가 만기({horizon:.0f}개월)를 넘습니다.")
    if tm.p_s > tm.p_e: w.append("조기상환 시작이 종료보다 늦습니다.")
    if tm.k_s > tm.k_e: w.append("매도청구 시작이 종료보다 늦습니다.")
    if tm.k_lock < tm.k_e: w.append("의무보유 전환지연이 매도청구 종료보다 이릅니다. 콜이 실효화될 수 있습니다.")
    # 할증금 산식은 보장수익률에서 표면이자율을 뺀다. 보장이 더 낮으면 음수가
    # 되어 상환금액이 액면 밑으로 내려간다. 0 에서 끊고는 있지만 입력 자체가
    # 계약과 맞지 않으므로 알려 준다.
    if tm.ytm > 0 and tm.ytm < tm.cpn - 1e-9:
        w.append(f"만기보장수익률({tm.ytm:.2%})이 표면이자율({tm.cpn:.2%})보다 낮습니다. "
                 "상환할증금이 음수가 되어 0 으로 끊었습니다. 계약서를 확인하십시오.")
    if tm.p_mode == "accrue" and tm.p_yield > 0 and tm.p_yield < tm.cpn - 1e-9:
        w.append(f"조기상환 보장수익률({tm.p_yield:.2%})이 표면이자율({tm.cpn:.2%})보다 "
                 "낮습니다. 조기상환금액이 액면 밑으로 내려갑니다.")
    if tm.k_w > 0 and tm.k_prem > 0 and tm.k_prem < tm.cpn - 1e-9:
        w.append(f"매도청구 프리미엄({tm.k_prem:.2%})이 표면이자율({tm.cpn:.2%})보다 "
                 "낮습니다. 매도청구금액이 액면 밑으로 내려갑니다.")
    if put_bdt_on(tm):
        if tm.bdt_sig <= 0:
            w.append("BDT 변동성이 0 입니다. 금리 고정 격자와 같은 값이 나옵니다.")
        elif tm.bdt_sig > 1.0:
            w.append(f"BDT 변동성이 {tm.bdt_sig:.0%} 입니다. 로그정규 변동성이라 "
                     "보통 10~30% 를 씁니다. 단위를 확인하십시오.")
        try:
            _g = pick(engine(tm, conv=False, put=True, call=False), tm.model)
            _b = bond_bdt(tm, True)
            if _b < _g - 1e-6:
                w.append(f"BDT 부채요소({_b:,.2f})가 금리 고정 격자({_g:,.2f})보다 "
                         "작습니다. 옵션가치는 음수가 될 수 없으므로 캘리브레이션을 "
                         "확인해야 합니다.")
        except Exception:
            pass
    if tm.floor > tm.K0: w.append("최저 조정가액이 최초 전환가액보다 큽니다.")
    if tm.par > tm.floor: w.append("액면가가 최저 조정가액보다 큽니다. 액면가가 하한으로 작동합니다.")
    if tm.rfx_mode > 0 and round(tm.rfx_cyc*tm.n/(tm.T*12)) < 1:
        w.append("조정 주기가 노드 간격보다 짧습니다. 노드를 늘리십시오.")
    if round(tm.p_f*tm.n/(tm.T*12)) < 1: w.append("조기상환 주기가 노드 간격보다 짧습니다.")
    if tm.k_w > 0 and round(tm.k_f*tm.n/(tm.T*12)) < 1:
        w.append("매도청구 주기가 노드 간격보다 짧습니다.")
    if tm.sig <= 0.01: w.append("변동성이 지나치게 낮습니다.")
    if tm.sig > 2: w.append("변동성이 200%를 넘습니다. 단위를 확인하십시오.")
    if tm.carry == 0 and tm.rfx_mode > 0 and tm.n > 120:
        w.append("상태확장은 노드 120개까지 권장합니다. 근사 방법을 고르거나 노드를 줄이십시오.")
    if tm.p_mode == "accrue" and tm.p_yield <= 0:
        w.append("조기상환 보장수익률이 0입니다. 복리 방식을 쓸 이유가 없습니다.")
    # 위험 곡선이 무위험보다 낮으면 두 곡선을 바꿔 넣은 것이다. 그대로 두면
    # 조기상환권과 전환권이 뒤집힌 값으로 나온다.
    if tm.rf_curve and tm.cr_curve:
        try:
            RF, CR = curves(tm)
            bad = [x for x in (0.25, 0.5, 1, 2, 3, 5, 7, 10)
                   if x <= tm.T + 1e-9 and CR(x) < RF(x) - 1e-9]
            if bad:
                w.append(f"신용스프레드가 음수인 구간이 있습니다 "
                         f"({bad[0]:g}년 무위험 {RF(bad[0]):.2%} > 위험 {CR(bad[0]):.2%}). "
                         f"두 곡선을 바꿔 넣지 않았는지 확인하십시오.")
        except Exception:
            pass
    return w


# ══════════════════════════════════════════════════════════
# 4. 주가·변동성
# ══════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def korean_font():
    """그림에 쓸 한글 글꼴 이름. 없으면 None.

    이름만 보고 고르면 글꼴이 있어도 한글 글리프가 없어 네모로 나온다.
    matplotlib 에 딸린 FT2Font 로 '가'(U+AC00) 가 실제로 있는지 확인한다.
    """
    try:
        from matplotlib import font_manager as fm
        from matplotlib.ft2font import FT2Font
    except Exception:
        return None
    pref = ("NanumGothic", "NanumBarunGothic", "NanumSquare", "Malgun Gothic",
            "AppleGothic", "Apple SD Gothic Neo", "Noto Sans CJK KR",
            "Noto Sans KR", "Source Han Sans KR", "UnDotum", "Baekmuk Gulim",
            "WenQuanYi Zen Hei", "Droid Sans Fallback")

    def has_hangul(path):
        try:
            return 0xAC00 in FT2Font(path).get_charmap()
        except Exception:
            return False

    byname = {}
    for f in fm.fontManager.ttflist:
        byname.setdefault(f.name, f.fname)
    for nm in pref:
        if nm in byname and has_hangul(byname[nm]):
            return nm
    for nm, path in sorted(byname.items()):
        if has_hangul(path):
            return nm
    return None


def use_korean_font():
    """그림을 그리기 전에 부른다. 글꼴을 찾았으면 이름, 못 찾았으면 None."""
    nm = korean_font()
    try:
        import matplotlib
        matplotlib.rcParams["axes.unicode_minus"] = False   # 음수 부호도 네모가 된다
        if nm:
            matplotlib.rcParams["font.family"] = nm
    except Exception:
        pass
    return nm


@st.cache_data(show_spinner=False, ttl=3600)
def fetch_prices(code: str, days: int, market: str, end: str = None):
    """야후 파이낸스에서 수정주가를 받는다.

    auto_adjust=True 라 유상증자·액면분할·배당이 반영된 종가가 온다.
    국내 종목은 코스닥 .KQ, 코스피 .KS 를 차례로 시도한다.
    """
    try:
        import yfinance as yf
    except ImportError:
        raise RuntimeError(
            "yfinance 가 설치되어 있지 않습니다. requirements.txt 에 "
            "yfinance>=0.2.40,<0.2.58 을 넣고 앱을 다시 시작하십시오. "
            "그동안은 아래에서 주가 파일을 넣으시면 됩니다.") from None
    d2 = dt.date.fromisoformat(end) if end else dt.date.today()
    d1 = d2 - dt.timedelta(days=int(days*1.7)+30)
    # 고른 시장을 먼저 보되 비면 다른 시장도 해 본다. 이전 상장이나 시장 이관이
    # 있으면 접미사가 어긋나는데, 화면에서는 "자료 없음" 으로만 보여 원인을 못 찾는다.
    if not code.isdigit():
        sufs = [""]
    else:
        sufs = [f".{market}"] if market else []
        sufs += [x for x in (".KQ", ".KS") if x not in sufs]
    errs = []
    for suf in sufs:
        sym = code + suf
        try:
            df = yf.download(sym, start=d1, end=d2+dt.timedelta(days=1),
                             progress=False, auto_adjust=True, threads=False)
            if df is None or df.empty:
                errs.append(f"{sym} 자료 없음"); continue
            if hasattr(df.columns, "nlevels") and df.columns.nlevels > 1:
                df = df.droplevel(1, axis=1)
            col = "Close" if "Close" in df.columns else df.columns[0]
            sr = df[col].dropna()
            rows = [(i.strftime("%Y-%m-%d"), float(v)) for i, v in sr.items() if v > 0]
            if len(rows) >= 10:
                return rows[-days:], f"야후 {sym} · 수정주가"
            errs.append(f"{sym} {len(rows)}개")
        except Exception as e:
            errs.append(f"{sym} {e}")
    # 무엇을 해 봤고 왜 안 됐는지 밝힌다. 옛 yfinance 는 야후가 API 를 바꾸면
    # 예외 없이 빈 표를 돌려주므로, 판 번호가 있어야 원인을 가릴 수 있다.
    raise RuntimeError((" / ".join(errs) or "자료를 찾지 못했습니다")
                       + f"  (yfinance {getattr(yf, '__version__', '?')})")


def vol_from(prices, tdays=250, drop_outlier=True):
    px = [p for _, p in prices]
    if len(px) < 10: return None
    r = np.diff(np.log(px))
    removed, lo, hi = 0, None, None
    if drop_outlier:
        M = float(np.median(r)); mad = float(np.median(np.abs(r-M)))*1.4826
        lo, hi = M-2.5*mad, M+2.5*mad   # 사례 5-2 와 같은 배수
        keep = (r >= lo) & (r <= hi); removed = int((~keep).sum()); r = r[keep]
    sd = float(np.std(r, ddof=1))
    return dict(daily=sd, annual=sd*math.sqrt(tdays), n=len(px)-1,
                removed=removed, lo=lo, hi=hi)


def read_upload(name: str, data: bytes) -> str:
    """업로드 파일을 텍스트로 편다.

    엑셀은 시트를 탭 구분 텍스트로 바꾼다. csv·txt 는 한글 인코딩을 차례로
    시도한다. 증권사·거래소에서 내려받은 파일은 대개 CP949 다.
    """
    low = name.lower()
    if low.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        out = []
        for row in wb.worksheets[0].iter_rows(values_only=True):
            cells = ["" if v is None else
                     (v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v))
                     for v in row]
            if any(cells): out.append("\t".join(cells))
        wb.close()
        return "\n".join(out)
    if low.endswith(".xls"):
        try:
            import xlrd                              # 옛 형식은 xlrd 만 읽는다
        except ImportError:
            raise ValueError(
                "옛 형식(.xls)을 읽으려면 xlrd 가 필요합니다. requirements.txt 에 "
                "xlrd>=2.0 을 넣고 앱을 다시 시작하시거나, 엑셀에서 .xlsx 로 "
                "저장해 다시 넣으십시오.") from None
        bk = xlrd.open_workbook(file_contents=data)
        sh = bk.sheet_by_index(0)
        out = []
        for r in range(sh.nrows):
            cells = []
            for c in range(sh.ncols):
                v, ty = sh.cell_value(r, c), sh.cell_type(r, c)
                if ty == xlrd.XL_CELL_DATE:
                    y, mo, d = xlrd.xldate_as_tuple(v, bk.datemode)[:3]
                    v = f"{y:04d}-{mo:02d}-{d:02d}"
                elif isinstance(v, float) and v == int(v) and abs(v) < 1e15:
                    v = int(v)
                cells.append("" if v is None else str(v))
            if any(cells): out.append("\t".join(cells))
        return "\n".join(out)
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            txt = data.decode(enc)
            if "�" not in txt: return txt
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", "ignore")


def parse_prices(txt: str):
    out, close_idx, start = [], -1, 0
    lines = [l.strip() for l in txt.splitlines() if l.strip()]
    import re
    def cut(l):
        l = re.sub(r'"([^"]*)"', lambda m: m.group(1).replace(",", ""), l)
        return [x.strip() for x in re.split(r"[\t,;]|\s{2,}", l) if x.strip()]
    isdate = lambda x: bool(re.match(r"^\d{4}[-./]\d{1,2}[-./]\d{1,2}$", x.strip()))
    for i, l in enumerate(lines[:3]):
        c = cut(l)
        k = [j for j, x in enumerate(c)
             if re.match(r"^(종가|현재가|close|adj\s*close)$", x.strip(), re.I)]
        if k: close_idx, start = k[0], i+1; break
        if any(re.search(r"[가-힣A-Za-z]{2,}", x) for x in c) and not any(isdate(x) for x in c):
            start = i+1
    for l in lines[start:]:
        c = cut(l)
        if not c: continue
        di = next((j for j, x in enumerate(c) if isdate(x)), -1)
        close = None
        if 0 <= close_idx < len(c):
            try: close = float(c[close_idx].replace(",", "").replace("원", ""))
            except Exception: close = None
        if close is None or close <= 0:
            for k in range(di+1 if di >= 0 else 0, len(c)):
                if isdate(c[k]): continue
                try:
                    v = float(c[k].replace(",", "").replace("원", ""))
                    if v > 0: close = v; break
                except Exception: pass
        if close is None or close <= 0: continue
        d = c[di].replace(".", "-").replace("/", "-") if di >= 0 else ""
        if d:
            p = d.split("-"); d = f"{p[0]}-{int(p[1]):02d}-{int(p[2]):02d}"
        out.append((d, close))
    if len(out) >= 2 and out[0][0] and out[-1][0] and out[0][0] > out[-1][0]:
        out.reverse()
    return out


def parse_prices_multi(txt: str):
    """여러 종목의 종가가 한 파일에 들어 있을 때 열마다 갈라 읽는다.

    첫 줄이 머리글이고 첫 열이 일자, 나머지 열이 종목별 종가인 형태를 본다.
    야후·거래소·증권사에서 여러 종목을 한 번에 내려받으면 대개 이 꼴이다.
    """
    import re
    lines = [l for l in (x.rstrip() for x in txt.splitlines()) if l.strip()]
    if not lines: return []

    def cut(l):
        l = re.sub(r'"([^"]*)"', lambda m: m.group(1).replace(",", ""), l)
        return [x.strip() for x in re.split(r"[\t,;]|\s{2,}", l)]

    isdate = lambda x: bool(re.match(r"^\d{4}[-./]\d{1,2}[-./]\d{1,2}$", x.strip()))
    hdr = cut(lines[0])
    body = [cut(l) for l in lines[1:]]
    body = [c for c in body if c and isdate(c[0])]
    if len(body) < 10 or len(hdr) < 3: return []
    out = []
    for j in range(1, len(hdr)):
        nm = hdr[j].strip() or f"열{j}"
        rows = []
        for c in body:
            if j >= len(c): continue
            try:
                v = float(c[j].replace(",", "").replace("원", ""))
            except Exception:
                continue
            if v <= 0: continue
            d = c[0].replace(".", "-").replace("/", "-").split("-")
            rows.append((f"{d[0]}-{int(d[1]):02d}-{int(d[2]):02d}", v))
        if len(rows) >= 10:
            if rows[0][0] > rows[-1][0]: rows.reverse()
            out.append((nm, rows))
    return out


def parse_yields(txt: str, unit: str = "auto"):
    """만기별 수익률을 읽는다.

    받는 형태
        만기(년)  수익률(%)                 →  2열
        만기일  만기(개월)  수익률(%)        →  3열 (한국은행·금투협 표를 그대로 붙여넣는 형태)
    unit 이 auto 면 만기 값이 40을 넘는 항목이 있을 때 개월로 본다.
    """
    import re
    rows = []
    isdate = lambda x: bool(re.match(r"^\d{4}[-./]\d{1,2}[-./]\d{1,2}$", x.strip()))
    for line in txt.splitlines():
        c = [x for x in re.split(r"[\t,;]|\s{2,}|\s", line.strip()) if x]
        if not c: continue
        c = [x.replace("%", "").replace(",", "") for x in c]
        c = [x for x in c if not isdate(x)]          # 날짜 열은 버린다
        nums = []
        for x in c:
            try: nums.append(float(x))
            except ValueError: pass
        if len(nums) >= 2:
            rows.append((nums[-2], nums[-1]))        # 뒤에서 둘 = 만기, 수익률
    if not rows: return []
    mx = max(r[0] for r in rows)
    months = (unit == "month") or (unit == "auto" and mx > 40)
    out = [((m/12 if months else m), y/100) for m, y in rows if m > 0]
    return sorted(out)


KIS_TENORS = {"3월": 3, "6월": 6, "9월": 9, "1년": 12, "1년6월": 18, "2년": 24,
              "2년6월": 30, "3년": 36, "4년": 48, "5년": 60, "7년": 84,
              "10년": 120, "15년": 180, "20년": 240, "30년": 360, "50년": 600}


def read_kisnet(name: str, data: bytes):
    """KIS-Net 채권시가평가 기준수익률 표를 읽는다.

    첫 시트가 ``종류 · 종류명 · 신용등급 · 고시기관 · 3월 · 6월 · … · 50년`` 이고
    금리는 % 단위, 값이 없으면 ``-`` 다. 국채 한 줄과 회사채 등급별 여러 줄이
    한 표에 같이 있으므로, 어느 줄을 무위험으로 쓰고 어느 줄을 위험으로 쓸지는
    화면에서 고른다.

    돌려주는 것은 ``[(라벨, [(만기(년), 수익률), …]), …]`` 이다.
    """
    ext = name.lower().rsplit(".", 1)[-1]
    if ext == "xls":
        try:
            import xlrd                              # 옛 형식은 xlrd 만 읽는다
        except ImportError:
            raise ValueError(
                "옛 형식(.xls)을 읽으려면 xlrd 가 필요합니다. requirements.txt 에 "
                "xlrd>=2.0 을 넣고 앱을 다시 시작하시거나, 엑셀에서 .xlsx 로 "
                "저장해 다시 넣으십시오.") from None
        sh = xlrd.open_workbook(file_contents=data).sheet_by_index(0)
        grid = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                for r in range(sh.nrows)]
    else:
        import openpyxl
        ws = openpyxl.load_workbook(io.BytesIO(data), data_only=True).worksheets[0]
        grid = [[c if c is not None else "" for c in row]
                for row in ws.iter_rows(values_only=True)]
    if not grid: raise ValueError("빈 파일입니다.")

    # 머리 행 — 만기 이름이 가장 많이 걸리는 줄
    hi, cols = -1, {}
    for i, row in enumerate(grid[:10]):
        got = {j: KIS_TENORS[str(v).strip()]
               for j, v in enumerate(row) if str(v).strip() in KIS_TENORS}
        if len(got) > len(cols): hi, cols = i, got
    if len(cols) < 3:
        raise ValueError("만기 열(3월·6월·1년 …)을 찾지 못했습니다. "
                         "KIS-Net 기준수익률 표의 첫 시트인지 확인하십시오.")

    out = []
    for row in grid[hi+1:]:
        head = [str(row[j]).strip() for j in range(min(3, len(row)))]
        head = [h for h in head if h and h != "-"]
        if not head: continue
        pts = []
        for j, mth in sorted(cols.items(), key=lambda x: x[1]):
            if j >= len(row): continue
            try: y = float(str(row[j]).replace(",", "").replace("%", ""))
            except ValueError: continue
            if y > 0: pts.append((mth/12, y/100))
        if len(pts) >= 2:
            out.append((" · ".join(head[:3]).replace("*", ""), pts))
    if not out: raise ValueError("수익률 행을 찾지 못했습니다.")
    return out


def curve_text(pts):
    """곡선을 화면 입력 형식(개월 · 수익률%)으로 되돌린다."""
    return "\n".join(f"{round(t*12):d}\t{y*100:.3f}%" for t, y in pts)


# ══════════════════════════════════════════════════════════
# 4-2. 금리변동성 — BDT 의 σ 를 시계열에서 뽑는다
# ══════════════════════════════════════════════════════════
# 할인율은 이미 등급보간(blend_curves) → 만기보간(_lin) 두 번을 거친다.
# 변동성도 같은 자료·같은 보간에서 나와야 조서가 하나로 이어진다.
#
# 순서가 중요하다. **보간을 먼저 하고 변동성을 나중에** 구한다. 변동성은
# 선형 함수가 아니라 등급별 σ 를 보간하면 값이 달라진다 — 내삽이면 2% 안쪽이지만
# 외삽하면 9% 가까이 벌어진다.

TENOR_MO = dict(KIS_TENORS)
TENOR_MO.update({"1개월": 1, "3개월": 3, "6개월": 6, "9개월": 9,
                 "1년6개월": 18, "2년6개월": 30, "18월": 18, "30월": 30})


def _tenor_years(label) -> float:
    """만기 머리글을 연 단위로. '3년' · '1년6개월' · '0.25' · '36'(개월) 을 받는다."""
    s = str(label).strip()
    if not s: return None
    if s in TENOR_MO: return TENOR_MO[s]/12
    m = re.match(r"^\s*(\d+)\s*년\s*(?:(\d+)\s*개?월)?\s*$", s)
    if m: return int(m.group(1)) + (int(m.group(2) or 0))/12
    m = re.match(r"^\s*(\d+)\s*개?월\s*$", s)
    if m: return int(m.group(1))/12
    try:
        v = float(s.replace(",", ""))
    except Exception:
        return None
    # 12 보다 크면 개월로 본다. 만기 곡선에 12년 이상은 드물다.
    return v/12 if v > 12 else (v if v > 0 else None)


def parse_rate_panel(txt: str):
    """일자 × (등급·만기) 표를 읽는다. 머리 줄이 여러 개여도 된다.

    금투협 시계열은 열마다 '회사채 I(공모사채) /무보증 / BBB0' 같은 등급명과
    '5년' 같은 만기가 **서로 다른 머리 줄**에 있다. 그래서 열별로 머리 줄을
    모두 모아 등급과 만기를 찾는다. 한 파일에 등급이 여럿이어도 갈라 읽고,
    등급 표기가 없으면 만기만 읽어 종전의 일자 × 만기 표와 같아진다.

    반환은 ([(등급, 만기(년), 열번호)], [(일자, [값…])]) 이고, 등급·만기는
    못 찾으면 None 이다. 수익률은 % 단위 그대로다.
    """
    lines = [l for l in (x.rstrip() for x in txt.splitlines()) if l.strip()]
    if not lines: return [], []

    def cut(l):
        l = re.sub(r'"([^"]*)"', lambda m: m.group(1).replace(",", ""), l)
        return [x.strip() for x in re.split(r"[\t,;]|\s{2,}", l)]

    isdate = lambda x: bool(re.match(r"^\d{4}[-./]\d{1,2}[-./]\d{1,2}$", x.strip()))
    heads, body = [], []
    for l in lines:
        c = cut(l)
        if c and isdate(c[0]):
            body.append(c)
        elif not body and len(c) >= 2:
            heads.append(c)                          # 자료가 나오기 전은 다 머리다
    if not heads or len(body) < 10: return [], []
    ncol = max([len(h) for h in heads] + [len(b) for b in body])
    cols = []
    for j in range(1, ncol):
        cells = [h[j] for h in heads if j < len(h)]
        rt = next((r for r in (rating_in(c) for c in cells) if r), None)
        tn = next((y for y in (_tenor_years(c) for c in cells) if y), None)
        if rt or tn: cols.append((rt, tn, j))
    if not cols: return [], []
    rows = []
    for c in body:
        vals = []
        for _, _, j in cols:
            try:
                vals.append(float(c[j].replace(",", "").replace("%", "")))
            except Exception:
                vals.append(None)
        if any(v is not None and v > 0 for v in vals):
            d = re.split(r"[-./]", c[0].strip())
            rows.append((f"{int(d[0]):04d}-{int(d[1]):02d}-{int(d[2]):02d}", vals))
    if len(rows) >= 2 and rows[0][0] > rows[-1][0]: rows.reverse()
    return cols, rows


def panel_series(cols, rows, T: float):
    """패널을 등급별 고정만기 시계열로 접는다.

    한 등급에 만기가 여럿이면 그날 곡선에서 잔존만기 T 지점을 뽑고, 하나뿐이면
    그 만기를 그대로 쓴다 — 고시표가 5년 한 열만 주는 일이 흔하다. 만기를
    먼저 맞추고 변동성은 나중에 계산해야 만기 이동 효과가 σ 에 섞이지 않는다.

    반환은 ({등급: [(일자, 금리)]}, {등급: [만기(년)…]}) 이다.
    """
    byr = {}
    for k, (rt, tn, _) in enumerate(cols):
        if tn is not None: byr.setdefault(rt, []).append((tn, k))
    ser, tens = {}, {}
    for rt, items in byr.items():
        items.sort()
        out = []
        for d, vals in rows:
            pts = [(tn, vals[k]) for tn, k in items
                   if vals[k] is not None and vals[k] > 0]
            if not pts: continue
            y = _lin(pts, T)
            if y and y > 0: out.append((d, y))
        if len(out) >= 10:
            ser[rt] = out
            tens[rt] = [tn for tn, _ in items]
    return ser, tens


def cm_series(tenors, rows, T: float):
    """고정만기 시계열 — 매일 그날 곡선에서 잔존만기 T 지점을 뽑는다.

    과거로 갈 때 만기를 함께 늘리면 금리 변동이 아니라 만기 이동 효과가
    변동성에 섞인다. 국고채 지표금리를 만기 고정으로 고시하는 것과 같은 이유다.
    """
    out = []
    for d, vals in rows:
        pts = [(t, v) for t, v in zip(tenors, vals) if v is not None and v > 0]
        if len(pts) < 2: continue
        pts.sort()
        y = _lin(pts, T)
        if y and y > 0: out.append((d, y))
    return out


def blend_series(sa, sb, ra: str, rb: str, rt: str):
    """두 등급의 고정만기 시계열을 노치 거리로 섞는다.

    blend_curves 와 같은 가중치를 쓴다. 날짜가 둘 다 있는 날만 남긴다.
    """
    ia, ib, it = rating_idx(ra), rating_idx(rb), rating_idx(rt)
    if not sb: return list(sa)
    if not sa: return list(sb)
    if ia < 0 or ib < 0 or it < 0 or ia == ib: return list(sa)
    w = (it-ia)/(ib-ia)
    db = dict(sb)
    return [(d, y + (db[d]-y)*w) for d, y in sa if d in db]


def rate_vol(series, tdays=250, drop=True):
    """금리 시계열의 변동성. 상대(로그정규)와 절대(정규)를 함께 준다.

    BDT 의 σ 는 **상대** 변동성이다. 실무에서 bp 로 말하는 절대 변동성을
    그대로 넣으면 크게 어긋나므로 둘을 나란히 보여 준다.

        상대 ≈ 절대 ÷ 평균금리
    """
    v = vol_from(series, tdays, drop)
    if not v: return None
    lv = np.array([y for _, y in series], dtype=float)
    lg = np.diff(np.log(lv))
    dif = np.diff(lv)
    # 상대 변동성이 채택분으로만 계산되므로 절대 변동성도 같은 날만 쓴다.
    # 전체로 계산하면 이상치를 뺀 상대값과 견줄 수 없고, 리포트 수식과도 어긋난다.
    if drop and v.get("lo") is not None:
        keep = (lg >= v["lo"]) & (lg <= v["hi"])
        dif = dif[keep]
    v = dict(v)
    v["abs_daily"] = float(np.std(dif, ddof=1)) if len(dif) > 1 else 0.0
    v["abs_annual"] = v["abs_daily"]*math.sqrt(tdays)
    v["mean"] = float(np.mean(lv))
    v["min"] = float(np.min(lv))
    v["neg"] = int((lv <= 0).sum())
    return v


# ══════════════════════════════════════════════════════════
# 4-1. 리포트 공통 서식
# ══════════════════════════════════════════════════════════
# 조서와 같은 손맛으로 보이도록 색·글꼴·번호서식을 한곳에 모았다.
RPT = dict(
    ink="1F3864", sub="44618C", grey="7F7F7F", amber="9A7200",
    green="1F6B44", red="A6301F", band="EFF3F8", light="F7F9FC",
    tint="E4EBF5", hair="D6DCE5", warm="FFF7E6",
)
R_N0, R_N2, R_N4, R_N6 = "#,##0", "#,##0.00", "#,##0.0000", "0.000000"
R_P2, R_P4 = "0.00%", "0.0000%"
R_YMD = "yyyy-mm-dd"


def report_kit(wb, font="맑은 고딕"):
    """리포트 한 권에 쓸 서식 도구를 만든다.

    put   — 셀 하나. 값·수식·서식·색을 한 번에 준다.
    head  — 표지 제목줄
    sec   — 구역 머리 (색 띠)
    cols  — 표 머리줄
    note  — 회색 주석
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gl

    thin = Side(style="thin", color=RPT["hair"])
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def put(ws, r, c, v, *, fmt=None, bold=False, size=10, color=None,
            fill=None, align=None, border=False, wrap=False, italic=False):
        cl = ws.cell(r, c, v)
        cl.font = Font(name=font, size=size, bold=bold, italic=italic,
                       color=color or RPT["ink"])
        if fmt: cl.number_format = fmt
        if fill: cl.fill = PatternFill("solid", fgColor=fill)
        if align or wrap:
            cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if border: cl.border = box
        return cl

    def head(ws, r, txt, sub=None, span=8):
        put(ws, r, 2, txt, bold=True, size=16, color=RPT["ink"])
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+span)
        if sub:
            put(ws, r+1, 2, sub, size=9.5, color=RPT["grey"], wrap=True)
            ws.merge_cells(start_row=r+1, start_column=2, end_row=r+1, end_column=1+span)
            ws.row_dimensions[r+1].height = 28
        ws.row_dimensions[r].height = 24

    def sec(ws, r, txt, span=8, tone=None):
        for c in range(2, 2+span):
            put(ws, r, c, txt if c == 2 else None, bold=(c == 2), size=10.5,
                color=RPT["ink"], fill=tone or RPT["band"])
        ws.row_dimensions[r].height = 20

    def cols(ws, r, names, widths=None, start=2):
        for i, nm in enumerate(names):
            put(ws, r, start+i, nm, bold=True, size=9, fill=RPT["tint"],
                align="center", border=True, wrap=True)
        if widths:
            for i, w in enumerate(widths):
                ws.column_dimensions[gl(start+i)].width = w
        ws.row_dimensions[r].height = 26

    def note(ws, r, txt, span=8, tone=None):
        put(ws, r, 2, txt, size=9, color=tone or RPT["grey"], wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1+span)
        ws.row_dimensions[r].height = max(16, 14*(1+len(txt)//95))

    def sheet(name, tab=None, widths=None, freeze=None, landscape=False):
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = tab or RPT["sub"]
        ws.column_dimensions["A"].width = 2.2
        if widths:
            for i, w in enumerate(widths):
                ws.column_dimensions[gl(2+i)].width = w
        if freeze: ws.freeze_panes = freeze
        ws.page_setup.orientation = "landscape" if landscape else "portrait"
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        return ws

    return dict(put=put, head=head, sec=sec, cols=cols, note=note, sheet=sheet, gl=gl)


def _brackets(pts, t):
    """t 를 감싸는 입력곡선 두 점의 번호. 범위 밖이면 (i, i) 로 한 점만 준다."""
    if not pts: return (0, 0)
    if t <= pts[0][0]: return (0, 0)
    if t >= pts[-1][0]: return (len(pts)-1, len(pts)-1)
    for i in range(1, len(pts)):
        if t <= pts[i][0]: return (i-1, i)
    return (len(pts)-1, len(pts)-1)


def lerp_formula(t, pts, col, row0, sh=None):
    """엑셀에서 선형보간. 입력 셀을 가리키므로 노란 셀을 고치면 따라 움직인다.

    col 은 값이 든 열 문자, row0 은 첫 점의 행, sh 는 그 표가 있는 시트다.
    범위 밖이면 끝점을 그대로 쓴다 — 앱의 _lin 과 같다.
    """
    q = f"'{sh}'!" if sh else ""
    i, j = _brackets(pts, t)
    if i == j: return f"={q}${col}${row0+i}"
    x0, x1 = pts[i][0], pts[j][0]
    return (f"={q}${col}${row0+i}+({q}${col}${row0+j}-{q}${col}${row0+i})"
            f"*({t:.12g}-{x0:.12g})/({x1:.12g}-{x0:.12g})")


def build_xlsx_vol(series, tdays=250, drop=True, mad_k=2.5, pick="median",
                   applied=None, asof=None, kind="stock", how=None):
    """변동성 산출내역 리포트. 계산이 전부 수식으로 들어간다.

    series 는 [(이름, [(일자, 종가), …]), …] 다. 하나면 대상회사만,
    여럿이면 비상장 평가에서 쓰는 피어 묶음이다.
    pick 은 여러 회사를 하나로 줄이는 방법 — median · mean · max · min.
    """
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    K = report_kit(wb)
    put, head, sec, cols, note, sheet = (K[x] for x in
        ("put", "head", "sec", "cols", "note", "sheet"))
    series = [(nm, px) for nm, px in series if px and len(px) >= 10]
    if not series: raise ValueError("종가가 10개 이상인 계열이 하나도 없습니다.")
    many = len(series) > 1
    YEL = "FFF9DB"
    # 금리 계열이면 절대(정규) 변동성을 한 줄 더 낸다. 실무에서 bp 로 말하는
    # 그 값이고, BDT 가 쓰는 상대(로그정규) 변동성과 헷갈리기 쉬워 나란히 둔다.
    _rate = (kind == "rate")
    UNIT = "금리 (%)" if _rate else "종가"
    RET = "로그변화율" if _rate else "로그수익률"
    PICKS = {"median": "중앙값", "mean": "단순평균", "max": "최댓값", "min": "최솟값"}

    # 회사별 시트의 행 자리. 한곳에서 정해 두고 수식이 이 이름만 쓴다.
    R_TD, R_MK, R_DR = 5, 6, 7          # 거래일수 · MAD 배수 · 이상치 제거
    R_NP, R_NR, R_MD = 8, 9, 10         # 종가 수 · 수익률 수 · 중앙값
    R_MA, R_LO, R_HI = 11, 12, 13       # MAD · 하한 · 상한
    R_EX, R_SD, R_AN = 14, 15, 16       # 제외 · 일 변동성 · 연 변동성
    R_AD, R_AA, R_MN = 17, 18, 19       # 절대 일·연 변동성 · 평균 (금리만)
    HDR = 22 if _rate else 19           # 표 머리
    R0 = HDR + 1                        # 첫 자료행

    names = [f"{i:02d} {_vsafe(nm)}"[:31] for i, (nm, _) in enumerate(series, 1)]

    # ── 표지 ──
    C = sheet("표지", tab=RPT["ink"], widths=[24, 20, 18, 18, 16, 16, 16, 16])
    head(C, 2, "변동성 산출내역",
         "전환사채 평가에 쓸 주가변동성을 로그수익률의 표본표준편차로 구한 내역이다. "
         "노란 셀만 입력이고 나머지는 수식이라, 거래일수나 이상치 배수를 바꾸면 "
         "표 전체가 다시 계산된다.")
    r = 5
    sec(C, r, "산출 요약"); r += 1
    cols(C, r, ["항목", "내용"], [26, 62]); r += 1
    for k, v in ([("평가기준일", (asof or dt.date.today()).isoformat()),
                 ("대상 계열", f"{len(series)}개 — " + " · ".join(nm for nm, _ in series)),
                 ("수익률", f"일별 {RET}  ln({UNIT} ÷ 직전 {UNIT})"),
                 ("이상치 처리", (f"중앙값 절대편차(MAD) × {mad_k:g} 밖을 제외"
                                if drop else "제외하지 않음")),
                 ("표준편차", "표본표준편차 STDEV.S — 자유도 n−1"),
                 ("연환산", f"일 변동성 × √{tdays:g}"),
                  ("종합 방법", PICKS.get(pick, pick) if many else "단일 계열")]
                 + ([("산출 경위", how)] if how else [])
                 + ([("변동성 종류", "상대(로그정규) — BDT 의 σ 다. 절대(정규) "
                                  "변동성도 함께 내되 모형에는 상대를 쓴다")]
                    if _rate else [])):
        put(C, r, 2, k, bold=True, border=True, fill=RPT["light"])
        put(C, r, 3, v, border=True, wrap=True)
        C.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        r += 1
    r += 1
    sec(C, r, "결과"); r += 1
    cols(C, r, ["구분", "연 변동성"], [26, 18]); r += 1
    res = r
    put(C, r, 2, "종합", bold=True, border=True, fill=RPT["warm"])
    put(C, r, 3, ("='종합'!$C$6" if many else f"='{names[0]}'!$C${R_AN}"),
        fmt=R_P2, bold=True, border=True, fill=RPT["warm"], align="right")
    r += 1
    if applied is not None:
        put(C, r, 2, "앱에 적용한 값", bold=True, border=True)
        put(C, r, 3, applied, fmt=R_P2, border=True, align="right", color=RPT["amber"])
        put(C, r+1, 2, "차이", bold=True, border=True)
        put(C, r+1, 3, f"=C{r}-C{res}", fmt=R_P4, border=True, align="right")
        r += 2
    r += 1
    note(C, r, "주황색 숫자는 앱이 넣은 값이고 노란 셀은 바꿔도 되는 입력이다. "
               "종가는 수정주가여야 한다 — 유상증자·액면분할·배당이 반영되지 않은 "
               "종가를 쓰면 그날 하루가 통째로 이상치가 된다.")

    # ── 회사별 시트 ──
    for (nm, px), sn in zip(series, names):
        W = sheet(sn, widths=[15, 13, 14, 13, 9, 14], freeze=f"B{R0}")
        head(W, 2, f"{nm} — 일별 로그수익률", span=6)
        n = len(px); last = R0 + n - 1
        D1, DN = f"$D${R0+1}", f"$D${last}"
        sec(W, 4, "입력과 결과", span=6)
        lab = [(R_TD, "연 거래일수", tdays, R_N0, True),
               (R_MK, "MAD 배수", mad_k, "0.0#", True),
               (R_DR, "이상치 제거 (1/0)", 1 if drop else 0, R_N0, True),
               (R_NP, "관측 종가", f"=COUNT($C${R0}:$C${last})", R_N0, False),
               (R_NR, "수익률", f"=COUNT({D1}:{DN})", R_N0, False),
               (R_MD, "중앙값", f"=MEDIAN({D1}:{DN})", R_N6, False),
               (R_MA, "MAD (×1.4826)", f"=MEDIAN($E${R0+1}:$E${last})*1.4826",
                R_N6, False),
               (R_LO, "정상범위 하한", f"=$C${R_MD}-$C${R_MK}*$C${R_MA}", R_N6, False),
               (R_HI, "정상범위 상한", f"=$C${R_MD}+$C${R_MK}*$C${R_MA}", R_N6, False),
               (R_EX, "제외 개수",
                f"=COUNT({D1}:{DN})-COUNT($G${R0+1}:$G${last})", R_N0, False),
               (R_SD, "일 변동성", f"=STDEV.S($G${R0+1}:$G${last})", R_P4, False),
               (R_AN, "연 변동성", f"=$C${R_SD}*SQRT($C${R_TD})", R_P2, False)]
        if _rate:
            lab += [(R_AD, "절대 일 변동성 (%p)",
                     f"=STDEV.S($H${R0+1}:$H${last})", R_N4, False),
                    (R_AA, "절대 연 변동성 (%p)",
                     f"=$C${R_AD}*SQRT($C${R_TD})", R_N4, False),
                    (R_MN, "평균 금리 (%)",
                     f"=AVERAGE($C${R0}:$C${last})", R_N4, False)]
        for rr, k, v, fm, inp in lab:
            fin = (rr == R_AN)
            put(W, rr, 2, k, bold=True, border=True,
                fill=(YEL if inp else (RPT["warm"] if fin else RPT["light"])))
            put(W, rr, 3, v, fmt=fm, border=True, align="right", bold=fin,
                fill=(YEL if inp else (RPT["warm"] if fin else None)))
        note(W, (R_MN if _rate else R_AN)+1,
             "MAD 는 중앙값 절대편차에 1.4826 을 곱해 정규분포의 표준편차와 눈금을 "
             "맞춘 값이다. 중앙값과 MAD 는 제외 전 전체 " + RET + " 로 구하고, "
             "표준편차만 채택분으로 구한다."
             + ("  절대 변동성은 로그가 아니라 금리 차이(%p)의 표준편차다. "
                "상대 ≈ 절대 ÷ 평균금리 로 환산된다 — BDT 에는 **상대**를 넣는다."
                if _rate else ""), span=(7 if _rate else 6))
        _hd = ["일자", UNIT, RET, "|편차|", "채택", f"채택 {RET}"]
        _wd = [15, 13, 14, 13, 9, 14]
        if _rate: _hd, _wd = _hd + ["채택 변화분"], _wd + [13]
        cols(W, HDR, _hd, _wd)
        for i, (d, v) in enumerate(px):
            rr = R0 + i
            put(W, rr, 2, (dt.date.fromisoformat(d) if d else None),
                fmt=R_YMD, border=True, align="center")
            put(W, rr, 3, v, fmt=R_N2, border=True, align="right")
            if i == 0: continue
            put(W, rr, 4, f"=LN(C{rr}/C{rr-1})", fmt=R_N6, border=True, align="right")
            put(W, rr, 5, f"=ABS(D{rr}-$C${R_MD})", fmt=R_N6, border=True, align="right")
            put(W, rr, 6, f"=IF($C${R_DR}=0,1,"
                          f"IF(AND(D{rr}>=$C${R_LO},D{rr}<=$C${R_HI}),1,0))",
                fmt=R_N0, border=True, align="center")
            put(W, rr, 7, f'=IF(F{rr}=1,D{rr},"")', fmt=R_N6, border=True, align="right")
            if _rate:
                # 절대 변동성용 — 로그가 아니라 금리 차이(%p) 다
                put(W, rr, 8, f'=IF(F{rr}=1,C{rr}-C{rr-1},"")', fmt=R_N4,
                    border=True, align="right")

    # ── 종합 ──
    if many:
        S = sheet("종합", tab=RPT["green"], widths=[8, 26, 18, 14, 12])
        head(S, 2, "피어 종합",
             "비상장이라 대상회사 주가가 없을 때, 유사기업의 변동성을 모아 하나로 줄인다.",
             span=5)
        fn = {"median": "MEDIAN", "mean": "AVERAGE",
              "max": "MAX", "min": "MIN"}.get(pick, "MEDIAN")
        r1, r2 = 9, 9 + len(series) - 1
        put(S, 5, 2, "종합 방법", bold=True, border=True, fill=YEL)
        put(S, 5, 3, PICKS.get(pick, pick), border=True, fill=YEL)
        put(S, 6, 2, "적용 변동성", bold=True, border=True, fill=RPT["warm"])
        put(S, 6, 3, f"={fn}($E${r1}:$E${r2})", fmt=R_P2, bold=True,
            border=True, align="right", fill=RPT["warm"])
        cols(S, 8, ["번호", "회사", "시트", "연 변동성", "수익률"], [8, 26, 18, 14, 12])
        for i, ((nm, px), sn) in enumerate(zip(series, names)):
            rr = r1 + i
            put(S, rr, 2, i+1, fmt=R_N0, border=True, align="center")
            put(S, rr, 3, nm, border=True)
            put(S, rr, 4, sn, border=True, size=9, color=RPT["grey"])
            put(S, rr, 5, f"='{sn}'!$C${R_AN}", fmt=R_P2, border=True, align="right")
            put(S, rr, 6, f"='{sn}'!$C${R_NR}", fmt=R_N0, border=True, align="right")
        note(S, r2+2, "중앙값은 한 회사의 급등락에 덜 흔들린다. 평균을 쓰려면 왜 그 "
                      "회사들이 대상회사와 같은 위험을 진다고 보는지 조서에 남긴다. "
                      "업종·규모·상장기간이 크게 다른 회사는 빼는 편이 낫다.", span=5)
    return _save(wb)


def _vsafe(s):
    """시트 이름에 쓸 수 없는 글자를 턴다."""
    out = "".join(c for c in str(s) if c not in r'[]:*?/\\')
    return out.strip() or "계열"


def polish_wb(wb):
    """조서·리포트 마무리. 탭 색으로 갈래를 나누고 인쇄를 폭 맞춤으로 둔다.

    시트가 스무 장을 넘으면 탭 이름만으로는 어디가 어딘지 안 보인다.
    입력(노랑) · 트리(회색) · 결과(초록) · 회계(빨강)로 갈라 놓는다.
    """
    tone = {"해설": RPT["ink"], "가정": RPT["amber"], "도달확률": "9AA4AE",
            "결과": RPT["green"], "회계처리": RPT["red"],
            "상각표": RPT["sub"], "이자율곡선": RPT["sub"], "표지": RPT["ink"]}
    for ws in wb.worksheets:
        nm = ws.title
        if ws.sheet_properties.tabColor is None:
            ws.sheet_properties.tabColor = tone.get(
                nm, "B7C0CC" if nm[:2].isdigit() else RPT["sub"])
        try:
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.print_options.horizontalCentered = True
            ws.oddHeader.left.text = nm
            ws.oddHeader.left.size = 9
            ws.oddHeader.left.color = "7F7F7F"
            ws.oddFooter.right.text = "&P / &N"
            ws.oddFooter.right.size = 9
        except Exception:
            pass
    if wb.worksheets: wb.active = 0
    return wb


def _save(wb):
    polish_wb(wb)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def build_xlsx_rate(tm: Terms):
    """선도이자율 산출내역 리포트.

    만기수익률 곡선 → 선형보간 → 부트스트래핑 → 연속복리 현물 → 구간 선도.
    조서 트리 시트 11·12행에 값으로 박히는 선도이자율이 어디서 나왔는지
    한 장씩 펼쳐 보여 준다. 노란 셀을 고치면 끝까지 따라 움직인다.
    """
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    K = report_kit(wb)
    put, head, sec, cols, note, sheet = (K[x] for x in
        ("put", "head", "sec", "cols", "note", "sheet"))
    YEL = "FFF9DB"
    derive(tm)
    n, T = int(tm.n), tm.T
    dt_ = T/n
    cc = credit_curve(tm)
    if len(tm.rf_curve) < 2 or len(cc) < 2:
        raise ValueError("무위험·위험 곡선을 각각 두 점 이상 넣어야 합니다.")
    spot_in = (tm.y_type == "spot")
    IN = "입력곡선"
    R0IN = 8                                   # 입력곡선 첫 자료행
    LEG = [("무위험", tm.rf_curve, int(tm.cmp_rf), "B", "C"),
           ("위험", cc, int(tm.cmp_cr), "E", "F")]

    # ── 표지 ──
    C = sheet("표지", tab=RPT["ink"], widths=[26, 22, 18, 18, 16, 16, 16, 16])
    head(C, 2, "이자율 산출내역",
         "만기수익률 곡선에서 할인계수를 순차로 풀고(부트스트래핑), 연속복리 "
         "현물이자율로 바꾼 뒤, 격자 한 구간의 선도이자율을 뽑는 과정이다. "
         "조서 트리 시트 11·12행에 값으로 들어가는 숫자가 여기서 나온다.")
    r = 5
    sec(C, r, "방법"); r += 1
    cols(C, r, ["단계", "내용"], [26, 64]); r += 1
    steps = ([("1. 입력", "현물이자율(제로커브)을 고시된 그대로 받는다"),
              ("2. 연속환산", "연속 = m · ln(1 + r ÷ m).  복리 횟수 m 을 무시하고 "
                            "ln(1+r) 로만 바꾸면 할인계수와 위험중립확률이 어긋난다"),
              ("3. 보간", "고시 만기 사이는 직선으로 잇는다"),
              ("4. 선도", "f(t₀,t₁) = [ r(t₁)·t₁ − r(t₀)·t₀ ] ÷ (t₁ − t₀)")]
             if spot_in else
             [("1. 입력", "만기수익률(YTM) 곡선. 고시 만기 사이는 직선으로 잇는다"),
              ("2. 부트스트래핑", "1 = c·(DF₁+…+DF_k) + DF_k 를 앞에서부터 순차로 푼다. "
                               "c 는 그 만기 수익률 ÷ 연 이표 횟수"),
              ("3. 현물", "연속복리 현물  r(t) = −ln(DF) ÷ t"),
              ("4. 선도", "f(t₀,t₁) = [ r(t₁)·t₁ − r(t₀)·t₀ ] ÷ (t₁ − t₀)")])
    for k, v in steps:
        put(C, r, 2, k, bold=True, border=True, fill=RPT["light"])
        put(C, r, 3, v, border=True, wrap=True)
        C.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        C.row_dimensions[r].height = 30
        r += 1
    r += 1
    sec(C, r, "설정"); r += 1
    cols(C, r, ["항목", "값"], [26, 24]); r += 1
    for k, v in [("입력 유형", "현물이자율(제로커브)" if spot_in else "만기수익률(YTM)"),
                 ("무위험 복리 횟수 (연)", int(tm.cmp_rf)),
                 ("위험 복리 횟수 (연)", int(tm.cmp_cr)),
                 ("위험 곡선 방식", {"pick": "표에서 등급 하나",
                                  "rating": "두 등급 보간"}.get(tm.rate_mode,
                                                          "직접 입력")),
                 ("위험 곡선 출처", tm.cr_src or "직접 입력"),
                 ("평가기준일", tm.d_base), ("만기일", tm.d_mat),
                 ("잔존기간 T (년)", round(T, 8)), ("노드 수 n", n),
                 ("한 구간 Δt (년)", round(dt_, 8)),
                 ("주가 변동성 σ", tm.sig)]:
        put(C, r, 2, k, bold=True, border=True, fill=RPT["light"])
        put(C, r, 3, v, border=True, fmt=(R_P2 if k == "주가 변동성 σ" else None),
            align=None if isinstance(v, str) else "right", wrap=True)
        r += 1
    r += 1
    note(C, r, "노란 셀만 입력이다. 만기와 수익률을 고치면 부트스트래핑부터 선도까지 "
               "전부 다시 계산된다. 다만 만기 칸을 늘리거나 줄이려면 앱에서 곡선을 "
               "바꿔 리포트를 다시 만들어야 한다 — 표의 길이는 구조라 수식으로 늘지 않는다.")
    r += 1
    note(C, r, "주가 변동성 σ 는 이자율 산출에 쓰이지 않는다. 선도이자율 시트의 "
               "위험중립가중치 q = [exp(f·Δt) − d] ÷ (u − d) 에서 u·d 를 만드는 데만 "
               "쓴다 — q 가 이자율과 주가를 잇는 자리라 여기 함께 적어 둔다.")
    r += 1
    note(C, r, "위험 곡선 출처는 앱의 이자율 칸에서 고른 그대로다. 두 등급 보간이면 "
               "아래 입력곡선 시트의 위험 열이 이미 섞인 곡선이고, 등급 하나를 "
               "고르셨으면 고시표의 그 줄이 그대로 들어간다.")

    # ── 입력곡선 ──
    I = sheet(IN, widths=[12, 16, 15, 5, 12, 16, 15])
    head(I, 2, "입력 곡선", "고시된 그대로 적는다. 이 두 표가 리포트 전체의 뿌리다.",
         span=7)
    cols(I, R0IN-1, ["무위험 만기", "수익률", "연속환산", "",
                     "위험 만기", "수익률", "연속환산"],
         [12, 16, 15, 5, 12, 16, 15])
    for (lbl, pts, cmp_, mcol, ycol) in LEG:
        for i, (mt, y) in enumerate(pts):
            rr = R0IN + i
            put(I, rr, 2 if mcol == "B" else 5, mt, fmt="0.####",
                border=True, align="right", fill=YEL)
            put(I, rr, 3 if mcol == "B" else 6, y, fmt=R_P4,
                border=True, align="right", fill=YEL)
            put(I, rr, 4 if mcol == "B" else 7,
                f"={cmp_}*LN(1+{ycol}{rr}/{cmp_})", fmt=R_P4,
                border=True, align="right")
    endr = R0IN + max(len(tm.rf_curve), len(cc)) + 1
    note(I, endr, "연속환산 열은 참고다. 만기수익률을 넣었다면 실제 계산은 다음 두 "
                  "시트의 부트스트래핑에서 하고, 현물이자율을 넣었다면 이 열이 곧 "
                  "쓰이는 값이다.", span=7)

    # ── 곡선별 산출 ──
    made = {}
    for (lbl, pts, cmp_, mcol, ycol) in LEG:
        sn = f"{lbl} 산출"
        W = sheet(sn, widths=[8, 13, 15, 13, 15, 15, 15], freeze="B9")
        R0 = 9
        if spot_in:
            head(W, 2, f"{lbl} — 현물이자율 연속환산", span=6)
            put(W, 5, 2, "연속 = m · ln(1 + r ÷ m)", bold=True, color=RPT["sub"])
            put(W, 6, 2, f"m = {cmp_}  (책 3.7.4.4)", color=RPT["grey"], size=9)
            cols(W, R0-1, ["번호", "만기 t", "고시 수익률", "연속복리 현물"],
                 [8, 13, 16, 16])
            for i, (mt, y) in enumerate(pts):
                rr = R0 + i
                put(W, rr, 2, i+1, fmt=R_N0, border=True, align="center")
                put(W, rr, 3, f"='{IN}'!${mcol}${R0IN+i}", fmt="0.####",
                    border=True, align="right")
                put(W, rr, 4, f"='{IN}'!${ycol}${R0IN+i}", fmt=R_P4,
                    border=True, align="right")
                put(W, rr, 5, f"={cmp_}*LN(1+D{rr}/{cmp_})", fmt=R_P4,
                    border=True, align="right")
            # 보간에 쓸 표 — (만기, 현물) 이 C·E 열에 있다
            grid = [(mt, None) for mt, _ in pts]
            made[lbl] = dict(sh=sn, r0=R0, tcol="C", rcol="E", pts=grid)
            note(W, R0+len(pts)+1, "고시 만기 사이는 다음 시트에서 직선으로 잇는다.",
                 span=6)
        else:
            head(W, 2, f"{lbl} — 부트스트래핑", span=7)
            put(W, 5, 2, "1 = c · (DF₁ + … + DF_k) + DF_k", bold=True, color=RPT["sub"])
            put(W, 6, 2, f"c = 그 만기 수익률 ÷ {cmp_}   (연 {cmp_}회 이표 가정) · "
                         f"현물 = −ln(DF) ÷ t", color=RPT["grey"], size=9)
            cols(W, R0-1, ["k", "만기 t", "보간 수익률", "c", "누적 DF", "DF",
                           "현물 (연속)"], [8, 13, 15, 13, 15, 15, 15])
            N = max(1, int(math.ceil(T*cmp_)))
            for k in range(1, N+1):
                rr, t_ = R0 + k - 1, k/cmp_
                put(W, rr, 2, k, fmt=R_N0, border=True, align="center")
                put(W, rr, 3, round(t_, 12), fmt="0.0000", border=True, align="right")
                put(W, rr, 4, lerp_formula(t_, pts, ycol, R0IN, IN), fmt=R_P4,
                    border=True, align="right")
                put(W, rr, 5, f"=D{rr}/{cmp_}", fmt=R_N6, border=True, align="right")
                put(W, rr, 6, ("=0" if k == 1 else f"=F{rr-1}+G{rr-1}"),
                    fmt=R_N6, border=True, align="right")
                put(W, rr, 7, f"=(1-E{rr}*F{rr})/(1+E{rr})", fmt=R_N6,
                    border=True, align="right")
                put(W, rr, 8, f"=-LN(G{rr})/C{rr}", fmt=R_P4, border=True, align="right")
            # 보간에 쓸 표 — 만기는 C, 현물은 H 열
            made[lbl] = dict(sh=sn, r0=R0, tcol="C", rcol="H",
                             pts=[(k/cmp_, None) for k in range(1, N+1)])
            note(W, R0+N+1, "DF 는 앞 회차 결과를 이어 받는다. 첫 줄의 누적 DF 가 0 인 "
                            "것은 그 앞에 이표가 없기 때문이다.", span=7)

    # ── 선도이자율 ──
    def spot_ref(leg, t):
        """산출 시트의 현물 표에서 t 의 값을 뽑는 수식."""
        d = made[leg]
        pts = [(x, 0.0) for x, _ in d["pts"]]
        return lerp_formula(t, pts, d["rcol"], d["r0"], d["sh"])

    F = sheet("선도이자율", tab=RPT["green"],
              widths=[8, 13, 13, 14, 14, 14, 14, 14, 14, 14, 12],
              freeze="B10", landscape=True)
    head(F, 2, "구간 선도이자율",
         "격자 한 칸을 건너갈 때 쓰는 이자율이다. 조서 트리 시트의 11행(무위험)과 "
         "12행(위험)에 이 값이 그대로 들어간다.", span=11)
    put(F, 5, 2, "f(t₀,t₁) = [ r(t₁)·t₁ − r(t₀)·t₀ ] ÷ (t₁ − t₀)",
        bold=True, color=RPT["sub"])
    put(F, 6, 2, "q = [ exp(f_무위험 · Δt) − d ] ÷ (u − d),   u = exp(σ√Δt),  d = 1/u",
        color=RPT["grey"], size=9)
    put(F, 7, 2, "Δt", bold=True, border=True, fill=RPT["light"])
    put(F, 7, 3, round(dt_, 12), fmt="0.00000000", border=True, align="right")
    put(F, 7, 4, "σ", bold=True, border=True, fill=YEL)
    put(F, 7, 5, tm.sig, fmt=R_P2, border=True, align="right", fill=YEL)
    put(F, 7, 6, "u", bold=True, border=True, fill=RPT["light"])
    put(F, 7, 7, "=EXP($E$7*SQRT($C$7))", fmt=R_N6, border=True, align="right")
    put(F, 7, 8, "d", bold=True, border=True, fill=RPT["light"])
    put(F, 7, 9, "=1/$G$7", fmt=R_N6, border=True, align="right")
    cols(F, 9, ["스텝", "t₀", "t₁", "무위험 r(t₀)", "무위험 r(t₁)", "무위험 선도",
                "위험 r(t₀)", "위험 r(t₁)", "위험 선도", "스프레드", "q"],
         [8, 13, 13, 14, 14, 14, 14, 14, 14, 14, 12])
    for i in range(n):
        rr = 10 + i
        t0, t1 = i*dt_, (i+1)*dt_
        put(F, rr, 2, i, fmt=R_N0, border=True, align="center")
        put(F, rr, 3, round(t0, 12), fmt="0.000000", border=True, align="right")
        put(F, rr, 4, round(t1, 12), fmt="0.000000", border=True, align="right")
        put(F, rr, 5, spot_ref("무위험", t0), fmt=R_P4, border=True, align="right")
        put(F, rr, 6, spot_ref("무위험", t1), fmt=R_P4, border=True, align="right")
        put(F, rr, 7, f"=(F{rr}*D{rr}-E{rr}*C{rr})/(D{rr}-C{rr})", fmt=R_P4,
            border=True, align="right", bold=True)
        put(F, rr, 8, spot_ref("위험", t0), fmt=R_P4, border=True, align="right")
        put(F, rr, 9, spot_ref("위험", t1), fmt=R_P4, border=True, align="right")
        put(F, rr, 10, f"=(I{rr}*D{rr}-H{rr}*C{rr})/(D{rr}-C{rr})", fmt=R_P4,
            border=True, align="right", bold=True)
        put(F, rr, 11, f"=J{rr}-G{rr}", fmt=R_P4, border=True, align="right")
        put(F, rr, 12, f"=(EXP(G{rr}*$C$7)-$I$7)/($G$7-$I$7)", fmt=R_N4,
            border=True, align="right")
    note(F, 10+n+1, "스프레드가 음수인 줄이 있으면 두 곡선을 바꿔 넣은 것이다. "
                    "q 가 0 과 1 밖으로 나가면 변동성이 너무 낮거나 노드가 너무 성긴 "
                    "것이다 — 격자가 무차익 조건을 못 맞춘다.", span=11)
    return _save(wb)


# ══════════════════════════════════════════════════════════
# 5. 엑셀 조서
# ══════════════════════════════════════════════════════════
def build_xlsx(tm: Terms, full, b0, b1, b2, ca, conv, eir):
    """트리 하나에 시트 하나. 엑셀 트리모델과 같은 구조로 내보낸다."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gl

    F = "KoPub돋움체 Medium"
    NAVY, SUB, LIGHT, BAND, RFXC = "1F3864", "44618C", "DCE6F1", "F2F5F8", "FCE4D6"
    RED, GREEN, GREY, AMB = "C00000", "006100", "6B7480", "BF8F00"
    thin = Side(style="thin", color="BFC7D0")
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
    N2, N0, P2, N4, N6 = '#,##0.00', '#,##0', '0.00%', '0.0000', '0.000000'
    DATE = 'yyyy-mm-dd'
    n = tm.n; dt_ = tm.T/n; mper = n/(tm.T*12); R0 = 20
    # 계약상 개월 → 평가기준일 기준 스텝. 엔진과 같아야 한다 (경과분을 뺀다).
    stp_lo, stp_hi = step_mapper(tm, n, dt_)
    per_ = lambda mth: max(1, int(round(mth*mper)))   # 주기는 뺄 것이 없다
    RF, CR = curves(tm)
    wb = Workbook(); wb.remove(wb.active)

    def put(ws, r, c, v, *, bold=False, color="000000", fill=None, fmt=None,
            size=10, align=None, border=False):
        cl = ws.cell(row=r, column=c, value=v)
        cl.font = Font(name=F, size=size, bold=bold, color=color)
        if fill: cl.fill = PatternFill("solid", fgColor=fill)
        if fmt: cl.number_format = fmt
        cl.alignment = Alignment(horizontal=align or "general", vertical="center")
        if border: cl.border = BOX
        return cl
    def title(ws, r, t, span=8):
        put(ws, r, 2, t, bold=True, color="FFFFFF", fill=NAVY, size=13)
        for c in range(3, 2+span): ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=NAVY)
        ws.row_dimensions[r].height = 22
    def sec(ws, r, t, span=8):
        put(ws, r, 2, t, bold=True, color="FFFFFF", fill=SUB, size=10)
        for c in range(3, 2+span): ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=SUB)

    # ── 노드 값 색인 (r = 하락 횟수) ──
    idx = {}
    for k, v in full["memo"].items():
        i, j = k[0], k[1]
        if (i, j) not in idx: idx[(i, j)] = v
    node = lambda i, r: idx.get((i, i-r))
    S = full["S"]
    # 다음 조정일은 평가기준일부터 (주기 − 경과분) 뒤다. 엔진과 같은 오프셋이다.
    rfx_per = max(1, int(round(tm.rfx_cyc*mper)))
    rfx_off = (stp_lo(tm.rfx_cyc*(math.floor(tm.elapsed_m/tm.rfx_cyc) + 1))
               if tm.rfx_cyc > 0 else 1)
    is_rfx = lambda i: (tm.rfx_mode > 0 and i > 0 and i >= rfx_off
                        and (i-rfx_off) % rfx_per == 0)
    REFIXC = {i for i in range(1, n+1) if is_rfx(i)}
    cpn_amt = 100*tm.cpn*tm.ipay/12
    ey = tm.elapsed_m/12                     # 경과 연수 — 행사금액은 발행일부터 붙는다
    red = 100*(1 + accrue_rate(tm.T + ey, tm.ytm, tm.cpn, tm.ytm_cmp))
    def in_set(i, a, b, fr):
        lo, hi = stp_lo(a), stp_hi(b)
        return lo <= i <= hi and (i-lo) % per_(fr) == 0
    def put_amt(i):
        if not in_set(i, tm.p_s, tm.p_e, tm.p_f): return 0.0
        if tm.p_mode == "accrue":
            return 100*(1 + accrue_rate(i*dt_ + ey, tm.p_yield, tm.cpn, tm.p_cmp))
        return tm.p_rate
    def call_amt(i, on=True):
        if not on or not in_set(i, tm.k_s, tm.k_e, tm.k_f): return 999999
        return 100*(1 + accrue_rate(i*dt_ + ey, tm.k_prem, tm.cpn, tm.k_cmp))

    HEAD = ["Date", "time-step", "Flag(전환)", "Flag(조기상환)", "Flag(매도청구)",
            "Flag(리픽싱)", "조기상환금액", "매도청구금액", "쿠폰", "만기상환",
            "무위험 선도이자율", "위험 선도이자율", "σ", "u", "d", "q", "1−q"]

    def newsheet(name, ttl, note, refs, call_on=True):
        W = wb.create_sheet(name); W.sheet_view.showGridLines = False
        W.column_dimensions["B"].width = 17
        for i in range(n+1): W.column_dimensions[gl(3+i)].width = 9
        for r, nm in enumerate(HEAD, start=1):
            put(W, r, 2, nm, bold=True, size=8, fill=LIGHT, border=True)
        d0 = dt.date.fromisoformat(tm.d_base)
        for i in range(n+1):
            g = lambda r, v, fm=None, col="000000": put(W, r, 3+i, v, fmt=fm,
                                                        align="center", size=8, color=col)
            g(1, d0 + dt.timedelta(days=round(i*dt_*365)), DATE, GREY)
            g(2, i, N0)
            g(3, 1 if stp_lo(tm.cv_s) <= i <= stp_hi(tm.cv_e) else 0, N0)
            g(4, 1 if in_set(i, tm.p_s, tm.p_e, tm.p_f) else 0, N0)
            g(5, 1 if (call_on and in_set(i, tm.k_s, tm.k_e, tm.k_f)) else 0, N0)
            g(6, 1 if i in REFIXC else 0, N0, RED)
            g(7, round(put_amt(i), 4), N2)
            g(8, round(call_amt(i, call_on), 4), N2)
            g(9, round(cpn_amt if (tm.cpn > 0 and i > 0
                                   and i % per_(tm.ipay) == 0) else 0.0, 4), N2)
            g(10, round(red if i == n else 0.0, 4), N2)
            if i < n:
                g(11, forward_rate(RF, i*dt_, (i+1)*dt_), P2)
                g(12, forward_rate(CR, i*dt_, (i+1)*dt_), P2)
            g(13, tm.sig, P2); g(14, full["u"], N4); g(15, full["d"], N4)
            g(16, full["q"], N4); g(17, 1-full["q"], N4)
        title(W, 18, ttl, span=min(n+1, 14))
        put(W, 19, 2, "r ＼ 스텝", bold=True, size=8, fill=LIGHT, border=True, align="center")
        for i in range(n+1):
            put(W, 19, 3+i, i, bold=True, size=8, fmt=N0, align="center",
                fill=(RFXC if i in REFIXC else LIGHT), border=True)
        for r in range(n+1):
            put(W, R0+r, 2, r, bold=True, size=8, fmt=N0, align="center", fill=LIGHT, border=True)
        put(W, R0+n+2, 2, note, color=GREY, size=9)
        put(W, R0+n+3, 2, "참조: " + refs, color=GREEN, size=9)
        put(W, R0+n+4, 2, "r 은 하락 횟수. 위로 갈수록 주가가 높다.", color=GREY, size=9)
        W.freeze_panes = "C20"
        return W

    def fill_tree(W, fn, fmt=N2, txt=False):
        for i in range(n+1):
            for r in range(i+1):
                v = fn(i, r)
                if v is None: continue
                put(W, R0+r, 3+i, v, fmt=(None if txt else fmt), size=8,
                    align=("center" if txt else "right"))

    # ── 가정 ──
    A = wb.create_sheet("가정"); A.sheet_view.showGridLines = False
    for cc, w in (("B", 32), ("C", 15), ("D", 13), ("E", 13), ("F", 52)):
        A.column_dimensions[cc].width = w
    title(A, 2, "전환사채 평가 조서", span=5)
    put(A, 3, 2, "생성 " + dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        + "   ·   금액은 전자등록금액 100 기준", color=GREY, size=9)
    blocks = [("1. 모형", [("신용위험 처리", tm.model, None),
        ("조정일 아닌 시점", ["상태확장(정확)", "경로가중치", "확률가중평균", "특정노드선택"][tm.carry], None),
        ("전환권 회계 분류", "파생상품부채" if tm.conv_class == "liability" else "자본", None)]),
      ("2. 계약조건", [("발행일", tm.d_issue, None), ("평가기준일", tm.d_base, None),
        ("만기일", tm.d_mat, None), ("경과기간 (개월)", tm.elapsed_m, N2),
        ("평가기준일 주가", tm.S0, N2), ("현재 전환가액", tm.K0, N2),
        ("잔존기간 (년)", tm.T, N4), ("노드 수", tm.n, N0), ("Δt", dt_, N4),
        ("표면이자율", tm.cpn, P2), ("이자 지급주기 (개월)", tm.ipay, N0),
        ("만기보장수익률", tm.ytm, P2), ("만기상환금액", red, N2)]),
      ("3. 전환가액 조정", [("조정 방식", ["조정 없음", "하향만", "하향+상향"][tm.rfx_mode], None),
        ("조정 주기 (개월)", tm.rfx_cyc, N0), ("최저 조정가액", tm.floor, N2), ("액면가", tm.par, N2)]),
      ("4. 옵션", [("전환 시작 / 종료 (개월)", tm.cv_s, N0), ("　", tm.cv_e, N0),
        ("조기상환 시작 / 종료 / 주기", tm.p_s, N0), ("　", tm.p_e, N0), ("　 ", tm.p_f, N0),
        ("조기상환 행사금액 산정", "보장수익률 복리" if tm.p_mode == "accrue" else "고정률", None),
        ("매도청구 시작 / 종료 / 주기", tm.k_s, N0), ("　  ", tm.k_e, N0), ("　   ", tm.k_f, N0),
        ("매도청구 프리미엄", tm.k_prem, P2),
        ("매도청구 복리 횟수 (연)", tm.k_cmp, N0), ("매도청구 한도", tm.k_w, P2),
        ("의무보유 전환지연 (개월)", tm.k_lock, N0),
        ("매도청구권 평가방법", K_METHODS[tm.k_method], None),
        ("매도청구권 회계 처리",
         "별도 금융상품" if tm.k_sep else "복합내재파생에 포함", None)]),
      ("5. 시장 인풋", [("변동성 σ", tm.sig, P2)]
        + ([("조기상환권 평가", "BDT 금리격자", None),
            ("BDT 단기이자율 변동성 σ", tm.bdt_sig, P2),
            ("BDT 기준 곡선", ("위험 곡선에 직접" if tm.bdt_base == 0
                            else "무위험 + 확정 스프레드"), None)]
           if put_bdt_on(tm) else [("조기상환권 평가", "금리 고정 격자", None)])
        + [
        ("무위험 이표 (연 회)", tm.cmp_rf, N0), ("위험 이표 (연 회)", tm.cmp_cr, N0),
        ("이자율 입력", ("만기수익률 곡선" if tm.y_type == "par"
                     else f"현물이자율 곡선 (무위험 {tm.cmp_rf}회·위험 {tm.cmp_cr}회 복리)"), None),
        (f"{tm.T:.2f}년 무위험 (연속)", RF(tm.T), P2),
        (f"{tm.T:.2f}년 위험 (연속)", CR(tm.T), P2)]),
      ("6. 격자 파라미터", [("상승계수 u", full["u"], N4), ("하락계수 d", full["d"], N4),
        ("위험중립가중치 q", full["q"], N4)])]
    r = 5
    for ttl, items in blocks:
        sec(A, r, ttl, span=5); r += 1
        for nm, v, fm in items:
            put(A, r, 2, nm, border=True)
            put(A, r, 3, v, color=(RED if fm in (None, N0, N2, P2, N4) else "000000"),
                fmt=fm, align="right", border=True)
            r += 1
        r += 1
    put(A, r, 2, "빨강은 입력값입니다. 트리 시트는 계산 결과를 값으로 담았습니다.",
        color=GREY, size=9)

    # ── 트리 시트 ──
    T01 = newsheet("01 주가", "① 주가트리  S = S0 × u^(스텝−r) × d^r",
        "위로 갈수록 상승이 많은 경로다.", "가정")
    fill_tree(T01, lambda i, r: round(S(i, i-r), 2), N2)

    T02 = newsheet("02 전환가격", "② 전환가격트리  조정일이면 주가를 자르고, 아니면 이어받는다",
        "조정일 열은 주황색이다.", "01 · 가정")
    if full["exact"]:
        put(T02, R0+n+6, 2, "상태확장을 골랐으므로 한 노드에 전환가격이 여럿일 수 있어 "
            "삼각형으로 펴지지 않는다. 대표값을 표시했다.", color=RED, size=9)
    fill_tree(T02, lambda i, r: (round(node(i, r)["K"], 2) if node(i, r) else None), N2)

    T03 = newsheet("03 전환비율", "③ 전환비율트리  100 ÷ 전환가격",
        "리픽싱으로 전환가격이 내려가면 받는 주식 수가 늘어난다.", "02")
    fill_tree(T03, lambda i, r: (round(100/node(i, r)["K"], 4) if node(i, r) else None), N4)

    T04 = newsheet("04 전환가치", "④ 전환가치트리  주가 × 전환비율",
        "전환청구기간 밖이면 0이다.", "01 · 03")
    fill_tree(T04, lambda i, r: (round(node(i, r)["cv"], 2) if node(i, r) else None))

    # 앱에서 고른 모형의 트리만 만든다. 값 조서의 GS 시트는 memo 에서 값을 직접
    # 받으므로 TF 트리를 참조하지 않는다. 그래서 서로 독립적으로 넣고 뺄 수 있다.
    _tf = tm.model != "GS"
    if _tf:
        T05 = newsheet("05 지분가치", "⑤ 지분가치트리  주식으로 받게 될 부분",
            "전환하면 전환가치, 상환하면 0, 보유하면 다음 열 값을 무위험이자율로 할인한 값이다.",
            "04 · 08 · 다음 열 05")
        fill_tree(T05, lambda i, r: (round(node(i, r)["E"], 2) if node(i, r) else None))

        T06 = newsheet("06 부채가치", "⑥ 부채가치트리  현금으로 받게 될 부분",
            "전환하면 0, 상환하면 그 금액, 보유하면 다음 열 값을 위험 선도이자율로 할인한 값이다.",
            "08 · 다음 열 06")
        fill_tree(T06, lambda i, r: (round(node(i, r)["B"], 2) if node(i, r) else None))

        T07 = newsheet("07 보유가치", "⑦ 보유가치트리  지금 행사하지 않을 때의 값",
            "지분은 무위험, 부채는 위험 선도이자율로 따로 할인해 더한다. 이것이 TF 모형이다.",
            "다음 열 05 · 06")
        fill_tree(T07, lambda i, r: (round(node(i, r)["hold"], 2) if node(i, r) else None))

        T08 = newsheet("08 의사결정", "⑧ 의사결정트리  전환 · 상환P · 상환C · 보유",
            "위쪽은 전환, 아래쪽은 상환이 몰린다. 매도청구는 중간 띠에 나타난다.", "04 · 07")
        lab = {"conv": "전환", "put": "상환P", "call": "상환C", "hold": "보유", "mat": "만기상환"}
        fill_tree(T08, lambda i, r: (lab.get(node(i, r)["kind"], "") if node(i, r) else None), txt=True)

        T09 = newsheet("09 금융상품가치", "⑨ 금융상품가치트리 = 지분가치 + 부채가치",
            "네 갈래 중 최적을 고른 뒤의 값이다. 07과 비교하면 어디서 행사가 일어났는지 보인다.",
            "05 · 06")
        fill_tree(T09, lambda i, r: (round(node(i, r)["E"]+node(i, r)["B"], 2) if node(i, r) else None))
    else:
        T10 = newsheet("05 GS 전환확률", "⑤ [GS] 전환확률트리  전환 1 · 현금 0 · 보유면 다음 두 칸의 평균",
            "이 확률로 할인율을 섞는다. 자식에서 가져오므로 순환참조가 없다.", "07 · 다음 열 05")
        fill_tree(T10, lambda i, r: (round(node(i, r)["P"], 4) if node(i, r) else None), N4)

        T11 = newsheet("06 GS 할인율", "⑥ [GS] 위험조정할인율트리  y = 확률 × 무위험 + (1−확률) × 위험",
            "위쪽은 무위험에, 아래쪽은 위험이자율에 가깝다.", "05")
        def gs_rate(i, r):
            o = node(i, r)
            if not o or i >= n: return None
            fr = forward_rate(RF, i*dt_, (i+1)*dt_); fc = forward_rate(CR, i*dt_, (i+1)*dt_)
            return o["P"]*fr + (1-o["P"])*fc
        fill_tree(T11, gs_rate, P2)

        T12 = newsheet("07 GS 금융상품가치", "⑦ [GS] 금융상품가치트리",
            "네 갈래 중 최적을 고른 뒤의 값이다.", "04 · 06")
        fill_tree(T12, lambda i, r: (round(node(i, r)["V"], 2) if node(i, r) else None))

    # ── BDT (조기상환권을 금리격자로 잴 때만) ──
    if put_bdt_on(tm):
        BP, BV = bdt_grid(tm, True)
        _, BV0 = bdt_grid(tm, False)
        d0 = dt.date.fromisoformat(tm.d_base)
        for nm, ttl, note, grid, rate in (
            ("BDT 단기이자율", "BDT 단기이자율격자  r(i,j) = a · exp(2σ·j·√Δt)",
             "로그정규라 이자율이 음수가 되지 않는다. j 는 상승 횟수이고 클수록 "
             "금리가 높다 — 주가 트리와 달리 위로 갈수록 낮다. 기준금리 a 는 곡선을 "
             "정확히 되돌리도록 역산한 값이다.", None, True),
            ("BDT 부채요소", "BDT 부채요소  전환 없는 사채 + 조기상환권",
             "MAX(조기상환금액, 계속보유) 를 고른다. 계속보유는 다음 두 칸을 0.5 씩 "
             "섞어 그 칸의 단기이자율로 할인한 값이다.", BV, False),
            ("BDT 주계약", "BDT 주계약  옵션이 없는 사채",
             "조기상환권을 빼고 같은 격자로 굴린 값이다. 곡선을 정확히 되돌리므로 "
             "⑩ 주계약과 같아야 한다 — 캘리브레이션 검산이다.", BV0, False)):
            W = wb.create_sheet(nm); W.sheet_view.showGridLines = False
            W.column_dimensions["B"].width = 18
            for i in range(n+1): W.column_dimensions[gl(3+i)].width = 9
            for r, h in enumerate(["Date", "time-step", "Flag(조기상환)", "조기상환금액",
                                   "쿠폰", "만기상환", "기준금리 a", "확정 스프레드"],
                                  start=1):
                put(W, r, 2, h, bold=True, size=8, fill=LIGHT, border=True)
            for i in range(n+1):
                g = lambda r, v, fm=None: put(W, r, 3+i, v, fmt=fm,
                                              align="center", size=8)
                g(1, d0 + dt.timedelta(days=round(i*dt_*365)), DATE)
                g(2, i, N0)
                g(3, 1 if BP["in_put"](i) else 0, N0)
                g(4, round(BP["put_a"](i), 4), N2)
                g(5, round(BP["cpn"] if BP["is_pay"](i) else 0.0, 4), N2)
                g(6, round(BP["red"] if i == n else 0.0, 4), N2)
                if i < n:
                    g(7, BP["a"][i], P2); g(8, BP["add"][i], P2)
            title(W, 10, ttl, span=min(n+1, 14))
            put(W, 11, 2, note, color=GREY, size=9)
            put(W, 12, 2, "j ＼ 스텝", bold=True, size=8, fill=LIGHT,
                border=True, align="center")
            for i in range(n+1):
                put(W, 12, 3+i, i, bold=True, size=8, fmt=N0, align="center",
                    fill=LIGHT, border=True)
            for j in range(n+1):
                put(W, 13+j, 2, j, bold=True, size=8, fmt=N0, align="center",
                    fill=LIGHT, border=True)
            for i in range(n+1):
                if rate and i == n: continue          # 만기에는 다음 구간이 없다
                for j in range(i+1):
                    v = BP["r"][i][j] if rate else grid[i][j]
                    put(W, 13+j, 3+i, round(v, 8), fmt=(P2 if rate else N2),
                        size=8, align="right")
            if not rate:
                put(W, 13+n+2, 2, "t = 0", bold=True)
                put(W, 13+n+2, 3, round(grid[0][0], 6), bold=True, fmt=N2,
                    align="right")
            else:
                # 캘리브레이션 검산 — 도달가격을 더하면 시장 무이표채 가격이다.
                # 기준금리 a 를 이 조건에 맞춰 역산했으므로 0 이 나와야 한다.
                QR = 13+n+3
                sec(W, QR-1, f"캘리브레이션 검산 — 격자가 {BP['base_nm']}을 되돌리는가")
                put(W, QR, 2, "j ＼ 스텝", bold=True, size=8, fill=LIGHT,
                    border=True, align="center")
                for i in range(n+1):
                    put(W, QR, 3+i, i, bold=True, size=8, fmt=N0, align="center",
                        fill=LIGHT, border=True)
                for j in range(n+1):
                    put(W, QR+1+j, 2, j, bold=True, size=8, fmt=N0,
                        align="center", fill=LIGHT, border=True)
                for i in range(n+1):
                    for j in range(len(BP["Q"][i])):
                        put(W, QR+1+j, 3+i, round(BP["Q"][i][j], 10), fmt=N6,
                            size=8, align="right")
                for k, (nm, fn) in enumerate((
                        ("모형 무이표채  Σ Q", lambda i: sum(BP["Q"][i])),
                        ("시장 할인계수", lambda i: BP["mkt"][i]),
                        ("차이", lambda i: sum(BP["Q"][i]) - BP["mkt"][i]))):
                    r2 = QR+n+2+k
                    put(W, r2, 2, nm, bold=True, size=8, fill=LIGHT, border=True)
                    for i in range(n+1):
                        put(W, r2, 3+i, round(fn(i), 12), fmt=N6, size=8,
                            align="right", bold=(k == 2))
                put(W, QR+n+5, 2,
                    "도달가격 Q(i,j) 는 그 칸에 이르는 경로의 확률을 그 경로의 "
                    "할인율로 할인해 더한 값이다. 스텝별로 모두 더하면 그 만기의 "
                    "무이표채 가격이 되고, 그것이 시장 할인계수와 같아야 한다 — "
                    "무차익거래 조건이다. 기준금리 a 를 이 조건에 맞춰 이분법으로 "
                    "역산했으므로 차이가 0 이다.", color=GREY, size=9)
            W.freeze_panes = "C13"

    # ── 이자율곡선 ──
    C = wb.create_sheet("이자율곡선"); C.sheet_view.showGridLines = False
    for cc, w in (("B", 12), ("C", 14), ("D", 14), ("E", 14), ("F", 14), ("G", 14)):
        C.column_dimensions[cc].width = w
    title(C, 2, "기간별 이자율", span=6)
    put(C, 3, 2, "선도이자율  f(t, t+Δt) = [ r(t+Δt)×(t+Δt) − r(t)×t ] ÷ Δt", color=GREY, size=9)
    for i, h in enumerate(["시점 (년)", "무위험 현물", "무위험 선도", "위험 현물", "위험 선도", "스프레드"]):
        put(C, 5, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    for k in range(9):
        t_ = tm.T*k/8; i = min(n-1, round(t_/dt_))
        fr = forward_rate(RF, i*dt_, (i+1)*dt_); fc = forward_rate(CR, i*dt_, (i+1)*dt_)
        for j2, v in enumerate([t_, RF(t_), fr, CR(t_), fc, fc-fr]):
            put(C, 6+k, 2+j2, v, fmt=(N2 if j2 == 0 else P2), align="right", border=True)
    if tm.y_type == "par" and len(tm.cr_curve) >= 2:
        sec(C, 16, "부트스트래핑 — 위험 곡선", span=6)
        for i, h in enumerate(["만기 (년)", "만기수익률", "할인계수", "현물 (연속)"]):
            put(C, 17, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
        for k, (t_, df) in enumerate([x for x in bootstrap_df(tm.cr_curve, tm.T, tm.cmp_cr) if x[0] > 0]):
            for j2, v in enumerate([t_, _lin(tm.cr_curve, t_), df, -math.log(df)/t_]):
                put(C, 18+k, 2+j2, v,
                    fmt=(N2 if j2 == 0 else (N6 if j2 == 2 else P2)), align="right", border=True)

    # ── 결과 ──
    R = wb.create_sheet("결과"); R.sheet_view.showGridLines = False
    for cc, w in (("B", 36), ("C", 14), ("D", 14), ("E", 12), ("F", 46)):
        R.column_dimensions[cc].width = w
    title(R, 2, "평가결과", span=5)
    sec(R, 4, "1. 순차 차감", span=5)
    for i, h in enumerate(["단계", "가치", "차액", "해당 옵션"]):
        put(R, 5, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    steps = [("B0  옵션 없는 사채", b0, None, "—"),
             ("B1  조기상환권 추가", b1, b1-b0, "조기상환청구권"),
             ("B2  전환권 추가", b2, b2-b1, "전환권"),
             ("B3  매도청구권 반영", b2-ca, -ca, f"매도청구권 ({tm.k_w*100:.0f}% 한도)")]
    for i, (k, v, dv, nm) in enumerate(steps):
        last = (i == 3); fl = BAND if last else None
        put(R, 6+i, 2, k, bold=last, fill=fl, border=True)
        put(R, 6+i, 3, v, bold=last, fill=fl, fmt=N2, align="right", border=True)
        put(R, 6+i, 4, dv if dv is not None else "", bold=last, fill=fl, fmt=N2,
            align="right", border=True)
        put(R, 6+i, 5, nm, bold=last, fill=fl, border=True)
    # 앱에서 고른 모형만 싣는다.
    sec(R, 11, "2. 신용위험 처리 — " + ("TF · 값을 쪼갠다" if _tf else "GS · 할인율을 섞는다"),
        span=5)
    for i, h in enumerate(["모형", "전체", "지분", "부채", "전환확률"]):
        put(R, 12, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    if _tf:
        put(R, 13, 2, "TF · 값을 쪼갠다", border=True)
        for j2, v in enumerate([full["TF"], full["E"], full["B"]]):
            put(R, 13, 3+j2, v, fmt=N2, align="right", border=True)
        put(R, 13, 6, "", border=True)
    else:
        put(R, 13, 2, "GS · 할인율을 섞는다", border=True)
        for j2, v in enumerate([full["GS"], full["GS"]*full["P"], full["GS"]*(1-full["P"])]):
            put(R, 13, 3+j2, v, fmt=N2, align="right", border=True)
        put(R, 13, 6, full["P"], fmt=N4, align="right", border=True)
    put(R, 15, 2, "앱에서 고른 방법만 싣습니다. 다른 방법의 값은 이 조서에 없습니다.",
        color=GREY, size=9)
    sec(R, 17, "3. 검산", span=5)
    imm = 100*tm.S0/tm.K0
    al = allocate(tm, full, b0, b1, b2, ca)[0]
    ck = [("위험중립가중치 q", full["q"], 0 < full["q"] < 1),
          ("상승계수 u", full["u"], full["u"] > 1),
          ("전체 ≥ 순수사채가치", b2-full["host"], b2 >= full["host"]-1e-6),
          ("전체 ≥ 즉시 전환가치", b2-imm, not (tm.cv_s <= 0 and b2 < imm-1e-6)),
          ("배분 합계 = 100", al[-1][1], abs(al[-1][1]-100) < 0.01)]
    for i, (k, v, ok) in enumerate(ck):
        put(R, 18+i, 2, k, border=True)
        put(R, 18+i, 3, v, fmt=N4, align="right", border=True)
        put(R, 18+i, 4, "적합" if ok else "확인 필요",
            color=(GREEN if ok else RED), align="center", border=True)

    # ── 회계처리 ──
    E = wb.create_sheet("회계처리"); E.sheet_view.showGridLines = False
    for cc, w in (("B", 34), ("C", 14), ("D", 14), ("E", 18), ("F", 18), ("G", 30)):
        E.column_dimensions[cc].width = w
    title(E, 2, "회계처리", span=6)
    put(E, 3, 2, "기업회계기준서 제1032호 문단 31·32 — 부채요소를 먼저 정하고 나머지를 자본에 배분한다. "
        "매도청구권은 제3자에게 이전될 수 있어 별도의 금융상품이다 (제1109호 문단 4.3.1, "
        "회계기준원 질의회신 2022-I-KQA006, 금융위 2022.5.3 감독지침). "
        "전환권이 부채이면 전환권과 조기상환권은 상호의존적이므로 하나의 복합내재파생상품으로 "
        "전체로서 측정한다 (제1109호 문단 B4.3.4).",
        color=GREY, size=9)
    if tm.elapsed_m > 0.01:
        put(E, 4, 2, "※ 평가기준일이 발행일보다 뒤입니다. 아래 배분은 최초 인식용이므로 "
            "결산 회계처리에 그대로 쓰지 마십시오. 결산일에 쓰는 것은 파생상품 공정가치뿐이고, "
            "주계약은 발행일 배분액을 유효이자율로 상각한 장부금액입니다.", color=RED, size=9)
    sec(E, 5, "1. 최초 인식 배분", span=5)
    fac = tm.face_total
    for i, h in enumerate(["항목", "100 기준", "전액 기준 (원)"]):
        put(E, 6, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    for i, (k, v) in enumerate(al[:-1]):
        put(E, 7+i, 2, k, border=True)
        put(E, 7+i, 3, v, fmt=N2, align="right", border=True)
        put(E, 7+i, 4, v/100*fac, fmt=N0, align="right", border=True)
    rr = 7+len(al)-1
    put(E, rr, 2, "합계", bold=True, fill=BAND, border=True)
    put(E, rr, 3, al[-1][1], bold=True, fill=BAND, fmt=N2, align="right", border=True)
    put(E, rr, 4, al[-1][1]/100*fac, bold=True, fill=BAND, fmt=N0, align="right", border=True)
    put(E, rr+1, 2, f"전자등록총액 {fac:,.0f}원 기준으로 환산했습니다.", color=GREY, size=9)
    sec(E, rr+3, "2. 분개", span=5)
    for i, h in enumerate(["계정", "차변 (100)", "대변 (100)", "차변 (원)", "대변 (원)"]):
        put(E, rr+4, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    # 분개는 배분표(al)를 그대로 뒤집어 만든다. 따로 계산하면 두 표가 어긋난다.
    # 음수 항목(매도청구권 자산)만 차변으로, 나머지는 대변으로 간다.
    je = [("현금", 100.0, None)]
    for k, v in al[:-1]:
        nm = k.split(" · ")[0]
        if v < 0: je.append((f"파생상품자산 ({nm})", -v, None))
        else:     je.append((f"　{nm}", None, v))
    for i, (k, dr, cr) in enumerate(je):
        put(E, rr+5+i, 2, k, size=9, border=True)
        put(E, rr+5+i, 3, dr if dr is not None else "", fmt=N2, align="right", border=True)
        put(E, rr+5+i, 4, cr if cr is not None else "", fmt=N2, align="right", border=True)
        put(E, rr+5+i, 5, dr/100*fac if dr is not None else "", fmt=N0, align="right", border=True)
        put(E, rr+5+i, 6, cr/100*fac if cr is not None else "", fmt=N0, align="right", border=True)
    tr = rr+5+len(je)
    sd = sum(x for _, x, _ in je if x); sc = sum(x for _, _, x in je if x)
    put(E, tr, 2, "합계", bold=True, fill=BAND, border=True)
    for j2, v2 in enumerate([sd, sc, sd/100*fac, sc/100*fac]):
        put(E, tr, 3+j2, v2, bold=True, fill=BAND, fmt=(N2 if j2 < 2 else N0),
            align="right", border=True)
    put(E, tr+2, 2, "최초 인식에는 어떠한 손익도 생기지 않는다. 차변과 대변 합계가 일치해야 한다.",
        color=GREY, size=9)
    put(E, tr+3, 2, "전환권 분류: " + ("파생상품부채 — 주계약을 잔여로"
        if tm.conv_class == "liability" else "자본 — 전환권대가를 잔여로"), color=GREY, size=9)

    # ── 상각표 ──
    r_eir, rows_eir, redm, nper = eir
    M = wb.create_sheet("상각표"); M.sheet_view.showGridLines = False
    for cc, w in (("B", 10), ("C", 13), ("D", 12), ("E", 16), ("F", 14),
                  ("G", 14), ("H", 16)):
        M.column_dimensions[cc].width = w
    title(M, 2, "주계약 상각표", span=7)
    put(M, 3, 2, "지급일은 계약상 일정이므로 발행일부터 센다. 회차 수는 노드가 아니라 "
        "이자 지급주기를 따른다.", color=GREY, size=9)
    sec(M, 4, "유효이자율 역산", span=7)
    for i, (k, v, fm) in enumerate([("주계약 (인식액)", rows_eir[0][2] if rows_eir else b0, N2),
                                    ("만기상환금액", redm, N2),
                                    ("표면이자 (회당)", cpn_amt, N2), ("상각 횟수", nper, N0)]):
        put(M, 5+i, 2, k, border=True); put(M, 5+i, 3, v, fmt=fm, align="right", border=True)
    put(M, 9, 2, "유효이자율 (연, 이산복리)", bold=True, fill=BAND, border=True)
    put(M, 9, 3, r_eir, bold=True, fill=BAND, fmt=P2, align="right", border=True)
    sec(M, 11, "상각 내역", span=7)
    for i, h in enumerate(["회차", "지급일", "경과연수", "기초", "이자비용",
                           "지급이자", "기말"]):
        put(M, 12, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    _di = dt.date.fromisoformat(tm.d_issue); _dm = dt.date.fromisoformat(tm.d_mat)
    for i, row in enumerate(rows_eir):
        last = (i == len(rows_eir)-1); fl = BAND if last else None
        # 마지막은 만기일, 나머지는 발행일 + 회차 × 지급주기다.
        pd_ = _dm if last else _add_months(_di, int(round(pay_index(tm, row[1])*tm.ipay)))
        put(M, 13+i, 2, row[0], bold=last, fill=fl, fmt=N0, align="right", border=True)
        put(M, 13+i, 3, pd_, bold=last, fill=fl, fmt=DATE, align="right", border=True)
        for j2, v in enumerate(row[1:], start=4):
            put(M, 13+i, j2, v, bold=last, fill=fl, fmt=N2, align="right", border=True)

    # ── 분리 판단 ──
    # 화면과 같은 함수가 만든 문안이라 둘이 어긋날 수 없다.
    SP = split_test(tm, full, b0, b1, b2, ca, eir[1])
    J = wb.create_sheet("분리 판단"); J.sheet_view.showGridLines = False
    J.column_dimensions["B"].width = 24; J.column_dimensions["C"].width = 92
    title(J, 2, "내재파생상품 분리 판단", span=2)
    put(J, 3, 2, "판단 순서 — 기업회계기준서 제1109호 문단 B4.3.5 말미는 제1032호에 "
                 "따라 전환채무상품의 자본요소를 분리하기 전에 내재된 콜옵션이나 "
                 "풋옵션이 주채무계약과 밀접하게 관련되어 있는지를 판단하라고 정한다.",
        color=GREY, size=9)
    _r = 5
    for _k, _nm in (("put", "조기상환청구권"), ("call", "매도청구권")):
        _d = SP[_k]
        sec(J, _r, _nm, span=2); _r += 1
        put(J, _r, 2, "결론", bold=True, border=True)
        put(J, _r, 3, _d["결론"], bold=True, border=True); _r += 1
        for _i, _x in enumerate(_d["이유"]):
            put(J, _r, 2, "판단 근거" if _i == 0 else "", border=True)
            put(J, _r, 3, _x, border=True); _r += 1
        put(J, _r, 2, "기준서", border=True)
        put(J, _r, 3, " · ".join(_d["근거"]) or "—", border=True); _r += 1
        put(J, _r, 2, "평가방법", border=True)
        put(J, _r, 3, _d["평가"].replace("**", ""), border=True); _r += 1
        for _a, _v in _d["지표"].items():
            put(J, _r, 2, _a, border=True)
            put(J, _r, 3, (f"{_v*100:.1f}%" if _a == "차이" else
                           ("예" if _v is True else "아니오" if _v is False
                            else f"{_v:,.4f}")), border=True); _r += 1
        _r += 1
    put(J, _r, 2, "이 시트는 앱의 「분리 판단」 화면과 같은 함수가 만든다. 계약 조항 "
                  "확인 항목을 바꾸면 결론과 문안이 함께 바뀐다.", color=GREY, size=9)
    for _row in J.iter_rows(min_row=5, max_row=_r, min_col=3, max_col=3):
        for _c in _row: _c.alignment = Alignment(wrap_text=True, vertical="top")

    # ── 해설 ──
    H = wb.create_sheet("해설", 0); H.sheet_view.showGridLines = False
    H.column_dimensions["B"].width = 22; H.column_dimensions["C"].width = 96
    title(H, 2, "이 조서를 읽는 법", span=2)
    ex = [("시트 순서", ""),
      ("구조", "트리 하나가 시트 하나다. 가정 → 01 주가 → … → 12 GS → 결과 → 회계처리."),
      ("따라가기", "시트 탭을 왼쪽부터 차례로 누르면 계산이 쌓이는 순서 그대로다."),
      ("", ""),
      ("머리 17행은 모두 같다", ""),
      ("1행 Date", "평가기준일 + 스텝 × Δt. 계약상 행사일과 대조해 보는 자리다."),
      ("1~2행", "날짜와 스텝 번호"),
      ("3~6행", "Flag — 전환 · 조기상환 · 매도청구 · 리픽싱이 가능한 열에 1이 뜬다"),
      ("7~10행", "조기상환금액 · 매도청구금액 · 쿠폰 · 만기상환"),
      ("11~12행", "무위험 선도이자율과 위험 선도이자율"),
      ("13~17행", "σ · u · d · q · 1−q"),
      ("", ""),
      ("행은 하락 횟수다", ""),
      ("r = 0", "한 번도 안 내린 경로. 맨 위이고 주가가 가장 높다."),
      ("r = 스텝", "계속 내린 경로. 맨 아래다."),
      ("빈칸", "그 시점에 존재하지 않는 노드다."),
      ("", ""),
      ("계산 순서", ""),
      ("만기부터", "만기에는 미래가 없어 전환가치·조기상환·만기상환만 비교하면 끝난다."),
      ("한 칸씩 왼쪽", "07 보유가치가 다음 열의 05·06을 가져와 할인한다."),
      ("그다음", "09 금융상품가치가 최적을 고르고, 08 의사결정이 이름을 붙이고, 05·06이 확정된다."),
      ("순환이 아닌 이유", "05·06은 같은 열의 08을 보지만, 07은 다음 열의 05·06을 본다."),
      ("", "오른쪽 열이 먼저 확정되고 왼쪽으로 오므로 고리가 닫히지 않는다."),
      ("", ""),
      ("TF와 GS", ""),
      ("TF", "05~09. 값을 지분과 부채로 쪼개 각각 다른 이자율로 할인한다."),
      ("GS", "10~12. 값은 하나로 두고 전환확률로 할인율을 섞는다."),
      ("비교", "09와 12의 같은 칸을 비교하면 두 모형의 차이가 그 노드에서 얼마인지 보인다."),
      ("", ""),
      ("이 파일의 성격", ""),
      ("값 조서", "평가앱이 계산한 결과를 값으로 담았다. 수식이 아니므로 셀을 바꿔도 다시 계산되지 않는다."),
      ("재계산", "인풋을 바꾸려면 앱에서 다시 계산한 뒤 조서를 새로 내려받으면 된다.")]
    r = 4
    for a2, b3 in ex:
        if a2 and not b3: sec(H, r, a2, span=2)
        elif a2:
            put(H, r, 2, a2, bold=True, size=9); put(H, r, 3, b3, size=9)
        r += 1
    for i in range(4, r):
        H.cell(row=i, column=3).alignment = Alignment(horizontal="left", vertical="center")

    polish_wb(wb)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()


def build_xlsx_formula(tm: Terms, full, b0, b1, b2, ca, conv, eir):
    """수식 조서 — 트리를 살아 있는 수식으로 내보낸다.
    가정 시트의 노란 셀을 바꾸면 엑셀 안에서 다시 계산된다.
    재결합 격자가 필요하므로 근사 방법에서만 만들 수 있다."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gl

    F = "KoPub돋움체 Medium"
    NAVY, SUB, LIGHT, BAND, RFXC = "1F3864", "44618C", "DCE6F1", "F2F5F8", "FCE4D6"
    RED, GREEN, GREY, AMB = "C00000", "006100", "6B7480", "BF8F00"
    thin = Side(style="thin", color="BFC7D0")
    BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
    N2, N0, P2, N4, N6 = '#,##0.00', '#,##0', '0.00%', '0.0000', '0.000000'
    DATE = 'yyyy-mm-dd'
    n = tm.n; dt_ = tm.T/n; mper = n/(tm.T*12); R0 = 20
    el = tm.elapsed_m
    # 트랜치 이름은 계약의 매도청구 한도에서 나온다. 30/70 으로 굳혀 두면
    # 한도가 다른 사채에서 시트 이름이 계약과 어긋난다.
    KW = f"{tm.k_w*100:,.0f}%"
    KW0 = f"{(1-tm.k_w)*100:,.0f}%"
    stp_lo, stp_hi = step_mapper(tm, n, dt_)
    RF, CR = curves(tm)
    # 다음 조정일은 평가기준일부터 (주기 − 경과분) 뒤다. 엔진과 같은 오프셋이다.
    rfx_per = max(1, int(round(tm.rfx_cyc*mper)))
    rfx_off = (stp_lo(tm.rfx_cyc*(math.floor(tm.elapsed_m/tm.rfx_cyc) + 1))
               if tm.rfx_cyc > 0 else 1)
    is_rfx = lambda i: (tm.rfx_mode > 0 and i > 0 and i >= rfx_off
                        and (i-rfx_off) % rfx_per == 0)
    REFIXSET = {i for i in range(1, n+1) if is_rfx(i)}
    wb = Workbook(); wb.remove(wb.active)

    def put(ws, r, c, v, *, bold=False, color="000000", fill=None, fmt=None,
            size=10, align=None, border=False):
        cl = ws.cell(row=r, column=c, value=v)
        cl.font = Font(name=F, size=size, bold=bold, color=color)
        if fill: cl.fill = PatternFill("solid", fgColor=fill)
        if fmt: cl.number_format = fmt
        cl.alignment = Alignment(horizontal=align or "general", vertical="center")
        if border: cl.border = BOX
        return cl

    def title(ws, r, t, span=8):
        put(ws, r, 2, t, bold=True, color="FFFFFF", fill=NAVY, size=13)
        for c in range(3, 2+span):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=NAVY)
        ws.row_dimensions[r].height = 22

    def sec(ws, r, t, span=8):
        put(ws, r, 2, t, bold=True, color="FFFFFF", fill=SUB, size=10)
        for c in range(3, 2+span):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=SUB)

    # ── 가정 ─────────────────────────────────────────────
    A = wb.create_sheet("가정"); A.sheet_view.showGridLines = False
    for cc, w in (("B", 30), ("C", 16), ("D", 46)): A.column_dimensions[cc].width = w
    title(A, 1, "전환사채 평가 조서 — 수식 포함", span=3)
    put(A, 2, 2, "노란 셀을 바꾸면 모든 트리 시트가 다시 계산된다.", color=GREY, size=9)
    spec = [
        # 진짜 날짜로 넣는다. 각 트리 머리 1행이 이 셀로 스텝 날짜를 계산한다.
        ("발행일", "d_issue", dt.date.fromisoformat(tm.d_issue), DATE, True),
        ("평가기준일", "d_base", dt.date.fromisoformat(tm.d_base), DATE, True),
        ("만기일", "d_mat", dt.date.fromisoformat(tm.d_mat), DATE, True),
        ("경과기간 (개월)", "elm", tm.elapsed_m, N2, True),
        ("평가기준일 주가", "S0", tm.S0, N2, True),
        ("현재 전환가액", "K0", tm.K0, N2, True),
        ("잔존기간 T (년)", "T", tm.T, N4, True),
        ("노드 수 n", "n", n, N0, True),
        ("Δt", "dt", "@=C{T}/C{n}", N4, False),
        ("표면이자율", "cpn", tm.cpn, P2, True),
        ("이자 지급주기 (스텝)", "ipay", max(1, int(round(tm.ipay*mper))), N0, True),
        ("이자 지급주기 (개월)", "ipaym", tm.ipay, N2, True),
        ("만기보장수익률", "ytm", tm.ytm, P2, True),
        ("만기보장 복리 횟수", "ycm", tm.ytm_cmp, N0, True),
        ("만기상환금액", "red",
         # 할증금은 음수가 될 수 없다. 엔진의 accrue_rate 와 같이 0 에서 끊는다.
         "@=IF(C{ytm}<=0,100*(1+MAX(0,(C{ytm}-C{cpn})*(C{T}+C{elm}/12))),"
         "100*(1+MAX(0,(C{ytm}-C{cpn})/C{ytm}*"
         "((1+C{ytm}/C{ycm})^(C{ycm}*(C{T}+C{elm}/12))-1))))", N2, False),
        ("최저 조정가액", "flr", tm.floor, N2, True),
        ("액면가", "par", tm.par, N2, True),
        ("리픽싱 상한", "cap", "@=C{K0}", N2, False),
        ("리픽싱 주기 (스텝)", "cyc", max(1, int(round(tm.rfx_cyc*mper))), N0, True),
        ("첫 조정 스텝", "roff", rfx_off, N0, True),
        ("전환 시작 (스텝)", "cvs", stp_lo(tm.cv_s), N0, True),
        ("전환 종료 (스텝)", "cve", stp_hi(tm.cv_e), N0, True),
        ("조기상환 시작 (스텝)", "pst", stp_lo(tm.p_s), N0, True),
        ("조기상환 종료 (스텝)", "pen", stp_hi(tm.p_e), N0, True),
        ("조기상환 주기 (스텝)", "frq", max(1, int(round(tm.p_f*mper))), N0, True),
        ("조기상환 행사금액", "prate", tm.p_rate, N2, True),
        ("조기상환 보장수익률", "pyld", tm.p_yield, P2, True),
        ("보장 복리 (연 회)", "pcmp", tm.p_cmp, N0, True),
        ("매도청구 시작 (스텝)", "kst", stp_lo(tm.k_s), N0, True),
        ("매도청구 종료 (스텝)", "ken", stp_hi(tm.k_e), N0, True),
        ("매도청구 주기 (스텝)", "kfrq", max(1, int(round(tm.k_f*mper))), N0, True),
        ("매도청구 프리미엄", "prem", tm.k_prem, P2, True),
        ("매도청구 복리 횟수 (연)", "kcmp", tm.k_cmp, N0, True),
        ("매도청구 한도", "cw", tm.k_w, P2, True),
        (f"{KW} 전환 시작 (스텝)", "cv30", stp_lo(max(tm.cv_s, tm.k_lock)), N0, True),
        ("변동성 σ", "sig", tm.sig, P2, True),
        ("상승계수 u", "u", "@=EXP(C{sig}*SQRT(C{dt}))", N4, False),
        ("하락계수 d", "dd", "@=1/C{u}", N4, False),
        ("위험중립가중치 q", "q", "@=(EXP(C{rfc}*C{dt})-C{dd})/(C{u}-C{dd})", N4, False),
        ("1 − q", "q1", "@=1-C{q}", N4, False),
        ("리픽싱 반영 (1/0)", "rfx", 1 if tm.rfx_mode > 0 else 0, N0, True),
        ("상향 조정 (1/0)", "up", 1 if tm.rfx_mode == 2 else 0, N0, True),
        ("조정일 처리 (1/2/3)", "mth", max(1, tm.carry), N0, True),
        ("전환권 분류 (1 자본 / 0 부채)", "eqcls", 1 if tm.conv_class == "equity" else 0, N0, True),
        ("매도청구권 평가방법 (0 유무가치 / 1 혼합할인율 / 2 지분·부채 분리)",
         "kmeth", tm.k_method, N0, True),
        ("매도청구권 처리 (1 별도 금융상품 / 0 내재파생 포함)", "ksep", tm.k_sep, N0, True),
        ("신용위험 처리 (0 TF / 1 GS)", "mdl", 1 if tm.model == "GS" else 0, N0, True),
        ("조기상환권 (0 금리고정 / 1 BDT)", "pbdt", 1 if put_bdt_on(tm) else 0, N0, True),
        ("BDT 변동성 σ", "bsig", tm.bdt_sig, P2, True),
        ("BDT 기준 (0 위험곡선 / 1 무위험+스프레드)", "bbase", tm.bdt_base, N0, True),
        ("전자등록총액 (원)", "face", tm.face_total, N0, True),
        ("무위험 (연속, 평탄)", "rfc", RF(tm.T), P2, False)]
    ROWN = {key: 3+i for i, (_, key, _, _, _) in enumerate(spec)}
    K = {key: f"가정!$C${r}" for key, r in ROWN.items()}
    for i, (nm, key, v, fm, inp) in enumerate(spec):
        r = 3+i
        put(A, r, 2, nm, border=True)
        val = v.lstrip("@").format(**ROWN) if (isinstance(v, str) and v.startswith("@")) else v
        put(A, r, 3, val, color=(RED if inp else "000000"),
            fill=("FDF6DD" if inp else None), fmt=fm, align="right", border=True)
    nb = 3+len(spec)+1
    put(A, nb, 2, "노란 셀이 입력값이다. 스텝은 평가기준일 기준이며, "
        "행사금액은 발행일부터 붙는다.", color=GREY, size=9)
    put(A, nb+1, 2, "선도이자율은 부트스트래핑 결과라 각 트리 시트 11·12행에 값으로 들어 있다.",
        color=AMB, size=9)

    HEAD = ["Date", "time-step", "Flag(전환)", "Flag(조기상환)", "Flag(매도청구)",
            "Flag(리픽싱)", "조기상환금액", "매도청구금액", "쿠폰", "만기상환",
            "무위험 선도이자율", "위험 선도이자율", "σ", "u", "d", "q", "1−q"]
    ey = el/12

    def newsheet(name, ttl, note, refs, call_on=True, conv_cell=None):
        W = wb.create_sheet(name); W.sheet_view.showGridLines = False
        W.column_dimensions["B"].width = 17
        for i in range(n+1): W.column_dimensions[gl(3+i)].width = 9
        cvs = conv_cell or K["cvs"]
        for r, nm in enumerate(HEAD, start=1):
            put(W, r, 2, nm, bold=True, size=8, fill=LIGHT, border=True)
        for i in range(n+1):
            L = gl(3+i); Lp = gl(2+i) if i > 0 else None
            g = lambda r, v, fm=None, col="000000": put(W, r, 3+i, v, fmt=fm,
                                                        align="center", size=8, color=col)
            # 머리는 모두 2행(스텝)을 참조한다. 2행 자신도 직전 열 + 1 이라,
            # 맨 앞 열의 0 하나에서 모든 열이 줄줄이 정해진다.
            st = f"{L}$2"                            # 이 열의 스텝
            yr = f"({st}*{K['dt']}+{K['elm']}/12)"   # 발행일부터 흐른 연수
            g(1, f"={K['d_base']}+{st}*{K['dt']}*365", DATE, GREY)
            g(2, (0 if i == 0 else f"={Lp}$2+1"), N0)
            g(3, f"=IF(AND({st}>={cvs},{st}<={K['cve']}),1,0)", N0)
            g(4, f"=IF(AND({st}>={K['pst']},{st}<={K['pen']},"
                 f"MOD({st}-{K['pst']},{K['frq']})=0),1,0)", N0)
            g(5, (f"=IF(AND({st}>={K['kst']},{st}<={K['ken']},"
                  f"MOD({st}-{K['kst']},{K['kfrq']})=0),1,0)" if call_on else 0), N0)
            g(6, f"=IF(AND({st}>0,{st}>={K['roff']},"
                 f"MOD({st}-{K['roff']},{K['cyc']})=0),1,0)", N0, RED)
            # 상환할증금 = (g−c)/g × ((1+g/m)^(m·t) − 1).  g 가 0 이면 (g−c)·t
            g(7, f"=IF({L}$4=1,IF({K['pyld']}>0,"
                 f"100*(1+MAX(0,({K['pyld']}-{K['cpn']})/{K['pyld']}*"
                 f"((1+{K['pyld']}/{K['pcmp']})^({K['pcmp']}*{yr})-1))),"
                 f"{K['prate']}),0)", N2)
            g(8, f"=IF({L}$5=1,IF({K['prem']}>0,"
                 f"100*(1+MAX(0,({K['prem']}-{K['cpn']})/{K['prem']}*"
                 f"((1+{K['prem']}/{K['kcmp']})^({K['kcmp']}*{yr})-1))),"
                 f"100*(1+MAX(0,-{K['cpn']}*{yr}))),999999)", N2)
            g(9, f"=IF(AND({st}>0,MOD({st},{K['ipay']})=0),"
                 f"100*{K['cpn']}*{K['ipaym']}/12,0)", N2)
            g(10, f"=IF({st}={K['n']},{K['red']},0)", N2)
            if i < n:
                g(11, forward_rate(RF, i*dt_, (i+1)*dt_), P2, AMB)
                g(12, forward_rate(CR, i*dt_, (i+1)*dt_), P2, AMB)
                g(16, f"=(EXP({L}$11*{K['dt']})-{L}$15)/({L}$14-{L}$15)", N4)
                g(17, f"=1-{L}$16", N4)
            else:
                g(16, f"={K['q']}", N4); g(17, f"={K['q1']}", N4)
            g(13, f"={K['sig']}", P2); g(14, f"={K['u']}", N4); g(15, f"={K['dd']}", N4)
        title(W, 18, ttl, span=min(n+1, 14))
        put(W, 19, 2, "r ＼ 스텝", bold=True, size=8, fill=LIGHT, border=True, align="center")
        for i in range(n+1):
            put(W, 19, 3+i, i, bold=True, size=8, fmt=N0, align="center",
                fill=(RFXC if i in REFIXSET else LIGHT), border=True)
        for r in range(n+1):
            put(W, R0+r, 2, r, bold=True, size=8, fmt=N0, align="center",
                fill=LIGHT, border=True)
        put(W, R0+n+2, 2, note, color=GREY, size=9)
        put(W, R0+n+3, 2, "참조: " + refs, color=GREEN, size=9)
        put(W, R0+n+4, 2, "r 은 하락 횟수. 11·12행 선도이자율만 값이다.", color=AMB, size=9)
        W.freeze_panes = "C20"
        return W

    def fill(W, fn, fmt=N2, txt=False):
        for i in range(n+1):
            L = gl(3+i); Lp = gl(2+i) if i > 0 else None; Ln = gl(4+i) if i < n else None
            for r in range(i+1):
                v = fn(i, r, L, Lp, Ln)
                if v is None: continue
                put(W, R0+r, 3+i, v, fmt=(None if txt else fmt), size=8,
                    align=("center" if txt else "right"))

    Q = lambda nm: f"'{nm}'"
    TOLX = repr(TOL)                # 엔진과 같은 동점 허용오차를 수식에도 쓴다
    S1, S2, S3, S4 = "01 주가", "02 전환가격", "03 전환비율", "04 전환가치"
    S5, S6, S7 = "05 지분가치", "06 부채가치", "07 보유가치"
    S8, S9, S10 = "08 금융상품가치", "09 의사결정", "10 주계약가치"
    S11, S12, S13, S14 = "11 GS 전환확률", "12 GS 할인율", "13 GS 보유가치", "14 GS 금융상품가치"
    S15, S16 = f"15 {KW} 트랜치", "16 부채요소"
    S17, S18 = "17 구성비율", "18 혼합할인율"
    S19, S20 = "19 콜 페이오프", "20 매도청구권가치"
    S21, S22 = "21 방법2 지분보유", "22 방법2 부채보유"
    S23, S24 = "23 방법2 지분몫", "24 방법2 부채몫"

    # ── 01 주가 ──
    W = newsheet(S1, "① 주가트리",
                 "맨 위는 직전 열 맨 위 × u, 나머지는 직전 열 한 칸 위 × d.", "가정")
    fill(W, lambda i, r, L, Lp, Ln: (f"={K['S0']}" if i == 0 else
         (f"={Lp}{R0}*{L}$14" if r == 0 else f"={Lp}{R0+r-1}*{L}$15")), N2)

    # ── 도달확률 ──
    W2 = wb.create_sheet("도달확률"); W2.sheet_view.showGridLines = False
    W2.column_dimensions["B"].width = 12
    title(W2, 2, "도달확률  P = COMBIN(스텝, r) × q^(스텝−r) × (1−q)^r", span=min(n+1, 12))
    put(W2, 3, 2, "q 는 그 열의 위험중립가중치다 (① 16행). 이월 계산에서 직전 열을 "
        "참조하므로 열마다 제 q 를 써야 엔진과 맞는다. "
        "스텝(4행)과 하락 횟수(B열)를 참조하므로 머리를 고치면 표 전체가 따라 움직인다.",
        color=GREY, size=9)
    put(W2, 4, 2, "r ＼ 스텝", bold=True, size=8, fill=LIGHT, border=True, align="center")
    for i in range(n+1):
        L = gl(3+i)
        W2.column_dimensions[L].width = 9
        # 스텝도 직전 열 + 1 이다. 맨 앞의 0 하나가 전체를 정한다.
        put(W2, 4, 3+i, (0 if i == 0 else f"={gl(2+i)}$4+1"),
            bold=True, size=8, fmt=N0, align="center", fill=LIGHT, border=True)
    for r in range(n+1):
        put(W2, 5+r, 2, (0 if r == 0 else f"=$B{4+r}+1"),
            bold=True, size=8, fmt=N0, align="center", fill=LIGHT, border=True)
        for i in range(r, n+1):
            L = gl(3+i)
            qq = f"{Q(S1)}!{L}$16"
            # 스텝은 이 열의 4행, 하락 횟수는 이 행의 B열에서 읽는다.
            put(W2, 5+r, 3+i,
                f"=COMBIN({L}$4,$B{5+r})*{qq}^({L}$4-$B{5+r})*(1-{qq})^$B{5+r}",
                fmt='0.000000', size=8, align="right")
    W2.freeze_panes = "C5"

    # ── 02 전환가격 ──
    W = newsheet(S2, "② 전환가격트리",
                 "조정일 열은 주황색이다. 처리 방법은 가정에서 고른다.", f"{S1} · 도달확률")
    def kf(i, r, L, Lp, Ln):
        if i == 0: return f"={K['K0']}"
        base = (f"IF({K['up']}=1,{Q(S1)}!{L}{R0+r},"
                f"MIN({Lp}{R0+max(r-1,0)},{Q(S1)}!{L}{R0+r}))")
        clip = f"MIN(MAX({base},{K['flr']},{K['par']}),{K['cap']})"
        up = f"{Lp}{R0+r}" if r <= i-1 else None
        dn = f"{Lp}{R0+r-1}" if r-1 >= 0 else None
        uP = f"도달확률!{Lp}{5+r}" if up else None
        dP = f"도달확률!{Lp}{5+r-1}" if dn else None
        if up is None: carry = dn
        elif dn is None: carry = up
        else:
            # q 는 직전 구간의 것이다. 엔진도 qi(i-1) 을 쓴다.
            m1 = (f"({up}*{uP}*{Lp}$16+{dn}*{dP}*{Lp}$17)"
                  f"/({uP}*{Lp}$16+{dP}*{Lp}$17)")
            m2 = f"({up}*{Lp}$16+{dn}*{Lp}$17)"
            carry = f"IF({K['mth']}=1,{m1},IF({K['mth']}=2,{m2},{dn}))"
        return f"=IF({K['rfx']}=0,{K['K0']},IF({L}$6=1,{clip},{carry}))"
    fill(W, kf, N2)

    W = newsheet(S3, "③ 전환비율트리  100 ÷ 전환가격", "받게 될 주식 수다.", S2)
    fill(W, lambda i, r, L, Lp, Ln: f"=100/{Q(S2)}!{L}{R0+r}", N4)

    W = newsheet(S4, "④ 전환가치트리  주가 × 전환비율",
                 "전환청구기간 밖이면 0이다.", f"{S1} · {S3}")
    fill(W, lambda i, r, L, Lp, Ln:
         f"=IF({L}$3=1,{Q(S1)}!{L}{R0+r}*{Q(S3)}!{L}{R0+r},0)")

    # 앱에서 고른 것만 만든다. 쓰이지 않는 트리는 아예 넣지 않는다.
    #   GS 시트     — 신용위험 처리가 GS 일 때
    #   ⑮ 트랜치    — 매도청구권을 유무가치비교법으로 잴 때
    #   ⑰~⑳       — 옵션차익 · 혼합할인율
    #   ⑰⑲㉑~㉔   — 옵션차익 · 지분·부채 분리
    _gs = tm.model == "GS"
    _hascall = tm.k_w > 0
    _km = tm.k_method
    _need15 = _hascall and _km == 0
    _need1 = _hascall and _km == 1
    _need2 = _hascall and _km == 2

    # ⑤~⑨ 는 TF 트리다. GS 를 골랐어도 옵션차익법(방법 1·2)의 기초자산이라
    # 그때는 남긴다. GS + 유무가치비교법이면 쓰이지 않으므로 만들지 않는다.
    _needTF = (not _gs) or _need1 or _need2
    _bdt = put_bdt_on(tm)      # 조기상환권을 BDT 로 재는가
    if _needTF:
        W = newsheet(S5, "⑤ 지분가치트리",
                     "전환이면 전환가치, 상환이면 0, 보유면 다음 열을 무위험이자율로 할인.",
                     f"{S4} · {S9} · 다음 열 {S5}", call_on=False)
        fill(W, lambda i, r, L, Lp, Ln: (
            f'=IF({Q(S9)}!{L}{R0+r}="전환",{Q(S4)}!{L}{R0+r},0)' if i == n else
            f'=IF({Q(S9)}!{L}{R0+r}="전환",{Q(S4)}!{L}{R0+r},'
            f'IF(OR({Q(S9)}!{L}{R0+r}="상환P",{Q(S9)}!{L}{R0+r}="상환C"),0,'
            f'({Ln}{R0+r}*{L}$16+{Ln}{R0+r+1}*{L}$17)*EXP(-{L}$11*{K["dt"]})))'))

        W = newsheet(S6, "⑥ 부채가치트리",
                     "전환이면 0, 상환이면 그 금액, 보유면 다음 열을 위험 선도이자율로 할인.",
                     f"{S9} · 다음 열 {S6}", call_on=False)
        fill(W, lambda i, r, L, Lp, Ln: (
            f'=IF({Q(S9)}!{L}{R0+r}="전환",0,MAX({L}$7,{L}$10)+{L}$9)' if i == n else
            f'=IF({Q(S9)}!{L}{R0+r}="상환P",{L}$7,IF({Q(S9)}!{L}{R0+r}="상환C",{L}$8,'
            f'IF({Q(S9)}!{L}{R0+r}="전환",0,'
            f'({Ln}{R0+r}*{L}$16+{Ln}{R0+r+1}*{L}$17)*EXP(-{L}$12*{K["dt"]})+{L}$9)))'))

        W = newsheet(S7, "⑦ 보유가치트리",
                     "지분은 무위험, 부채는 위험 선도이자율로 따로 할인해 더한다. 이것이 TF다.",
                     f"다음 열 {S5} · {S6}", call_on=False)
        fill(W, lambda i, r, L, Lp, Ln: (f"={L}$10+{L}$9" if i == n else
             f"=({Q(S5)}!{Ln}{R0+r}*{L}$16+{Q(S5)}!{Ln}{R0+r+1}*{L}$17)"
             f"*EXP(-{L}$11*{K['dt']})"
             f"+({Q(S6)}!{Ln}{R0+r}*{L}$16+{Q(S6)}!{Ln}{R0+r+1}*{L}$17)"
             f"*EXP(-{L}$12*{K['dt']})+{L}$9"))

        W = newsheet(S8, "⑧ 금융상품가치트리",
                     f"{KW0} 트랜치 — 매도청구권이 걸리지 않는다. 콜은 ⑮에서만 반영한다.",
                     f"{S4} · {S7}", call_on=False)
        fill(W, lambda i, r, L, Lp, Ln:
             f"=MAX({Q(S7)}!{L}{R0+r},{Q(S4)}!{L}{R0+r},{L}$7"
             + (f"+{L}$9)" if i == n else ")"))

        W = newsheet(S9, "⑨ 의사결정트리",
                     f"전환가치·조기상환금액·보유가치를 직접 견준다. {KW0} 트랜치라 "
                     "상환C 는 없다. 리픽싱 조정일에는 전환가치가 정확히 100 이 되어 상환금액과 동점이 되므로 "
                     "전환은 허용오차만큼 앞설 때만 이긴다. 동점이면 현금이다.",
                     f"{S4} · {S7}", call_on=False)
        fill(W, lambda i, r, L, Lp, Ln:
             f'=IF({Q(S4)}!{L}{R0+r}>=MAX({L}$7,{Q(S7)}!{L}{R0+r})+{TOLX},"전환",'
             f'IF({L}$7>={Q(S7)}!{L}{R0+r}-{TOLX},"상환P","보유"))', txt=True)

    W = newsheet(S10, "⑩ 주계약가치트리  옵션이 전혀 없는 순수 사채",
                 "주가와 무관하므로 같은 열의 값이 모두 같다.", "가정")
    fill(W, lambda i, r, L, Lp, Ln: (f"={L}$10+{L}$9" if i == n else
         f"=({Ln}{R0+r}*{L}$16+{Ln}{R0+r+1}*{L}$17)*EXP(-{L}$12*{K['dt']})"
         f"+{L}$9+{L}$10"))

    if _gs:
        W = newsheet(S11, "⑪ [GS] 전환확률트리",
                     "전환 1, 현금 0, 보유면 다음 두 칸의 평균.", f"{S14} · 다음 열 {S11}")
        # 현금(상환P·상환C)이 동점이면 0 이다. 전환은 허용오차만큼 앞설 때만 1 이다.
        _cash = lambda L, r: (f'OR(ABS({Q(S14)}!{L}{R0+r}-{L}$7)<{TOLX},'
                              f'ABS({Q(S14)}!{L}{R0+r}-{L}$8)<{TOLX})')
        fill(W, lambda i, r, L, Lp, Ln: (
            f'=IF({_cash(L, r)},0,'
            f'IF(ABS({Q(S14)}!{L}{R0+r}-{Q(S4)}!{L}{R0+r})<{TOLX},1,'
            + ('0))' if i == n else
               f'{Ln}{R0+r}*{L}$16+{Ln}{R0+r+1}*{L}$17))')), N4)

        W = newsheet(S12, "⑫ [GS] 위험조정할인율트리",
                     "이 칸을 직전 시점으로 할인할 때 쓰는 이자율이다. "
                     "전환확률로 무위험과 위험을 섞되 직전 구간의 선도이자율을 쓴다.", S11)
        fill(W, lambda i, r, L, Lp, Ln:
             (f"={Q(S11)}!{L}{R0+r}*{Lp}$11+(1-{Q(S11)}!{L}{R0+r})*{Lp}$12"
              if i > 0 else "=0"), P2)

        W = newsheet(S13, "⑬ [GS] 보유가치트리",
                     "다음 두 칸을 각 칸의 할인율로 따로 할인한다.", f"다음 열 {S12} · {S14}", call_on=False)
        fill(W, lambda i, r, L, Lp, Ln: (f"={L}$10+{L}$9" if i == n else
             f"={Q(S14)}!{Ln}{R0+r}*{L}$16*EXP(-{Q(S12)}!{Ln}{R0+r}*{K['dt']})"
             f"+{Q(S14)}!{Ln}{R0+r+1}*{L}$17*EXP(-{Q(S12)}!{Ln}{R0+r+1}*{K['dt']})+{L}$9"))

        W = newsheet(S14, "⑭ [GS] 금융상품가치트리",
                     "⑧과 같은 칸을 비교하면 모형 차이가 보인다.", f"{S4} · {S13}", call_on=False)
        fill(W, lambda i, r, L, Lp, Ln:
             f"=MAX({Q(S13)}!{L}{R0+r},{Q(S4)}!{L}{R0+r},{L}$7"
             + (f"+{L}$9)" if i == n else ")"))

    if _need15:
        # ── 15 트랜치 ──
        # 고른 모형의 블록만 담는다. GS 블록은 TF 다섯 블록을 참조하지 않으므로
        # 어느 쪽을 골라도 나머지 절반은 만들 필요가 없다.
        W = newsheet(S15, f"⑮ {KW} 트랜치  매도청구권이 걸리는 부분",
                     f"의무보유 때문에 전환 시작이 늦다. {KW0}와의 차이가 "
                     "매도청구권의 가치다. "
                     + ("전환가치 뒤에 GS 네 블록을 둔다. 고른 모형이 GS 이기 때문이다."
                        if _gs else
                        "전환가치 뒤에 TF 다섯 블록을 둔다. 고른 모형이 TF 이기 때문이다."),
                     "가정", conv_cell=K["cv30"])
        HH = n+3
        _bs = {"row": R0, "i": 0}
        def blk(t):
            _bs["row"] += HH
            row = _bs["row"]
            sec(W, row-1, f"{'가나다라마바'[_bs['i']]}  {t}", span=min(n+1, 14))
            _bs["i"] += 1
            put(W, row, 2, "r ＼ 스텝", bold=True, size=8, fill=LIGHT, border=True, align="center")
            for i in range(n+1):
                put(W, row, 3+i, i, bold=True, size=8, fmt=N0, align="center",
                    fill=LIGHT, border=True)
            for r in range(n+1):
                put(W, row+1+r, 2, r, bold=True, size=8, fmt=N0, align="center",
                    fill=LIGHT, border=True)
            return row+1

        # 전환가치는 두 모형이 함께 쓴다.
        c1 = blk("전환가치")
        for i in range(n+1):
            L = gl(3+i)
            for r in range(i+1):
                put(W, c1+1+r, 3+i,
                    f"=IF({L}$3=1,{Q(S1)}!{L}{R0+r}*{Q(S3)}!{L}{R0+r},0)",
                    fmt=N2, size=8, align="right")

        if not _gs:
            c2 = blk("지분가치")
            c3 = blk("부채가치")
            c4 = blk("보유가치")
            c5 = blk("금융상품가치")
            c6 = blk("의사결정")
            last = c5
            for i in range(n+1):
                L = gl(3+i); Ln = gl(4+i) if i < n else None
                for r in range(i+1):
                    p = lambda base, v, fm=N2, tx=False: put(W, base+1+r, 3+i, v,
                            fmt=(None if tx else fm), size=8,
                            align=("center" if tx else "right"))
                    if i == n:
                        p(c5, f"=MAX({L}{c1+1+r},MAX({L}$7,{L}$10)+{L}$9)")
                        p(c6, f'=IF({L}{c1+1+r}>=MAX({L}$7,{L}$10)+{L}$9+{TOLX},"전환",'
                              f'IF({L}$7>={L}$10-{TOLX},"상환P","만기상환"))', tx=True)
                        p(c2, f'=IF({L}{c6+1+r}="전환",{L}{c1+1+r},0)')
                        p(c3, f'=IF({L}{c6+1+r}="전환",0,MAX({L}$7,{L}$10)+{L}$9)')
                        p(c4, f"={L}$10+{L}$9")
                        continue
                    e = f"({Ln}{c2+1+r}*{L}$16+{Ln}{c2+2+r}*{L}$17)*EXP(-{L}$11*{K['dt']})"
                    b = f"({Ln}{c3+1+r}*{L}$16+{Ln}{c3+2+r}*{L}$17)*EXP(-{L}$12*{K['dt']})"
                    p(c4, f"={e}+{b}+{L}$9")
                    p(c5, f"=IF({L}$5=1,MAX(MIN({L}{c4+1+r},{L}$8),{L}{c1+1+r},{L}$7),"
                          f"MAX({L}{c4+1+r},{L}{c1+1+r},{L}$7))")
                    p(c6, f'=IF({L}{c1+1+r}>=MAX({L}$7,MIN({L}{c4+1+r},{L}$8))+{TOLX},"전환",'
                          f'IF({L}$7>=MIN({L}{c4+1+r},{L}$8)-{TOLX},"상환P",'
                          f'IF({L}{c4+1+r}<={L}$8+{TOLX},"보유","상환C")))', tx=True)
                    p(c2, f'=IF({L}{c6+1+r}="전환",{L}{c1+1+r},'
                          f'IF(OR({L}{c6+1+r}="상환P",{L}{c6+1+r}="상환C"),0,{e}))')
                    p(c3, f'=IF({L}{c6+1+r}="상환P",{L}$7,IF({L}{c6+1+r}="상환C",{L}$8,'
                          f'IF({L}{c6+1+r}="전환",0,{b}+{L}$9)))')
        else:
            # ⑪~⑭ 와 같은 순서지만 트랜치 자신의 헤더와 전환가치를 쓴다.
            c7 = blk("[GS] 전환확률")
            c8 = blk("[GS] 위험조정할인율")
            c9 = blk("[GS] 보유가치")
            c10 = blk("[GS] 금융상품가치")
            last = c10
            for i in range(n+1):
                L = gl(3+i); Lp = gl(2+i) if i > 0 else None; Ln = gl(4+i) if i < n else None
                for r in range(i+1):
                    p = lambda base, v, fm=N2: put(W, base+1+r, 3+i, v, fmt=fm,
                                                   size=8, align="right")
                    if i == n:
                        # 만기에는 TF 와 GS 가 같다.
                        p(c9, f"={L}$10+{L}$9")
                        p(c10, f"=MAX({L}{c1+1+r},MAX({L}$7,{L}$10)+{L}$9)")
                        p(c7, f"=IF({L}{c10+1+r}={L}{c1+1+r},1,0)", N4)
                    else:
                        p(c9, f"={Ln}{c10+1+r}*{L}$16*EXP(-{Ln}{c8+1+r}*{K['dt']})"
                              f"+{Ln}{c10+2+r}*{L}$17*EXP(-{Ln}{c8+2+r}*{K['dt']})+{L}$9")
                        p(c10, f"=IF({L}$5=1,MAX(MIN({L}{c9+1+r},{L}$8),{L}{c1+1+r},{L}$7),"
                               f"MAX({L}{c9+1+r},{L}{c1+1+r},{L}$7))")
                        p(c7, f"=IF({L}{c10+1+r}={L}{c1+1+r},1,"
                              f"IF(OR({L}{c10+1+r}={L}$7,{L}{c10+1+r}={L}$8),0,"
                              f"{Ln}{c7+1+r}*{L}$16+{Ln}{c7+2+r}*{L}$17))", N4)
                    p(c8, (f"={L}{c7+1+r}*{Lp}$11+(1-{L}{c7+1+r})*{Lp}$12"
                           if i > 0 else "=0"), P2)

        # 마지막으로 놓인 블록 아래에 둔다. TF 는 의사결정(바)이 값 블록(마) 뒤에
        # 오므로 last 를 그대로 쓰면 그 위에 겹쳐 쓰게 된다.
        RT = _bs["row"] + n + 3
        sec(W, RT, "결과", span=6)
        put(W, RT+1, 2, f"{KW} 트랜치 금융상품가치 · {tm.model} (t=0)", bold=True)
        put(W, RT+1, 3, f"=C{last+1}", bold=True, fmt=N2, align="right")

    # ── 16 부채요소 ──
    D = wb.create_sheet(S16); D.sheet_view.showGridLines = False
    D.column_dimensions["B"].width = 20
    for i in range(n+1): D.column_dimensions[gl(3+i)].width = 9
    title(D, 2, "⑯ 부채요소  전환권이 없는 사채에 조기상환권만 붙인 값", span=min(n+1, 14))
    put(D, 3, 2, "전환이 없으면 주가와 무관하므로 한 줄로 끝난다. "
        "만기부터 왼쪽으로 오며 MAX(조기상환금액, 계속보유)를 고른다.", color=GREY, size=9)
    for r, nm in enumerate(["Date", "time-step", "Flag(조기상환)", "조기상환금액", "쿠폰",
                            "만기상환", "위험 선도이자율", "부채요소"], start=4):
        put(D, r, 2, nm, bold=True, size=8, fill=LIGHT, border=True)
    for i in range(n+1):
        L = gl(3+i); Lp = gl(2+i) if i > 0 else None; Ln = gl(4+i) if i < n else None
        g = lambda r, v, fm=None, col="000000": put(D, r, 3+i, v, fmt=fm,
                                                    align="center", size=8, color=col)
        st = f"{L}$5"
        yr = f"({st}*{K['dt']}+{K['elm']}/12)"
        g(4, f"={K['d_base']}+{st}*{K['dt']}*365", DATE, GREY)
        g(5, (0 if i == 0 else f"={Lp}$5+1"), N0)
        g(6, f"=IF(AND({st}>={K['pst']},{st}<={K['pen']},"
             f"MOD({st}-{K['pst']},{K['frq']})=0),1,0)", N0)
        g(7, f"=IF({L}$6=1,IF({K['pyld']}>0,"
             f"100*(1+MAX(0,({K['pyld']}-{K['cpn']})/{K['pyld']}*"
             f"((1+{K['pyld']}/{K['pcmp']})^({K['pcmp']}*{yr})-1))),"
             f"{K['prate']}),0)", N2)
        g(8, f"=IF(AND({st}>0,MOD({st},{K['ipay']})=0),"
             f"100*{K['cpn']}*{K['ipaym']}/12,0)", N2)
        g(9, f"=IF({st}={K['n']},{K['red']},0)", N2)
        if i < n: g(10, forward_rate(CR, i*dt_, (i+1)*dt_), P2, AMB)
        g(11, (f"=MAX({L}$7,{L}$9)+{L}$8" if i == n else
               f"=MAX({L}$7,{Ln}11*EXP(-{L}$10*{K['dt']})+{L}$8)"), N2)
    put(D, 13, 2, "부채요소 (t=0)", bold=True)
    put(D, 13, 3, "=C11", bold=True, fmt=N2, align="right")

    if _bdt:
        # ── BDT 단기이자율 · BDT 부채요소 ──
        # 기준금리 a_i 는 곡선에 맞추려고 역산한 값이라 수식으로 펼 수 없다.
        # 선도이자율과 같은 자리다. 격자와 역진은 전부 수식이다.
        BP, BV = bdt_grid(tm, True)
        SB1, SB2 = "BDT 단기이자율", "BDT 부채요소"
        RB = 14                                  # 표 첫 자료행 (j = 0)

        def bhead(W, ttl, note):
            W.sheet_view.showGridLines = False
            W.column_dimensions["B"].width = 20
            for i in range(n+1): W.column_dimensions[gl(3+i)].width = 9
            title(W, 2, ttl, span=min(n+1, 14))
            put(W, 3, 2, note, color=GREY, size=9)
            for r, nm in enumerate(["Date", "time-step", "Flag(조기상환)", "조기상환금액",
                                    "쿠폰", "만기상환", "기준금리 a", "확정 스프레드"],
                                   start=4):
                put(W, r, 2, nm, bold=True, size=8, fill=LIGHT, border=True)
            for i in range(n+1):
                L = gl(3+i); Lp = gl(2+i) if i > 0 else None
                g = lambda r, v, fm=None, col="000000": put(W, r, 3+i, v, fmt=fm,
                                                            align="center", size=8, color=col)
                st = f"{L}$5"
                yr = f"({st}*{K['dt']}+{K['elm']}/12)"
                g(4, f"={K['d_base']}+{st}*{K['dt']}*365", DATE, GREY)
                g(5, (0 if i == 0 else f"={Lp}$5+1"), N0)
                g(6, f"=IF(AND({st}>={K['pst']},{st}<={K['pen']},"
                     f"MOD({st}-{K['pst']},{K['frq']})=0),1,0)", N0)
                g(7, f"=IF({L}$6=1,IF({K['pyld']}>0,"
                     f"100*(1+MAX(0,({K['pyld']}-{K['cpn']})/{K['pyld']}*"
                     f"((1+{K['pyld']}/{K['pcmp']})^({K['pcmp']}*{yr})-1))),"
                     f"{K['prate']}),0)", N2)
                g(8, f"=IF(AND({st}>0,MOD({st},{K['ipay']})=0),"
                     f"100*{K['cpn']}*{K['ipaym']}/12,0)", N2)
                g(9, f"=IF({st}={K['n']},{K['red']},0)", N2)
                if i < n:
                    g(10, BP["a"][i], P2, AMB)
                    g(11, BP["add"][i], P2, AMB)
            put(W, RB-1, 2, "j ＼ 스텝", bold=True, size=8, fill=LIGHT,
                border=True, align="center")
            for i in range(n+1):
                put(W, RB-1, 3+i, i, bold=True, size=8, fmt=N0, align="center",
                    fill=LIGHT, border=True)
            for j in range(n+1):
                put(W, RB+j, 2, j, bold=True, size=8, fmt=N0, align="center",
                    fill=LIGHT, border=True)
            W.freeze_panes = f"C{RB}"

        W = wb.create_sheet(SB1)
        bhead(W, "BDT 단기이자율격자  r(i,j) = a · exp(2σ·j·√Δt)",
              "로그정규라 이자율이 음수가 되지 않는다. j 는 상승 횟수이고 클수록 "
              "금리가 높다 — 주가 트리와 달리 위로 갈수록 낮다. 기준금리 a 는 곡선을 "
              "정확히 되돌리도록 역산한 값이라 주황색이다. σ 를 바꾸시려면 앱에서 "
              "조서를 다시 만드셔야 한다.")
        for i in range(n):
            L = gl(3+i)
            for j in range(i+1):
                put(W, RB+j, 3+i,
                    f"={L}$10*EXP(2*{K['bsig']}*$B{RB+j}*SQRT({K['dt']}))",
                    fmt=P2, size=8, align="right")

        # ── 캘리브레이션 검산 ──
        # 기준금리 a 만 값이라 조서만 보면 근거 없는 상수처럼 보인다. 그래서
        # 도달가격 Q 를 수식으로 쌓아 보인다. Σ_j Q(k,j) 가 시장 할인계수와
        # 같아야 하고, a 를 바로 그 조건에 맞춰 역산했다 — 무차익거래 조건이다.
        QR = RB+n+3
        sec(W, QR-2, f"캘리브레이션 검산 — 격자가 {BP['base_nm']}을 되돌리는가",
            span=min(n+1, 14))
        put(W, QR-1, 2, "j ＼ 스텝", bold=True, size=8, fill=LIGHT,
            border=True, align="center")
        for i in range(n+1):
            put(W, QR-1, 3+i, i, bold=True, size=8, fmt=N0, align="center",
                fill=LIGHT, border=True)
        for j in range(n+1):
            put(W, QR+j, 2, j, bold=True, size=8, fmt=N0, align="center",
                fill=LIGHT, border=True)
        for i in range(n+1):
            L, Lp = gl(3+i), (gl(2+i) if i > 0 else None)
            for j in range(i+1):
                if i == 0:
                    v = 1
                else:
                    # 아래에서 올라온 몫 + 위에서 내려온 몫. 삼각형 밖은 비어
                    # 있으므로 엑셀에서 0 으로 읽힌다.
                    up = (f"0.5*{Lp}{QR+j-1}*EXP(-{Lp}{RB+j-1}*{K['dt']})"
                          if j > 0 else "")
                    dn = (f"0.5*{Lp}{QR+j}*EXP(-{Lp}{RB+j}*{K['dt']})"
                          if j <= i-1 else "")
                    v = "=" + "+".join(x for x in (up, dn) if x)
                put(W, QR+j, 3+i, v, fmt=N6, size=8, align="right")
        for k2, nm in enumerate(("모형 무이표채  Σ Q", "시장 할인계수", "차이")):
            r2 = QR+n+1+k2
            put(W, r2, 2, nm, bold=True, size=8, fill=LIGHT, border=True)
            for i in range(n+1):
                L = gl(3+i)
                v = (f"=SUM({L}{QR}:{L}{QR+n})" if k2 == 0 else
                     (BP["mkt"][i] if k2 == 1 else f"={L}{r2-2}-{L}{r2-1}"))
                put(W, r2, 3+i, v, fmt=N6, size=8, align="right",
                    bold=(k2 == 2), color=(AMB if k2 == 1 else "000000"))
        put(W, QR+n+5, 2,
            "도달가격 Q(i,j) 는 그 칸에 이르는 경로의 확률을 그 경로의 할인율로 "
            "할인해 더한 값이다. 스텝별로 모두 더하면 그 만기의 무이표채 가격이 "
            "되고, 그것이 시장 할인계수(주황)와 같아야 한다 — 무차익거래 조건이다. "
            "기준금리 a 를 이 조건에 맞춰 이분법으로 역산했으므로 차이가 0 이다. "
            "σ 를 바꾸면 a 도 함께 바뀌어야 하므로 앱에서 조서를 다시 만드셔야 "
            "한다 — 이 시트에서 σ 만 바꾸면 차이가 0 에서 벗어난다.", color=GREY, size=9)

        W = wb.create_sheet(SB2)
        bhead(W, "BDT 부채요소  전환 없는 사채 + 조기상환권",
              "MAX(조기상환금액, 계속보유) 를 고른다. 계속보유는 다음 두 칸을 "
              "0.5 씩 섞어 그 칸의 단기이자율로 할인한 값이다. 확정 격자와의 차이가 "
              "곧 금리에서 나오는 옵션의 시간가치다.")
        for i in range(n, -1, -1):
            L = gl(3+i); Ln = gl(4+i) if i < n else None
            for j in range(i+1):
                if i == n:
                    v = f"=MAX({L}$7,{L}$9)+{L}$8"
                else:
                    d = f"EXP(-('{SB1}'!{L}{RB+j}+{L}$11)*{K['dt']})"
                    v = (f"=MAX({L}$7,(0.5*{Ln}{RB+j+1}+0.5*{Ln}{RB+j})*{d}+{L}$8)")
                put(W, RB+j, 3+i, v, fmt=N2, size=8, align="right")
        put(W, RB+n+2, 2, "부채요소 (t=0)", bold=True)
        put(W, RB+n+2, 3, f"=C{RB}", bold=True, fmt=N2, align="right")
        put(W, RB+n+3, 2, "금리 고정 격자 (⑯)", bold=True)
        put(W, RB+n+3, 3, f"={Q(S16)}!C13", fmt=N2, align="right")
        put(W, RB+n+4, 2, "차이 = 금리에서 나온 옵션가치", bold=True)
        put(W, RB+n+4, 3, f"=C{RB+n+2}-C{RB+n+3}", bold=True, fmt=N2, align="right")

    if _need1 or _need2:
        # ── 17~20 옵션차익혼합할인법 ──
        # 제3자 지정 가능 콜옵션은 전환사채를 기초자산으로 하는 복합옵션이다.
        # 기초자산은 콜과 부속조항(의무보유)을 뺀 ⑧ 이다 (책 4.4.3, 부속예제 4-4).
        W = newsheet(S17, "⑰ 구성비율트리  지분 몫 ÷ 전환사채 가치",
                     "노드 가치 중 주식에서 온 몫의 비율이다.", f"{S5} · {S8}", call_on=False)
        fill(W, lambda i, r, L, Lp, Ln:
             f"=IF({Q(S8)}!{L}{R0+r}=0,0,{Q(S5)}!{L}{R0+r}/{Q(S8)}!{L}{R0+r})", N4)

        if _need1:                      # ⑱·⑳ 은 혼합할인율(방법 1) 전용이다
            W = newsheet(S18, "⑱ 혼합할인율트리  구성비율 × 무위험 + (1−구성비율) × 위험",
                         "이 칸을 직전 시점으로 할인할 때 쓰는 이자율이다. "
                         "지분 몫에는 무위험, 채권 몫에는 위험 선도이자율을 섞는다.",
                         S17, call_on=False)
            fill(W, lambda i, r, L, Lp, Ln:
                 (f"={Q(S17)}!{L}{R0+r}*{Lp}$11+(1-{Q(S17)}!{L}{R0+r})*{Lp}$12"
                  if i > 0 else "=0"), P2)

        W = newsheet(S19, "⑲ 콜 페이오프트리  MAX(전환사채 가치 − 매도청구금액, 0)",
                     "매도청구 행사기간에만 값이 생긴다. 기초자산은 ⑧ 이다.", S8)
        fill(W, lambda i, r, L, Lp, Ln:
             f"=IF({L}$5=1,MAX({Q(S8)}!{L}{R0+r}-{L}$8,0),0)")

        if _need1:
            W = newsheet(S20, "⑳ 매도청구권가치트리  미국형 복합옵션",
                         "자식의 구성비율로 섞은 할인율로 자식을 각각 할인한다.",
                         f"{S18} · {S19} · 다음 열 {S20}")
            fill(W, lambda i, r, L, Lp, Ln: (f"={Q(S19)}!{L}{R0+r}" if i == n else
                 f"=MAX({Q(S19)}!{L}{R0+r},"
                 f"{Ln}{R0+r}*{L}$16*EXP(-{Q(S18)}!{Ln}{R0+r}*{K['dt']})"
                 f"+{Ln}{R0+r+1}*{L}$17*EXP(-{Q(S18)}!{Ln}{R0+r+1}*{K['dt']}))"))
            put(W, R0+n+3, 2, "매도청구권 (한도 반영 전, t=0)", bold=True)
            put(W, R0+n+3, 3, f"=C{R0}", bold=True, fmt=N2, align="right")

    if _need2:
        # ── 21~24 옵션차익혼합할인법 · 방법 2 (지분·부채 분리) ──
        # 값 하나를 섞은 할인율로 할인하는 대신, 콜옵션 가치를 지분 몫과 부채 몫으로
        # 쪼개 각각 무위험·위험 선도이자율로 할인한다 (책 4.4.3, 부속예제 4-4 방법2).
        W = newsheet(S21, "㉑ [방법2] 지분 몫 보유가치",
                     "다음 두 칸의 지분 몫을 무위험 선도이자율로 할인한다.",
                     f"다음 열 {S23}")
        fill(W, lambda i, r, L, Lp, Ln: ("=0" if i == n else
             f"=({Q(S23)}!{Ln}{R0+r}*{L}$16+{Q(S23)}!{Ln}{R0+r+1}*{L}$17)"
             f"*EXP(-{L}$11*{K['dt']})"), N4)

        W = newsheet(S22, "㉒ [방법2] 부채 몫 보유가치",
                     "다음 두 칸의 부채 몫을 위험 선도이자율로 할인한다.",
                     f"다음 열 {S24}")
        fill(W, lambda i, r, L, Lp, Ln: ("=0" if i == n else
             f"=({Q(S24)}!{Ln}{R0+r}*{L}$16+{Q(S24)}!{Ln}{R0+r+1}*{L}$17)"
             f"*EXP(-{L}$12*{K['dt']})"), N4)

        ex2 = (lambda L, r: f"{Q(S19)}!{L}{R0+r}>={Q(S21)}!{L}{R0+r}+{Q(S22)}!{L}{R0+r}")
        W = newsheet(S23, "㉓ [방법2] 매도청구권 · 지분 몫",
                     "행사하면 페이오프의 지분 몫, 아니면 보유가치의 지분 몫이다. "
                     "만기에는 보유가치가 0 이라 언제나 페이오프를 쪼갠다.",
                     f"{S17} · {S19} · {S21} · {S22}")
        fill(W, lambda i, r, L, Lp, Ln:
             f"=IF({ex2(L, r)},{Q(S19)}!{L}{R0+r}*{Q(S17)}!{L}{R0+r},"
             f"{Q(S21)}!{L}{R0+r})", N4)

        W = newsheet(S24, "㉔ [방법2] 매도청구권 · 부채 몫",
                     "행사 판단은 ㉓ 과 같다. 지분 몫의 나머지가 부채 몫이다.",
                     f"{S17} · {S19} · {S21} · {S22}")
        fill(W, lambda i, r, L, Lp, Ln:
             f"=IF({ex2(L, r)},{Q(S19)}!{L}{R0+r}*(1-{Q(S17)}!{L}{R0+r}),"
             f"{Q(S22)}!{L}{R0+r})", N4)
        put(W, R0+n+3, 2, "매도청구권 · 방법2 (한도 반영 전, t=0)", bold=True)
        put(W, R0+n+3, 3, f"={Q(S23)}!C{R0}+C{R0}", bold=True, fmt=N2, align="right")

    # ── 결과 ──
    R = wb.create_sheet("결과"); R.sheet_view.showGridLines = False
    for cc, w in (("B", 36), ("C", 14), ("D", 16), ("E", 12), ("F", 42)):
        R.column_dimensions[cc].width = w
    title(R, 2, "평가결과", span=5)
    put(R, 3, 2, "모든 값이 앞의 트리 시트에서 수식으로 넘어온다.", color=GREY, size=9)
    sec(R, 5, "1. 트랜치", span=5)
    # 주계약과 부채요소에는 전환이 없어 TF 와 GS 가 항상 같다. 모형 선택이
    # 갈라지는 곳은 트랜치 둘뿐이므로 여기서 한 번만 고른다.
    # 앱에서 고른 것만 값이 든다. 행 자리는 그대로 두어 아래 참조가 깨지지 않게 한다.
    B_ = ""
    _t70 = f"={Q(S14)}!C{R0}" if _gs else f"={Q(S8)}!C{R0}"
    _t30 = f"={Q(S15)}!C{RT+1}" if _need15 else B_
    for i, (nm, fx) in enumerate([
            (f"{KW0} 트랜치 · TF", B_ if _gs else _t70),
            (f"{KW0} 트랜치 · GS", _t70 if _gs else B_),
            (f"{KW} 트랜치 · TF", B_ if _gs else _t30),
            (f"{KW} 트랜치 · GS", _t30 if _gs else B_),
            (f"적용 · {KW0} 트랜치", _t70),
            (f"적용 · {KW} 트랜치", _t30),
            ("가중 평균", f"=(1-{K['cw']})*C10+{K['cw']}*C11" if _need15 else B_)]):
        bold = (i >= 4)
        put(R, 6+i, 2, nm, bold=bold, border=True)
        put(R, 6+i, 3, fx, bold=bold, fmt=N2, align="right", border=True)
    sec(R, 14, "2. 구성요소", span=5)
    put(R, 15, 3, "100 기준", bold=True, fill=LIGHT, align="center", border=True)
    put(R, 15, 4, "전액 기준 (원)", bold=True, fill=LIGHT, align="center", border=True)
    put(R, 15, 5, "공시", bold=True, fill=LIGHT, align="center", border=True)
    items = [("주계약", f"={Q(S10)}!C{R0}"),
             ("부채요소 (사채 + 조기상환권)",
              f"='BDT 부채요소'!C{14+n+2}" if _bdt else f"={Q(S16)}!C13"),
             ("조기상환청구권", "=C17-C16"),
             ("매도청구권 · 유무가치비교법",
              f"={K['cw']}*(C10-C11)" if _need15 else B_),
             ("매도청구권 · 옵션차익 · 혼합할인율",
              f"={K['cw']}*{Q(S20)}!C{R0+n+3}" if _need1 else B_),
             ("매도청구권 · 옵션차익 · 지분·부채 분리",
              f"={K['cw']}*{Q(S24)}!C{R0+n+3}" if _need2 else B_),
             ("매도청구권자산 (적용값)", f"=C{19+_km}" if _hascall else "=0"),
             ("전환권대가 (자본일 때)", f'=IF({K["eqcls"]}=1,100-C17+C22,"")'),
             ("복합내재파생상품 (부채일 때)", f'=IF({K["eqcls"]}=0,C10-C16,"")'),
             ("주계약 잔여 (부채일 때)", f'=IF({K["eqcls"]}=0,100+C22-C24,"")')]
    for i, (nm, fx) in enumerate(items):
        r = 16+i
        put(R, r, 2, nm, bold=True, border=True)
        put(R, r, 3, fx, bold=True, fmt=N2, align="right", border=True)
        put(R, r, 4, f'=IF(ISNUMBER(C{r}),C{r}/100*{K["face"]},"")',
            fmt=N0, align="right", border=True)
    sec(R, 27, "3. 검산", span=5)
    for i, (nm, fx, jd) in enumerate([
            ("위험중립가중치 q", f"={K['q']}", '=IF(AND(C28>0,C28<1),"적합","확인 필요")'),
            ("도달확률 합계 (마지막 열)",
             f"=SUM(도달확률!{gl(3+n)}5:{gl(3+n)}{5+n})",
             '=IF(ABS(C29-1)<0.0001,"적합","확인 필요")'),
            ("전체 ≥ 주계약", "=C10-C16", '=IF(C30>=0,"적합","확인 필요")'),
            ("배분 합계 = 100",
             f'=IF({K["ksep"]}=1,IF({K["eqcls"]}=1,C16+C18-C22+C23,C25+C24-C22),'
             f'IF({K["eqcls"]}=1,C16+C18-C22+C23,C25+C24-C22))',
             '=IF(ABS(C31-100)<0.01,"적합","확인 필요")'),
            # 상태확장 격자는 재결합하지 않아 엑셀 트리 한 장으로 옮길 수 없다.
            # 앱이 상태확장으로 계산했다면 이 조서는 근사값이므로 그 사실을 밝힌다.
            ("앱 계산값 · 상태확장 격자", (b2 if tm.carry == 0 else ""),
             '=IF(C32="","해당 없음 (앱과 조서가 같은 방법)",'
             'IF(ABS(C32-C10)<0.01,"적합","★ 조서는 근사값 — 아래 설명"))')]):
        put(R, 28+i, 2, nm, border=True)
        put(R, 28+i, 3, fx, fmt=N4, align="right", border=True,
            color=(AMB if i == 4 else "000000"))
        put(R, 28+i, 5, jd, align="center", border=True)
    if tm.carry == 0:
        put(R, 33, 2,
            "★ 앱은 상태확장 격자로 계산했다. 조정일마다 전환가액이 갈라져 같은 칸에 "
            "여러 값이 존재하므로 엑셀 트리 한 장으로는 옮길 수 없다. 이 조서는 "
            "경로가중치 근사로 다시 계산한 값이다. 위 두 숫자의 차이가 근사 오차이며, "
            "정확한 값은 앱 계산값(주황)이다.", color=RED, size=9)
    put(R, 34, 2, "이 조서에는 앱에서 고른 방법만 들어 있습니다. 다른 신용위험 처리나 "
        "다른 매도청구권 평가방법의 값은 이 조서에 없습니다.", color=GREY, size=9)
    put(R, 35, 2, "주황색 숫자만 값이다. 선도이자율은 부트스트래핑 결과라 엑셀에서 재현하지 않는다.",
        color=AMB, size=9)

    # ── 이자율곡선 ──
    # 각 트리 11·12행의 선도이자율이 어디서 왔는지 남긴다. 부트스트래핑은
    # 엑셀에서 재현하지 않으므로 여기서도 값이다.
    C = wb.create_sheet("이자율곡선"); C.sheet_view.showGridLines = False
    for cc, w in (("B", 12), ("C", 14), ("D", 14), ("E", 14), ("F", 14), ("G", 14)):
        C.column_dimensions[cc].width = w
    title(C, 2, "기간별 이자율", span=6)
    put(C, 3, 2, "선도이자율  f(t, t+Δt) = [ r(t+Δt)×(t+Δt) − r(t)×t ] ÷ Δt   "
        "· 각 트리 시트 11·12행이 이 값이다.", color=GREY, size=9)
    for i, h in enumerate(["시점 (년)", "무위험 현물", "무위험 선도",
                           "위험 현물", "위험 선도", "스프레드"]):
        put(C, 5, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    for k in range(9):
        t_ = tm.T*k/8; i = min(n-1, round(t_/dt_))
        fr = forward_rate(RF, i*dt_, (i+1)*dt_); fc = forward_rate(CR, i*dt_, (i+1)*dt_)
        for j2, v in enumerate([t_, RF(t_), fr, CR(t_), fc, fc-fr]):
            put(C, 6+k, 2+j2, v, fmt=(N2 if j2 == 0 else P2), align="right",
                border=True, color=(None if j2 == 0 else AMB))
    rr = 16
    if tm.y_type == "par" and len(tm.cr_curve) >= 2:
        sec(C, rr, "부트스트래핑 — 위험 곡선", span=6)
        for i, h in enumerate(["만기 (년)", "만기수익률", "할인계수", "현물 (연속)"]):
            put(C, rr+1, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
        bs = [x for x in bootstrap_df(tm.cr_curve, tm.T, tm.cmp_cr) if x[0] > 0]
        for k, (t_, df) in enumerate(bs):
            for j2, v in enumerate([t_, _lin(tm.cr_curve, t_), df, -math.log(df)/t_]):
                put(C, rr+2+k, 2+j2, v,
                    fmt=(N2 if j2 == 0 else (N6 if j2 == 2 else P2)),
                    align="right", border=True)
        rr = rr+3+len(bs)
    put(C, rr, 2, "주황색은 값이다. 곡선을 바꾸려면 앱에서 조서를 다시 만들어야 한다.",
        color=AMB, size=9)

    # ── 상각표 ──
    # 유효이자율만 역산 결과(값)이고, 나머지는 살아 있는 수식이다.
    r_eir, rows_eir, redm, nper = eir
    M = wb.create_sheet("상각표"); M.sheet_view.showGridLines = False
    for cc, w in (("B", 10), ("C", 13), ("D", 12), ("E", 16), ("F", 14),
                  ("G", 14), ("H", 16)):
        M.column_dimensions[cc].width = w
    title(M, 2, "주계약 상각표", span=7)
    put(M, 3, 2, "주계약(옵션 없는 사채)을 유효이자율법으로 상각한다. "
        "기말 잔액이 만기상환금액과 맞아떨어져야 한다. "
        "지급일은 계약상 일정이므로 발행일부터 센다. 평가기준일이 발행일보다 뒤이면 "
        "첫 회차만 짧고 나머지는 온전한 한 주기다. 회차 수는 노드가 아니라 "
        "이자 지급주기를 따른다.", color=GREY, size=9)
    sec(M, 5, "유효이자율 역산", span=6)
    for i, (k, fx, fm, val) in enumerate([
            # 부채로 분류하면 잔여로 떨어진 금액이 인식액이다. 이론값(C16)이 아니다.
            ("주계약 (인식액)", f'=IF({K["eqcls"]}=1,결과!C16,결과!C25)', N2, None),
            ("만기상환금액", f"={K['red']}", N2, None),
            ("표면이자 (회당)", f"=100*{K['cpn']}*{K['ipaym']}/12", N2, None),
            ("상각 횟수", None, N0, nper)]):
        put(M, 6+i, 2, k, border=True)
        put(M, 6+i, 3, fx if fx else val, fmt=fm, align="right", border=True)
    put(M, 10, 2, "유효이자율 (연, 이산복리)", bold=True, fill=BAND, border=True)
    put(M, 10, 3, r_eir, bold=True, fill=BAND, fmt=P2, align="right",
        border=True, color=AMB)
    put(M, 11, 2, "주황색은 역산 결과라 값이다. 아래 표는 이 이자율로 도는 수식이다.",
        color=AMB, size=9)
    sec(M, 13, "상각 내역", span=7)
    for i, h in enumerate(["회차", "지급일", "경과연수", "기초", "이자비용",
                           "지급이자", "기말"]):
        put(M, 14, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    for i, row in enumerate(rows_eir):
        r = 15+i; last = (i == len(rows_eir)-1); fl = BAND if last else None
        prev = r-1
        put(M, r, 2, row[0], fmt=N0, align="right", border=True, bold=last, fill=fl)
        # 계약상 지급일이다. 마지막은 만기일, 나머지는 발행일에 달을 더한다.
        pdf = (f"={K['d_mat']}" if last else
               f"=EDATE({K['d_issue']},{int(round(pay_index(tm, row[1])*tm.ipay))})")
        put(M, r, 3, pdf, fmt=DATE, align="right", border=True, bold=last, fill=fl)
        put(M, r, 4, row[1], fmt=N2, align="right", border=True, bold=last, fill=fl)
        put(M, r, 5, ("=$C$6" if i == 0 else f"=H{prev}"), fmt=N2, align="right",
            border=True, bold=last, fill=fl)
        # 이자비용 = 기초 × (1+r)^기간 − 기초.  회차마다 기간이 달라 이렇게 쓴다.
        gap = f"(D{r}" + ("" if i == 0 else f"-D{prev}") + ")"
        put(M, r, 6, f"=E{r}*((1+$C$10)^{gap}-1)", fmt=N2, align="right",
            border=True, bold=last, fill=fl)
        put(M, r, 7, "=$C$8", fmt=N2, align="right", border=True, bold=last, fill=fl)
        put(M, r, 8, f"=E{r}+F{r}-G{r}", fmt=N2, align="right",
            border=True, bold=last, fill=fl)
    lr = 15+len(rows_eir)
    put(M, lr, 2, "검산 · 기말 잔액 = 만기상환금액", bold=True, border=True)
    put(M, lr, 4, f"=H{lr-1}-$C$7", fmt=N2, align="right", border=True)
    put(M, lr, 6, f'=IF(ABS(H{lr-1}-$C$7)<0.01,"적합","확인 필요")',
        align="center", border=True)
    # 유효이자율은 값이라, 가정을 고쳐 인식액이 움직이면 표가 닫히지 않는다.
    put(M, lr+1, 2, "검산 · 인식액이 역산 당시와 같은가", bold=True, border=True)
    put(M, lr+1, 4, f"=C6-{rows_eir[0][2] if rows_eir else 0!r}", fmt=N2,
        align="right", border=True)
    put(M, lr+1, 6, f'=IF(ABS(C6-{rows_eir[0][2] if rows_eir else 0!r})<0.0001,'
        f'"적합","앱에서 다시 만드십시오")', align="center", border=True)

    # ── 회계처리 ──
    E = wb.create_sheet("회계처리"); E.sheet_view.showGridLines = False
    for cc, w in (("B", 34), ("C", 14), ("D", 14), ("E", 18), ("F", 18)):
        E.column_dimensions[cc].width = w
    title(E, 2, "회계처리", span=5)
    put(E, 3, 2, "기업회계기준서 제1032호 문단 31·32 — 부채요소를 먼저 정하고 나머지를 자본에 배분한다. "
        "매도청구권은 제3자에게 이전될 수 있어 별도의 금융상품이다 (제1109호 문단 4.3.1, "
        "회계기준원 질의회신 2022-I-KQA006, 금융위 2022.5.3 감독지침). "
        "전환권이 부채이면 전환권과 조기상환권은 상호의존적이므로 하나의 복합내재파생상품으로 "
        "전체로서 측정한다 (제1109호 문단 B4.3.4).",
        color=GREY, size=9)
    if tm.elapsed_m > 0.01:
        put(E, 4, 2, "※ 평가기준일이 발행일보다 뒤입니다. 아래 배분은 최초 인식용이므로 "
            "결산 회계처리에 그대로 쓰지 마십시오. 결산일에 쓰는 것은 파생상품 공정가치뿐이고, "
            "주계약은 발행일 배분액을 유효이자율로 상각한 장부금액입니다.", color=RED, size=9)
    sec(E, 5, "1. 최초 인식 배분", span=5)
    for i, h in enumerate(["항목", "100 기준", "전액 기준 (원)"]):
        put(E, 6, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    # 매도청구권을 별도 금융상품으로 볼지 내재파생에 넣을지에 따라 표가 갈린다.
    # 넣는 쪽이면 파생 줄에서 콜을 빼고 자산 줄을 비운다. 합계는 어느 쪽이든 100.
    KS = K["ksep"]
    al2 = [("주계약", f'=IF({K["eqcls"]}=1,결과!C16,결과!C25)'),
           ("조기상환청구권 · 파생상품부채",
            f'=IF({K["eqcls"]}=0,"",IF({KS}=1,결과!C18,결과!C18-결과!C22))'),
           ("복합내재파생상품 · 파생상품부채",
            f'=IF({K["eqcls"]}=1,"",IF({KS}=1,결과!C24,결과!C24-결과!C22))'),
           ("매도청구권 · 파생상품자산", f'=IF({KS}=1,-결과!C22,"")'),
           ("전환권대가 · 자본", f'=IF({K["eqcls"]}=1,결과!C23,"")')]
    for i, (nm, fx) in enumerate(al2):
        r = 7+i
        put(E, r, 2, nm, border=True)
        put(E, r, 3, fx, fmt=N2, align="right", border=True)
        put(E, r, 4, f'=IF(ISNUMBER(C{r}),C{r}/100*{K["face"]},"")',
            fmt=N0, align="right", border=True)
    put(E, 12, 2, "합계", bold=True, fill=BAND, border=True)
    put(E, 12, 3, "=SUM(C7:C11)", bold=True, fill=BAND, fmt=N2, align="right", border=True)
    put(E, 12, 4, "=SUM(D7:D11)", bold=True, fill=BAND, fmt=N0, align="right", border=True)
    sec(E, 14, "2. 분개", span=5)
    for i, h in enumerate(["계정", "차변 (100)", "대변 (100)", "차변 (원)", "대변 (원)"]):
        put(E, 15, 2+i, h, bold=True, fill=LIGHT, align="center", border=True, size=9)
    je2 = [("현금", "=100", None),
           ("파생상품자산 (매도청구권)", f'=IF({KS}=1,결과!C22,"")', None),
           ("　전환사채 (주계약)", None, f'=IF({K["eqcls"]}=1,결과!C16,결과!C25)'),
           ("　파생상품부채 (조기상환청구권)", None,
            f'=IF({K["eqcls"]}=0,"",IF({KS}=1,결과!C18,결과!C18-결과!C22))'),
           ("　파생상품부채 (복합내재파생상품)", None,
            f'=IF({K["eqcls"]}=1,"",IF({KS}=1,결과!C24,결과!C24-결과!C22))'),
           ("　전환권대가 (자본)", None, f'=IF({K["eqcls"]}=1,결과!C23,"")')]
    for i, (nm, dr, cr) in enumerate(je2):
        r = 16+i
        put(E, r, 2, nm, size=9, border=True)
        put(E, r, 3, dr if dr else "", fmt=N2, align="right", border=True)
        put(E, r, 4, cr if cr else "", fmt=N2, align="right", border=True)
        put(E, r, 5, f'=IF(ISNUMBER(C{r}),C{r}/100*{K["face"]},"")',
            fmt=N0, align="right", border=True)
        put(E, r, 6, f'=IF(ISNUMBER(D{r}),D{r}/100*{K["face"]},"")',
            fmt=N0, align="right", border=True)
    put(E, 22, 2, "합계", bold=True, fill=BAND, border=True)
    for j2, col in enumerate("CDEF"):
        put(E, 22, 3+j2, f"=SUM({col}16:{col}21)", bold=True, fill=BAND,
            fmt=(N2 if j2 < 2 else N0), align="right", border=True)
    put(E, 23, 2, "차변과 대변이 일치해야 한다", bold=True, border=True)
    put(E, 23, 3, '=IF(ABS(C22-D22)<0.01,"적합","오류")', align="center", border=True)
    put(E, 25, 2, "최초 인식에는 어떠한 손익도 생기지 않는다.", color=GREY, size=9)

    # ── 분리 판단 ──
    # 화면과 같은 함수가 만든 문안이라 둘이 어긋날 수 없다.
    SP = split_test(tm, full, b0, b1, b2, ca, eir[1])
    J = wb.create_sheet("분리 판단"); J.sheet_view.showGridLines = False
    J.column_dimensions["B"].width = 24; J.column_dimensions["C"].width = 92
    title(J, 2, "내재파생상품 분리 판단", span=2)
    put(J, 3, 2, "판단 순서 — 기업회계기준서 제1109호 문단 B4.3.5 말미는 제1032호에 "
                 "따라 전환채무상품의 자본요소를 분리하기 전에 내재된 콜옵션이나 "
                 "풋옵션이 주채무계약과 밀접하게 관련되어 있는지를 판단하라고 정한다.",
        color=GREY, size=9)
    _r = 5
    for _k, _nm in (("put", "조기상환청구권"), ("call", "매도청구권")):
        _d = SP[_k]
        sec(J, _r, _nm, span=2); _r += 1
        put(J, _r, 2, "결론", bold=True, border=True)
        put(J, _r, 3, _d["결론"], bold=True, border=True); _r += 1
        for _i, _x in enumerate(_d["이유"]):
            put(J, _r, 2, "판단 근거" if _i == 0 else "", border=True)
            put(J, _r, 3, _x, border=True); _r += 1
        put(J, _r, 2, "기준서", border=True)
        put(J, _r, 3, " · ".join(_d["근거"]) or "—", border=True); _r += 1
        put(J, _r, 2, "평가방법", border=True)
        put(J, _r, 3, _d["평가"].replace("**", ""), border=True); _r += 1
        for _a, _v in _d["지표"].items():
            put(J, _r, 2, _a, border=True)
            put(J, _r, 3, (f"{_v*100:.1f}%" if _a == "차이" else
                           ("예" if _v is True else "아니오" if _v is False
                            else f"{_v:,.4f}")), border=True); _r += 1
        _r += 1
    put(J, _r, 2, "이 시트는 앱의 「분리 판단」 화면과 같은 함수가 만든다. 계약 조항 "
                  "확인 항목을 바꾸면 결론과 문안이 함께 바뀐다.", color=GREY, size=9)
    for _row in J.iter_rows(min_row=5, max_row=_r, min_col=3, max_col=3):
        for _c in _row: _c.alignment = Alignment(wrap_text=True, vertical="top")

    # ── 해설 ──
    H = wb.create_sheet("해설", 0); H.sheet_view.showGridLines = False
    H.column_dimensions["B"].width = 22; H.column_dimensions["C"].width = 96
    title(H, 2, "수식 조서 사용 안내", span=2)
    ex = [("성격", ""),
      ("살아 있는 수식", "가정 시트의 노란 셀을 바꾸면 모든 트리가 다시 계산된다."),
      ("값으로 들어간 것", "각 시트 11·12행의 선도이자율. 부트스트래핑 결과라 엑셀에서 재현하기 어렵다."),
      ("머리와 스텝", "머리 1·3~10행은 모두 2행(스텝)을 참조한다. 2행 자신도 직전 열 + 1 이라 "
                    "맨 앞 열의 0 하나에서 모든 열이 정해진다. 날짜도 스텝에서 나온다."),
      ("바꿀 수 없는 것", "노드 수와 리픽싱 주기는 격자 구조를 정하므로 앱에서 다시 만들어야 한다."),
      ("옮길 수 없는 것", "상태확장 격자는 재결합하지 않아 엑셀 트리 한 장으로 표현할 수 없다. "
                     "앱에서 상태확장을 골랐다면 이 조서는 경로가중치 근사이고, "
                     "결과 시트 3번 검산 마지막 줄에 앱 계산값과의 차이가 나온다."),
      ("", ""),
      ("시트 순서", ""),
      ("흐름", f"가정 → 01 주가 → 02 전환가격 → … → 09 의사결정 → 10 주계약 → "
              f"11~14 GS → {S15} → 16 부채요소 → 17~20 매도청구권 방법1 → "
              "21~24 방법2 → 이자율곡선 → 상각표 → 결과 → 회계처리"),
      ("도달확률", "02 전환가격이 경로가중치 방법을 쓸 때 참조한다."),
      ("16 부채요소", "전환이 없으면 주가와 무관해 한 줄로 끝난다. 결과 시트가 이 값을 쓴다."),
      ("", ""),
      ("05~09가 순환처럼 보이는 이유", ""),
      ("사실", "05·06은 같은 열의 09를 보지만, 07은 다음 열의 05·06을 본다."),
      ("결과", "오른쪽 열이 먼저 확정되고 왼쪽으로 오므로 고리가 닫히지 않는다."),
      ("확인", "아무 칸에서 F2를 누르면 참조 테두리가 오른쪽이나 같은 열에만 생긴다."),
      ("", ""),
      ("동점을 어떻게 깨는가", ""),
      ("언제 생기나", "리픽싱이 주가로 재설정되는 날에는 전환가액 = 주가이므로 "
                  "전환가치가 정확히 100 이 된다. 같은 날 조기상환금액도 100 이면 값이 같다."),
      ("규칙", "전환은 허용오차(1e-9)만큼 앞설 때만 이긴다. 동점이면 현금(상환)이다. "
             "지분으로 보면 무위험이자율로 할인되어 값이 올라가므로 현금 쪽이 보수적이다."),
      ("왜 정해야 하나", "정해 두지 않으면 부동소수 잡음이 갈라 놓는다. "
                    "TF 는 지분과 부채를 다른 이자율로 할인하므로 한 노드의 판정이 "
                    "전체 값을 몇 포인트씩 움직인다."),
      ("", ""),
      ("주의", ""),
      ("상태확장", "수식 조서는 재결합 격자에서만 만들 수 있다. 앱이 경로가중치로 대체한다."),
      ("매도청구권", "⑰~⑳ 이 혼합할인율(방법1), ㉑~㉔ 가 지분·부채 분리(방법2)다. "
                    "결과 시트에 세 방법이 나란히 나오고, "
                 "가정 시트의 평가방법 값으로 어느 쪽을 적용할지 고른다."),
      ("신용위험 처리", "가정 시트의 TF/GS 플래그가 결과 시트 1번의 '적용' 두 행을 고른다. "
                       "주계약과 부채요소는 전환이 없어 두 모형이 같다."),
      ("이자율곡선", "각 트리 11·12행 선도이자율의 출처다. 부트스트래핑 표까지 남긴다."),
      ("상각표", "유효이자율만 역산 결과라 값이고, 상각 내역은 수식이다. "
                "기말 잔액이 만기상환금액과 맞는지 마지막 줄에서 검산한다."),
      ("검산", "결과 시트 3번과 상각표 마지막 줄을 먼저 보고 모두 적합인지 확인한다.")]
    r = 4
    for a2, b3 in ex:
        if a2 and not b3: sec(H, r, a2, span=2)
        elif a2:
            put(H, r, 2, a2, bold=True, size=9); put(H, r, 3, b3, size=9)
        r += 1
    for i in range(4, r):
        H.cell(row=i, column=3).alignment = Alignment(horizontal="left", vertical="center")

    polish_wb(wb)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()


# ══════════════════════════════════════════════════════════
# 6. 화면
# ══════════════════════════════════════════════════════════
st.set_page_config(page_title="전환사채 평가", layout="wide")
st.markdown("""<style>
.block-container{padding-top:2.2rem;max-width:1250px}
h1{font-size:1.7rem !important;letter-spacing:-.02em}
[data-testid="stMetricValue"]{font-size:1.9rem}
</style>""", unsafe_allow_html=True)

st.title("전환사채 평가")
st.caption("계약조건과 시장자료를 넣으면 이항격자로 옵션을 분리해 계산하고 조서를 엑셀로 내보냅니다. "
           "금액은 전자등록금액 100 기준입니다.")

if "tm" not in st.session_state: st.session_state.tm = Terms()
if "prices" not in st.session_state: st.session_state.prices = []
if "peers" not in st.session_state: st.session_state.peers = []
if "rate_series" not in st.session_state: st.session_state.rate_series = []

with st.sidebar:
    st.subheader("계약조건")

    up = st.file_uploader("시나리오 불러오기", type=["json"], key="scen")
    if up is not None:
        try:
            o = json.load(up)
            st.session_state.tm = Terms(**{k: v for k, v in o.items()
                                           if k in Terms.__dataclass_fields__})
            st.success("불러왔습니다.")
        except Exception as ex:
            st.error(f"읽지 못했습니다 — {ex}")
    t = st.session_state.tm

    with st.expander("모형", expanded=True):
        t.model = st.selectbox("신용위험 처리", ["TF", "GS"],
                               index=0 if t.model == "TF" else 1,
                               format_func=lambda x: "TF · 값을 쪼갠다" if x == "TF" else "GS · 할인율을 섞는다")
        t.carry = st.selectbox("조정일 아닌 시점", [0, 1, 2, 3], index=t.carry,
                               format_func=lambda i: ["상태확장 (정확)", "경로가중치",
                                                      "확률가중평균", "특정노드선택"][i])
        t.conv_class = st.selectbox("전환권 회계 분류", ["equity", "liability"],
                                    index=0 if t.conv_class == "equity" else 1,
                                    format_func=lambda x: "자본 · 전환권대가를 잔여로"
                                    if x == "equity" else "파생상품부채 · 주계약을 잔여로")
        st.caption("분류에 따라 무엇을 공정가치로 재고 무엇을 잔여로 두는지가 뒤바뀝니다.")

    with st.expander("날짜 · 기간", expanded=True):
        c1, c2 = st.columns(2)
        t.d_issue = c1.date_input("발행일", value=dt.date.fromisoformat(t.d_issue)).isoformat()
        t.d_mat = c2.date_input("만기일", value=dt.date.fromisoformat(t.d_mat)).isoformat()
        t.d_base = st.date_input("평가기준일", value=dt.date.fromisoformat(t.d_base),
                                 help="발행일과 같으면 최초 인식, 뒤면 후속 측정입니다.").isoformat()
        gaps = {"월": 1.0, "2주": 12/26, "주": 12/52}
        gname = min(gaps, key=lambda k: abs(gaps[k]-t.gap_m))
        gname = st.selectbox("노드 간격", list(gaps), index=list(gaps).index(gname))
        t.gap_m = gaps[gname]
        derive(t)
        c3, c4, c5 = st.columns(3)
        c3.metric("경과", f"{t.elapsed_m:.1f}개월")
        c4.metric("잔존", f"{t.T:.2f}년")
        c5.metric("노드", f"{t.n}")
        st.caption("행사 시점은 아래에서 **발행일 기준 개월**로 넣으십시오. "
                   "앱이 평가기준일 기준으로 옮기고, 행사금액도 발행일부터 복리로 붙입니다.")

    with st.expander("기본", expanded=True):
        t.S0 = st.number_input("평가기준일 주가 (원)", value=float(t.S0), step=1.0)
        t.K0 = st.number_input("현재 전환가액 (원)", value=float(t.K0), step=1.0,
                               help="리픽싱이 이미 일어났으면 조정된 값을 넣으십시오.")
        t.cpn = st.number_input("표면이자율 (%)", value=t.cpn*100, step=0.5)/100
        t.ipay = st.number_input("이자 지급주기 (개월)", value=float(t.ipay), step=1.0)
        t.ytm = st.number_input("만기보장수익률 (%)", value=t.ytm*100, step=0.1, format="%.4f")/100
        t.ytm_cmp = int(st.number_input("만기보장 복리 횟수 (연)", value=int(t.ytm_cmp),
                                        step=1, min_value=1, max_value=12,
                                        help="공시 상환율이 분기복리면 4, 반기면 2."))
        t.face_total = st.number_input("전자등록총액 (원)", value=float(t.face_total),
                                       step=1e8, format="%.0f",
                                       help="회계처리 탭의 전액 기준 금액을 계산합니다.")
        st.caption(f"만기상환금액 = {100*(1+accrue_rate(t.T+t.elapsed_m/12, t.ytm, t.cpn, t.ytm_cmp)):,.4f}   "
                   "공시 만기상환율과 대조하십시오.")

    with st.expander("전환 · 조정"):
        st.caption("모두 **발행일 기준 개월**입니다. 계약서 그대로 넣으십시오.")
        t.cv_s = st.number_input("전환 시작 (개월)", value=float(t.cv_s), step=1.0)
        t.cv_e = st.number_input("전환 종료 (개월)", value=float(t.cv_e), step=1.0)
        t.rfx_mode = st.selectbox("조정 방식", [2, 1, 0], index=[2, 1, 0].index(t.rfx_mode),
                                  format_func=lambda i: ["조정 없음", "하향만", "하향 + 상향"][i])
        t.rfx_cyc = st.number_input("조정 주기 (개월)", value=float(t.rfx_cyc), step=1.0)
        t.floor = st.number_input("최저 조정가액 (원)", value=float(t.floor), step=1.0)
        t.par = st.number_input("액면가 (원)", value=float(t.par), step=100.0)

    with st.expander("조기상환청구권"):
        t.p_s = st.number_input("시작 (개월)", value=float(t.p_s), step=1.0, key="ps")
        t.p_e = st.number_input("종료 (개월)", value=float(t.p_e), step=1.0, key="pe")
        t.p_f = st.number_input("주기 (개월)", value=float(t.p_f), step=1.0, key="pf")
        t.p_mode = st.selectbox("행사금액 산정", ["fixed", "accrue"],
                                index=0 if t.p_mode == "fixed" else 1,
                                format_func=lambda x: "고정률" if x == "fixed" else "보장수익률 복리")
        if t.p_mode == "fixed":
            t.p_rate = st.number_input("행사금액 (%)", value=float(t.p_rate), step=1.0)
        else:
            t.p_yield = st.number_input("조기상환 보장수익률 (%)", value=t.p_yield*100, step=0.5)/100
            t.p_cmp = int(st.number_input("복리 횟수 (연)", value=int(t.p_cmp), step=1, min_value=1))
            st.caption("행사금액 = 100 × (1 + 실효수익률)^경과연수")

        st.divider()
        st.markdown("**평가 방법**")
        _ok = (t.conv_class == "equity" and t.model == "TF" and t.p_s <= t.p_e)
        t.put_bdt = int(st.checkbox(
            "BDT 금리격자로 평가", value=bool(t.put_bdt), disabled=not _ok,
            help="전환을 끄면 격자가 주가와 무관해져 조기상환권이 확정 계산이 "
                 "됩니다. 금리를 확률변수로 두면 옵션의 시간가치가 생깁니다."))
        if not _ok:
            st.caption("전환권을 **자본**으로 두고 **TF** 를 쓸 때만 켤 수 있습니다. "
                       "자본이면 전환권대가가 잔여라 부채요소만 바꿔도 배분이 "
                       "성립하지만, 부채이면 복합내재파생을 전체로서 재야 해서 "
                       "전체 가치까지 함께 손봐야 합니다.")
        elif t.put_bdt:
            # 키를 두지 않는다. 키가 있으면 위젯이 저장해 둔 값이 value 를
            # 이겨서, 아래 「이 변동성 적용」 도 시나리오 불러오기도 화면에
            # 반영되지 않는다. 주가 변동성 칸도 같은 이유로 키가 없다.
            t.bdt_sig = st.number_input("단기이자율 변동성 (%)",
                                        value=t.bdt_sig*100, step=1.0,
                                        min_value=0.0)/100
            t.bdt_base = st.selectbox(
                "기준 곡선", [0, 1], index=int(t.bdt_base),
                format_func=lambda x: ("위험 곡선에 직접" if x == 0
                                       else "무위험 + 확정 스프레드"))
            if t.bdt_base == 0:
                st.caption("단기이자율이 곧 위험이자율입니다. 변동성이 신용스프레드 "
                           "변동까지 안고 갑니다. 옵션 없는 사채가 격자의 주계약과 "
                           "정확히 같아져 검산이 쉽습니다.")
            else:
                st.caption("국고채에 변동성을 태우고 구간 선도 스프레드를 확정으로 "
                           "얹습니다. 변동성을 국고채에서 관측한 값으로 쓸 수 있지만, "
                           "**스프레드가 금리와 무관하다고 본 것**이므로 그 한계를 "
                           "조서에 적으십시오.")
            st.caption("로그정규 변동성입니다. 실무에서는 10~30% 를 씁니다. "
                       "0 이면 지금과 같은 값이 나옵니다.")

            st.markdown("**시계열로 σ 산출**")
            st.caption("할인율은 이미 등급보간 → 만기보간을 거칩니다. 변동성도 같은 "
                       "자료·같은 보간에서 나와야 조서가 하나로 이어집니다.")
            _rmode = st.radio("자료", ["single", "blend"], horizontal=True,
                              key="rvmode",
                              format_func=lambda x: "단일 시계열" if x == "single"
                              else "등급 보간")
            _rt = int(st.number_input("연 거래일수", value=250, step=5,
                                      min_value=30, key="rvtd"))
            _rdrop = st.checkbox("이상치 제거 (MAD 2.5배)", value=True, key="rvdrop")
            _ser, _how = None, ""
            _RVX = ["xls", "xlsx", "xlsm", "csv", "txt", "tsv"]
            if _rmode == "single":
                _f1 = st.file_uploader("금리 시계열 (일자 · 금리)", type=_RVX,
                                       key="rv1")
                if _f1 is not None:
                    try:
                        _ser = parse_prices(read_upload(_f1.name, _f1.getvalue()))
                        _how = f"{_f1.name} · 단일 시계열"
                    except Exception as ex:
                        st.error(str(ex))
            else:
                st.caption("첫 열이 일자이고 머리 줄에 **등급과 만기**가 있으면 "
                           "한 파일에 등급이 여럿이어도 갈라 읽습니다. 등급을 "
                           "따로 받으셨으면 두 번째 칸에 넣으십시오.")
                _fa = st.file_uploader("금리 시계열", type=_RVX, key="rva")
                _fb = st.file_uploader("두 번째 파일 (선택)", type=_RVX, key="rvb")
                _pool, _tens, _seen = {}, {}, False
                for _f in (_fa, _fb):
                    if _f is None: continue
                    _seen = True
                    try:
                        _c, _r = parse_rate_panel(read_upload(_f.name, _f.getvalue()))
                        _s2, _t2 = panel_series(_c, _r, t.T)
                        _pool.update(_s2); _tens.update(_t2)
                    except Exception as ex:
                        st.error(f"{_f.name} — {ex}")
                if _seen and not _pool:
                    st.error("등급이나 만기를 찾지 못했습니다. 머리 줄에 "
                             "`… / BBB0` 같은 등급이나 `5년` 같은 만기가 "
                             "있어야 합니다. 자료는 10줄 이상이어야 합니다.")
                elif _pool:
                    # 고를 수 있는 것은 **파일에 실제로 있는 등급**뿐이다.
                    # 고시표에 없는 등급을 곡선으로 고르면 붙일 자료가 없다.
                    _opt = sorted(_pool, key=lambda r: (r is None, rating_idx(r)))
                    _nm = lambda r: (r or "등급 미상") + (
                        f" ({'·'.join(f'{x:g}년' for x in _tens.get(r, []))})"
                        if _tens.get(r) else "")
                    st.success("찾은 등급 — " + " · ".join(_nm(r) for r in _opt))
                    g1, g2 = st.columns(2)
                    _ia = _opt.index(t.rt_a) if t.rt_a in _opt else 0
                    _ra = g1.selectbox("곡선 A", _opt, index=_ia,
                                       format_func=_nm, key="rvga")
                    _rest = [x for x in _opt if x != _ra]
                    _ib = (_rest.index(t.rt_b) + 1) if t.rt_b in _rest else 0
                    _rb = g2.selectbox("곡선 B", [None] + _rest, index=_ib,
                                       format_func=lambda r: "(없음)" if r is None
                                       else _nm(r), key="rvgb")
                    _rtg = st.selectbox(
                        "평가대상", RATINGS, key="rvgt",
                        index=RATINGS.index(t.rt_tgt) if t.rt_tgt in RATINGS else 8,
                        help="두 곡선 사이면 내삽, 밖이면 같은 기울기로 외삽합니다.")
                    if _rb is None or _ra is None:
                        _ser = _pool[_ra]
                        _how = f"{_nm(_ra)} 단일 · 고정만기 {t.T:.2f}년"
                        st.info("곡선이 하나뿐이라 등급 보간 없이 그대로 씁니다.")
                    else:
                        _ser = blend_series(_pool[_ra], _pool[_rb], _ra, _rb, _rtg)
                        _ja, _jb = rating_idx(_ra), rating_idx(_rb)
                        _w = ((rating_idx(_rtg)-_ja)/(_jb-_ja)) if _ja != _jb else 0.0
                        _how = (f"{_ra}·{_rb} → {_rtg} (가중치 {_w:.2f}) · "
                                f"고정만기 {t.T:.2f}년")
                        st.caption(f"가중치 — {_ra} {1-_w:.0%} · {_rb} {_w:.0%}")
                        if not 0 <= _w <= 1:
                            st.warning(
                                f"평가대상 **{_rtg}** 가 두 곡선 **밖**입니다 "
                                f"(가중치 {_w:.2f}). 등급 간 스프레드는 아래로 "
                                "갈수록 가속해서 벌어지므로, 직선으로 뻗는 외삽은 "
                                "금리를 낮게 잡습니다. 평가대상을 사이에 끼우는 "
                                "등급을 받아 오시는 편이 낫습니다.")
                    _tn = _tens.get(_ra) or []
                    if len(_tn) == 1 and abs(_tn[0] - t.T) > 0.5:
                        st.warning(f"파일의 만기가 **{_tn[0]:g}년** 한 열뿐인데 "
                                   f"잔존만기는 **{t.T:.2f}년** 입니다. 만기 보간을 "
                                   "할 수 없으므로 그 차이를 조서에 적으십시오.")
            if _ser and len(_ser) >= 10:
                _v = rate_vol(_ser, _rt, _rdrop)
                if _v:
                    st.session_state.rate_series = _ser
                    st.session_state.rate_how = _how
                    st.session_state.rate_opt = dict(tdays=_rt, drop=_rdrop)
                    st.metric("상대 변동성 (BDT 의 σ)", f"{_v['annual']*100:.2f}%",
                              f"절대 {_v['abs_annual']:.3f}%p")
                    st.caption(f"평균 금리 {_v['mean']:.3f}% · 관측 {_v['n']}개 · "
                               f"{_ser[0][0]} ~ {_ser[-1][0]}"
                               + (f" · 이상치 {_v['removed']}개 제거" if _v['removed'] else "")
                               + f"  ·  절대 ÷ 평균 = {_v['abs_annual']/max(_v['mean'],1e-9)*100:.2f}%")
                    if _v["neg"]:
                        st.error(f"0 이하인 금리가 {_v['neg']}개 있습니다. 로그를 쓸 수 "
                                 "없어 그 구간이 빠집니다.")
                    if _v["min"] < 1.0:
                        st.warning(f"최저 금리가 {_v['min']:.3f}% 입니다. 저금리 구간에서는 "
                                   "작은 변동도 로그로는 크게 잡혀 σ 가 부풀어 오릅니다.")
                    _exp = 0 if t.bdt_base else 1
                    st.caption("기준 곡선이 **"
                               + ("무위험 + 확정 스프레드" if t.bdt_base else "위험 곡선 직접")
                               + "** 이므로 "
                               + ("국고채" if t.bdt_base else "회사채(평가대상 등급)")
                               + " 시계열을 쓰셔야 맞습니다.")
                    if st.button("이 변동성 적용", use_container_width=True,
                                 type="primary", key="rvapply"):
                        t.bdt_sig = _v["annual"]
                        st.rerun()

    with st.expander("매도청구권"):
        t.k_s = st.number_input("시작 (개월)", value=float(t.k_s), step=1.0, key="ks")
        t.k_e = st.number_input("종료 (개월)", value=float(t.k_e), step=1.0, key="ke")
        t.k_f = st.number_input("주기 (개월)", value=float(t.k_f), step=1.0, key="kf")
        t.k_prem = st.number_input("프리미엄 (연 %)", value=t.k_prem*100, step=0.5)/100
        t.k_cmp = int(st.number_input("복리 횟수 (연)", 1, 12, int(t.k_cmp), 1,
                                      help="분기복리 4 · 반기 2 · 연 1. 계약서의 "
                                           "매수대금 표와 맞는지 확인하십시오."))
        t.k_w = st.number_input("행사 한도 (%)", value=t.k_w*100, step=5.0)/100
        t.k_lock = st.number_input("의무보유 전환지연 (개월)", value=float(t.k_lock), step=1.0)
        t.k_sep = 1 if st.selectbox(
            "회계 처리", ["별도 금융상품", "복합내재파생에 포함"],
            index=0 if t.k_sep else 1,
            help="발행회사가 지정하는 제3자가 살 수 있으면 거래상대방이 달라지므로 "
                 "별도의 금융상품입니다 (기준서 1109 문단 4.3.1). 발행회사만 "
                 "행사할 수 있으면 내재파생상품이라 전환권·조기상환권과 하나로 "
                 "묶습니다 (문단 B4.3.4). 주계약과 전환권대가는 어느 쪽이든 같고 "
                 "파생을 총액으로 볼지 순액으로 볼지가 다릅니다."
            ) == "별도 금융상품" else 0
        t.k_method = st.selectbox("평가방법", [0, 1, 2],
                                  index=[0, 1, 2].index(t.k_method),
                                  format_func=lambda i: K_METHODS[i])
        if t.k_method:
            st.caption("발행회사가 **지정하는 제3자**도 행사할 수 있는 콜옵션은 별도의 "
                       "금융상품이고 기초자산이 전환사채인 복합옵션입니다 "
                       "(기준서 1109 문단 4.3.1). 기초자산에서 **의무보유는 빠집니다.**")
        else:
            st.caption("콜을 넣고 뺀 두 평가액의 차이로 봅니다. 의무보유 효과가 콜 값에 "
                       "포함됩니다.")

    with st.expander("변동성", expanded=True):
        c1, c2 = st.columns([2, 1])
        code = c1.text_input("종목코드 · 티커", value="057680",
                             help="국내는 6자리 숫자, 해외는 티커")
        mkt = c2.selectbox("시장", ["KQ", "KS", ""], index=0,
                           format_func=lambda x: {"KQ": "코스닥", "KS": "코스피", "": "해외"}[x])
        c3, c4 = st.columns(2)
        pdays = int(c3.number_input("조회 일수", value=250, step=10, min_value=30))
        tdays = int(c4.number_input("연 거래일수", value=250, step=5))
        st.caption("야후 파이낸스 수정주가를 씁니다. 유상증자·액면분할·배당이 반영된 종가입니다.")
        asof = st.date_input("조회 종료일", value=dt.date.today())
        drop = st.checkbox("이상치 제거 (중앙값 절대편차 2.5배)", value=True,
                           help="MAD × 1.4826 × 2.5 밖의 일간수익률을 뺍니다. "
                                "책 사례 5-2 와 같은 배수입니다.")
        if st.button("주가 수집", use_container_width=True, type="secondary"):
            with st.spinner("받는 중"):
                try:
                    px, src = fetch_prices(code.strip(), pdays, mkt, asof.isoformat())
                    st.session_state.prices = px
                    st.session_state.px_src = src
                    st.success(f"{src} · {len(px)}개 · {px[0][0]} ~ {px[-1][0]}")
                except Exception as ex:
                    st.error(f"받지 못했습니다 — {ex}\n\n"
                             "종목코드와 시장을 확인하시거나 아래에서 파일을 넣으십시오.")
        pf = st.file_uploader("주가 파일 (엑셀 · csv · txt)",
                              type=["xlsx", "xlsm", "xls", "csv", "txt", "tsv"], key="pxf")
        if pf is not None:
            try:
                rows = parse_prices(read_upload(pf.name, pf.getvalue()))
            except Exception as ex:
                st.error(str(ex))
            else:
                if len(rows) >= 10:
                    st.session_state.prices = rows
                    st.session_state.px_src = pf.name
                    st.success(f"{pf.name} · {len(rows)}개")
                else:
                    st.error(f"종가를 {len(rows)}개밖에 찾지 못했습니다. "
                             "머리글에 '종가' 또는 'Close' 가 있는지 확인하십시오.")
        if st.session_state.prices:
            v = vol_from(st.session_state.prices, tdays, drop)
            if v:
                st.metric("연 변동성", f"{v['annual']*100:.2f}%",
                          f"일 {v['daily']*100:.2f}%")
                st.caption(f"{st.session_state.get('px_src','')} · 수익률 {v['n']}개"
                           + (f" · 이상치 {v['removed']}개 제거 "
                              f"(정상범위 {v['lo']*100:.2f}% ~ {v['hi']*100:.2f}%)"
                              if v['removed'] else ""))
                if st.button("이 변동성 적용", use_container_width=True, type="primary"):
                    t.sig = v["annual"]
                    st.rerun()

        st.divider()
        st.markdown("**비상장 — 피어로 산출**")
        st.caption("대상회사 주가가 없으면 유사기업 여럿의 변동성을 모아 씁니다. "
                   "업종·규모·상장기간이 비슷한 회사를 고르고, 왜 골랐는지 조서에 남기십시오.")
        ptxt = st.text_area("피어 목록 — 한 줄에 하나, `코드` 또는 `코드,이름`",
                            value=st.session_state.get("peer_txt", ""),
                            height=90, placeholder="122870,와이지엔터\n035900,JYP\n041510,SM")
        st.session_state.peer_txt = ptxt
        pc1, pc2 = st.columns(2)
        pmkt = pc1.selectbox("피어 시장", ["KQ", "KS", ""], index=0, key="pmkt",
                             format_func=lambda x: {"KQ": "코스닥", "KS": "코스피",
                                                    "": "해외"}[x])
        vpick = pc2.selectbox("종합 방법", ["median", "mean", "max", "min"],
                             format_func=lambda x: {"median": "중앙값", "mean": "단순평균",
                                                    "max": "최댓값", "min": "최솟값"}[x])
        if st.button("피어 주가 수집", use_container_width=True):
            got, fail = [], []
            with st.spinner("받는 중"):
                for line in ptxt.splitlines():
                    line = line.strip()
                    if not line: continue
                    parts = [x.strip() for x in line.replace("\t", ",").split(",")]
                    code = parts[0]
                    nm = parts[1] if len(parts) > 1 and parts[1] else code
                    try:
                        rows, _src = fetch_prices(code, pdays, pmkt, asof.isoformat())
                        got.append((nm, rows))
                    except Exception as ex:
                        fail.append(f"{nm} — {ex}")
            st.session_state.peers = got
            if got: st.success(f"{len(got)}개 수집 · " + " · ".join(n for n, _ in got))
            if fail: st.error("못 받은 것: " + " / ".join(fail))
        mf = st.file_uploader("여러 종목 종가 파일 (첫 열 일자, 나머지 열 종목)",
                              type=["xlsx", "xlsm", "csv", "txt", "tsv"], key="mpxf")
        if mf is not None:
            try:
                got = parse_prices_multi(read_upload(mf.name, mf.getvalue()))
            except Exception as ex:
                st.error(str(ex))
            else:
                if got:
                    st.session_state.peers = got
                    st.success(f"{mf.name} · {len(got)}개 — "
                               + " · ".join(n for n, _ in got))
                else:
                    st.error("종목별 종가를 찾지 못했습니다. 첫 줄이 머리글이고 "
                             "첫 열이 일자인지 확인하십시오.")
        peers = st.session_state.get("peers") or []
        if peers:
            pv = [(nm, vol_from(px, tdays, drop)) for nm, px in peers]
            pv = [(nm, x) for nm, x in pv if x]
            if pv:
                ann = sorted(x["annual"] for _, x in pv)
                agg = {"median": (ann[len(ann)//2] if len(ann) % 2
                                  else (ann[len(ann)//2-1]+ann[len(ann)//2])/2),
                       "mean": sum(ann)/len(ann), "max": ann[-1], "min": ann[0]}[vpick]
                st.dataframe(pd.DataFrame(
                    [[nm, x["annual"], x["n"], x["removed"]] for nm, x in pv],
                    columns=["회사", "연 변동성", "수익률", "제외"]).style.format(
                    {"연 변동성": "{:.2%}"}), use_container_width=True, hide_index=True)
                st.metric("피어 종합", f"{agg*100:.2f}%",
                          {"median": "중앙값", "mean": "단순평균",
                           "max": "최댓값", "min": "최솟값"}[vpick])
                if st.button("피어 종합 적용", use_container_width=True, type="primary"):
                    t.sig = agg
        st.session_state.vol_opt = dict(tdays=tdays, drop=drop, pick=vpick,
                                        asof=asof.isoformat())
        t.sig = st.number_input("변동성 (%)", value=t.sig*100, step=0.5)/100

    with st.expander("이자율", expanded=True):
        st.caption("무위험·위험 모두 만기수익률(YTM) 곡선을 넣습니다. "
                   "앱이 선형보간 → 부트스트래핑 → 선도이자율 순으로 처리합니다.")
        t.y_type = st.selectbox("입력 유형", ["par", "spot"],
                                index=0 if t.y_type == "par" else 1,
                                format_func=lambda x: "만기수익률 (YTM)" if x == "par"
                                else "현물이자율 (zero rate)")
        if t.y_type == "spot":
            st.caption("아래 **이표 횟수**를 고시된 곡선의 **복리 횟수**로 맞추십시오. "
                       "회사채 제로커브가 분기복리인데 연복리로 두면 할인계수가 어긋납니다.")
        unit = st.selectbox("만기 단위", ["auto", "month", "year"], index=0,
                            format_func=lambda x: {"auto": "자동 인식", "month": "개월",
                                                   "year": "년"}[x])
        cc1, cc2 = st.columns(2)
        t.cmp_rf = int(cc1.number_input("무위험 이표 (연 회)", value=int(t.cmp_rf),
                                        step=1, min_value=1, max_value=12))
        t.cmp_cr = int(cc2.number_input("위험 이표 (연 회)", value=int(t.cmp_cr),
                                        step=1, min_value=1, max_value=12))
        st.caption("국고채는 6개월 이표(2회), 회사채는 3개월 이표(4회)가 발행 관행입니다. "
                   "부트스트래핑에서 현금흐름 시점을 잡는 데 쓰입니다.")
        # ── KIS-Net 기준수익률 표에서 바로 채우기 ──
        kf = st.file_uploader("KIS-Net 기준수익률 표 (선택)", type=["xls", "xlsx"],
                              key="kisnet",
                              help="채권시가평가 기준수익률 표를 그대로 올리면 "
                                   "무위험·위험 곡선을 골라 아래 칸에 채웁니다.")
        if kf is not None:
            try:
                _rows = read_kisnet(kf.name, kf.getvalue())
            except Exception as ex:
                st.error(f"읽지 못했습니다 — {ex}")
            else:
                _lbl = [x[0] for x in _rows]

                def _first(*cands):
                    """앞에서부터 걸리는 첫 줄. 0 번이 정답일 수 있어 None 으로 가른다."""
                    for ws in cands:
                        for i, L in enumerate(_lbl):
                            if all(w in L for w in ws): return i
                    return 0

                st.success(f"{len(_rows)}개 곡선을 찾았습니다.")
                k1, k2 = st.columns(2)
                _ri = k1.selectbox(
                    "무위험 곡선 (국공채)", range(len(_rows)),
                    index=_first(("국채", "국고채권"), ("국채",)),
                    format_func=lambda i: _lbl[i])
                _ci = k2.selectbox(
                    "위험 곡선 (신용위험 반영)", range(len(_rows)),
                    index=_first(("회사채 II", "무보증"), ("회사채", "무보증"), ("회사채",)),
                    format_func=lambda i: _lbl[i])
                st.caption("사모 CB 는 **회사채 II(사모사채)** 줄이 성격에 가깝습니다. "
                           "무등급이면 추정 등급을 고르고 근거를 조서에 남기십시오.")
                if st.button("이 곡선 적용", use_container_width=True, type="primary"):
                    st.session_state.rf_txt = curve_text(_rows[_ri][1])
                    st.session_state.cr_txt = curve_text(_rows[_ci][1])
                    st.session_state.kis_src = (_lbl[_ri], _lbl[_ci])
                    st.rerun()
                # 등급 보간 칸이 같은 표에서 곡선을 고를 수 있도록 남긴다
                st.session_state.kis_rows = _rows
        if st.session_state.get("kis_src"):
            st.caption("적용된 곡선 — 무위험 **%s** · 위험 **%s**"
                       % st.session_state["kis_src"])

        # 두 곡선 모두 같은 형식이다 — 만기(개월) 다음에 수익률(%).
        # 날짜 열이 앞에 붙어 있어도 그대로 읽는다.
        st.caption("형식은 두 곡선이 같습니다 — 한 줄에 **만기 · 수익률**. "
                   "고시표를 날짜 열까지 통째로 붙여 넣어도 날짜는 알아서 버립니다.")
        if "rf_txt" not in st.session_state:
            st.session_state.rf_txt = ("3\t2.40%\n6\t2.38%\n12\t2.25%\n"
                                       "24\t2.33%\n36\t2.34%\n60\t2.50%")
        if "cr_txt" not in st.session_state:
            st.session_state.cr_txt = ("12\t5.10%\n24\t5.70%\n36\t6.20%\n"
                                       "48\t6.65%\n60\t7.05%")
        rf_txt = st.text_area("무위험 곡선 (국공채 YTM)", key="rf_txt", height=130)
        t.rf_curve = parse_yields(rf_txt, unit)
        _MODES = ["pick", "direct", "rating"]
        _MODE_NM = {"pick": "표에서 등급 하나 고르기",
                    "direct": "YTM 직접 입력",
                    "rating": "두 등급 곡선으로 보간"}
        if t.rate_mode not in _MODES: t.rate_mode = "direct"
        t.rate_mode = st.selectbox("위험 곡선", _MODES,
                                   index=_MODES.index(t.rate_mode),
                                   format_func=lambda x: _MODE_NM[x],
                                   help="고시표를 올리셨으면 **등급 하나 고르기**가 "
                                        "가장 짧습니다. 평가대상 등급이 표에 없을 "
                                        "때만 두 등급 보간을 쓰십시오.")
        _kr = st.session_state.get("kis_rows") or []
        _kg = [(i, rating_in(L)) for i, (L, _) in enumerate(_kr)]
        _kg = [(i, g) for i, g in _kg if g]
        _pick = sorted({g for _, g in _kg}, key=rating_idx)

        if t.rate_mode == "pick":
            # 고시표의 한 줄을 그대로 위험 곡선으로 쓴다. 텍스트 칸을 거치지
            # 않으므로 "고른 등급과 실제로 쓰인 곡선이 다른" 사고가 없다.
            if not _kr:
                st.info("고시표를 아직 올리지 않으셨습니다. 위에서 KIS-Net 표를 "
                        "올리시거나, **YTM 직접 입력**으로 바꾸십시오. 그동안은 "
                        "아래 직접 입력 칸을 씁니다.")
                cr_txt = st.text_area("위험 곡선 (등급별 회사채 YTM)",
                                      key="cr_txt", height=130)
                t.cr_curve = parse_yields(cr_txt, unit); t.cr_curve_b = []
                t.cr_src = "직접 입력"
            else:
                _idx = [i for i, _ in _kg] or list(range(len(_kr)))
                _prev = next((i for i in _idx if _kr[i][0] == t.cr_src), _idx[0])
                _pi = st.selectbox("위험 곡선으로 쓸 줄", _idx,
                                   index=_idx.index(_prev),
                                   format_func=lambda i: _kr[i][0])
                t.cr_curve = list(_kr[_pi][1]); t.cr_curve_b = []
                t.cr_src = _kr[_pi][0]
                _g = rating_in(t.cr_src)
                if _g: t.rt_tgt = _g
                st.success(f"**{t.cr_src}** 을 그대로 씁니다 — "
                           + " · ".join(f"{int(round(m*12))}월 {y*100:.2f}%"
                                        for m, y in t.cr_curve[:6])
                           + (" …" if len(t.cr_curve) > 6 else ""))
                st.caption("표의 만기가 그대로 들어갑니다. 아래 직접 입력 칸은 이 "
                           "방식에서는 쓰이지 않습니다.")
        elif t.rate_mode == "direct":
            cr_txt = st.text_area("위험 곡선 (등급별 회사채 YTM)", key="cr_txt", height=130)
            t.cr_curve = parse_yields(cr_txt, unit)
            t.cr_curve_b = []
            t.cr_src = "직접 입력"
        else:
            # 표를 올리셨으면 **그 표에 있는 등급만** 고르게 한다. 고시표에 없는
            # 등급을 곡선으로 고르면 붙여 넣을 자료가 없다 — 무보증 공모사채는
            # 대개 BBB- 까지만 고시한다.
            _opts = _pick or RATINGS
            _at = lambda r, d: _opts.index(r) if r in _opts else min(d, len(_opts)-1)
            r1, r2, r3 = st.columns(3)
            # 표가 없으면 종전 기본값(BBB+ · BBB-), 있으면 가장 위·아래 등급
            t.rt_a = r1.selectbox("곡선 A", _opts, index=_at(t.rt_a, 0 if _pick else 7))
            t.rt_b = r2.selectbox("곡선 B", _opts,
                                  index=_at(t.rt_b, len(_opts)-1 if _pick else 9))
            t.rt_tgt = r3.selectbox(
                "평가대상", RATINGS,
                index=RATINGS.index(t.rt_tgt) if t.rt_tgt in RATINGS else 8,
                help="곡선 두 개 사이면 내삽, 밖이면 같은 기울기로 외삽합니다. "
                     "평가대상은 표에 없어도 고를 수 있습니다.")
            if _pick:
                st.caption("올리신 표에 있는 등급 — **" + " · ".join(_pick)
                           + "**. 두 곡선은 이 안에서만 고를 수 있습니다.")
                if st.button("고른 두 등급으로 채우기", use_container_width=True):
                    _byg = {}
                    for i, g in _kg: _byg.setdefault(g, i)
                    st.session_state.ca_txt = curve_text(_kr[_byg[t.rt_a]][1])
                    st.session_state.cb_txt = curve_text(_kr[_byg[t.rt_b]][1])
                    st.rerun()
            if "ca_txt" not in st.session_state:
                st.session_state.ca_txt = "12\t4.60%\n36\t5.40%\n60\t6.10%"
            if "cb_txt" not in st.session_state:
                st.session_state.cb_txt = "12\t5.80%\n36\t7.20%\n60\t8.30%"
            ca_txt = st.text_area(f"{t.rt_a} 곡선", key="ca_txt", height=100)
            cb_txt = st.text_area(f"{t.rt_b} 곡선", key="cb_txt", height=100)
            t.cr_curve = parse_yields(ca_txt, unit)
            t.cr_curve_b = parse_yields(cb_txt, unit)
            t.cr_src = f"{t.rt_a}·{t.rt_b} 두 등급 보간 → {t.rt_tgt}"
            ia, ib, it2 = rating_idx(t.rt_a), rating_idx(t.rt_b), rating_idx(t.rt_tgt)
            if ia >= 0 and ib >= 0 and it2 >= 0 and ia != ib:
                w = (it2-ia)/(ib-ia)
                if not 0 <= w <= 1:
                    st.warning(f"평가대상 **{t.rt_tgt}** 가 두 곡선 **밖**입니다 "
                               f"(가중치 {w:.2f}). 등급 간 스프레드는 아래로 갈수록 "
                               "가속해서 벌어지므로, 직선으로 뻗는 외삽은 금리를 "
                               "낮게 잡습니다. 평가대상을 사이에 끼우는 등급을 "
                               "받아 오시는 편이 낫습니다.")
                st.caption(f"가중치 — {t.rt_a} {1-w:.0%} · {t.rt_b} {w:.0%}"
                           + ("   (등급 범위 밖이라 외삽합니다)" if not 0 <= w <= 1 else ""))
        cc = credit_curve(t)
        if len(t.rf_curve) < 2 or len(cc) < 2:
            st.error(f"읽힌 줄 — 무위험 {len(t.rf_curve)}개, 위험 {len(cc)}개. "
                     "만기가 다른 값이 각각 두 개 이상 필요합니다.")
        else:
            RFq, CRq = curves(t)
            st.success(f"부트스트래핑 완료 · {t.T:.2f}년 무위험 {math.exp(RFq(t.T))-1:.2%} "
                       f"위험 {math.exp(CRq(t.T))-1:.2%} "
                       f"(스프레드 {math.exp(CRq(t.T))-math.exp(RFq(t.T)):.2%})")

    st.download_button("시나리오 저장",
                       json.dumps(asdict(t), ensure_ascii=False, indent=2).encode(),
                       f"CB평가_시나리오_{dt.date.today()}.json", "application/json",
                       use_container_width=True)

# ── 계산 ──
t = st.session_state.tm
warn = validate(t)
if warn:
    st.warning("확인이 필요합니다\n\n" + "\n".join(f"- {w}" for w in warn))

with st.spinner("계산 중"):
    full, b0, b1, b2, ca, conv = decompose(t)

if not (0 < full["q"] < 1):
    st.error(f"위험중립가중치가 {full['q']:.4f}로 범위를 벗어났습니다. "
             "변동성을 올리거나 노드 수를 늘리십시오.")
    st.stop()

eq = full["GS"]*full["P"] if t.model == "GS" else full["E"]
dv = full["GS"]*(1-full["P"]) if t.model == "GS" else full["B"]

c1, c2, c3 = st.columns([2, 1, 1])
c1.metric(f"전환사채 공정가치 · {t.model}", f"{b2:,.2f}", help="매도청구권 미반영 기준")
c2.metric("지분가치", f"{eq:,.2f}", help="주식으로 받게 될 부분")
c3.metric("부채가치", f"{dv:,.2f}", help="현금으로 받게 될 부분")

tabs = st.tabs(["구성요소", "회계처리", "분리 판단", "이자율곡선", "주가·변동성",
                "의사결정", "상각표", "민감도", "검산", "조서"])

with tabs[0]:
    df = pd.DataFrame([
        ["B0  옵션 없는 사채", b0, None, "—"],
        ["B1  조기상환권 추가", b1, b1-b0, "조기상환청구권"],
        ["B2  전환권 추가", b2, b2-b1, "전환권"],
        ["B3  매도청구권 반영", b2-ca, -ca, f"매도청구권 ({t.k_w*100:.0f}% 한도)"]],
        columns=["단계", "가치", "차액", "해당 옵션"])
    st.dataframe(df.style.format({"가치": "{:,.2f}", "차액": "{:+,.2f}"}, na_rep="—"),
                 use_container_width=True, hide_index=True)
    st.caption("옵션은 서로 대체 관계라 각각 따로 평가해 더하면 총액이 부풀려집니다. "
               "하나씩 얹으며 차액을 보면 합계가 항상 맞습니다.")
    st.dataframe(pd.DataFrame([
        ["TF · 값을 쪼갠다", full["TF"], full["E"], full["B"], None],
        ["GS · 할인율을 섞는다", full["GS"], full["GS"]*full["P"], full["GS"]*(1-full["P"]), full["P"]],
        ["차이", full["TF"]-full["GS"], None, None, None]],
        columns=["모형", "전체", "지분", "부채", "전환확률"]).style.format(
        {"전체": "{:,.2f}", "지분": "{:,.2f}", "부채": "{:,.2f}", "전환확률": "{:.4f}"}, na_rep=""),
        use_container_width=True, hide_index=True)
    st.caption("전환확률이 0과 1 사이 중간이면 두 모형이 갈립니다. "
               "한쪽으로 몰리면 사실상 같은 값이 나옵니다."
               if 0.15 < full["P"] < 0.85 else
               "전환확률이 한쪽으로 몰려 두 모형이 사실상 같은 값을 냅니다.")

with tabs[1]:
    alloc_rows, alloc_note = allocate(t, full, b0, b1, b2, ca)
    af = allocate_full(t, alloc_rows)
    st.dataframe(pd.DataFrame(af, columns=["항목", "100 기준", "전액 기준 (원)"]).style.format(
        {"100 기준": "{:,.2f}", "전액 기준 (원)": "{:,.0f}"}),
        use_container_width=True, hide_index=True)
    st.caption(f"전자등록총액 {t.face_total:,.0f}원 기준으로 환산했습니다.")
    st.caption(alloc_note)
    if t.conv_class == "liability":
        deriv = b2 - b0 - ca
        host_acc = 100 - deriv + ca
        je = f"""[최초 인식]
차) 현금                                {100:>10,.2f}
차) 파생상품자산 (매도청구권)             {ca:>10,.2f}
    대) 전환사채 (주계약 · 잔여)              {host_acc:>10,.2f}
    대) 파생상품부채 (내재파생상품)            {deriv:>10,.2f}

[후속 결산]
차) 이자비용                   주계약 × 유효이자율
    대) 전환사채
차) 파생상품평가손익            매 결산 공정가치로 재측정
    대) 파생상품부채
※ 전환권이 부채이므로 주가가 오르면 평가손실이 납니다."""
    else:
        je = f"""[최초 인식]
차) 현금                                {100:>10,.2f}
차) 파생상품자산 (매도청구권)             {ca:>10,.2f}
    대) 전환사채 (주계약)                    {b0:>10,.2f}
    대) 파생상품부채 (조기상환청구권)          {b1-b0:>10,.2f}
    대) 전환권대가 (자본 · 잔여)              {conv:>10,.2f}

[후속 결산]
차) 이자비용                   주계약 × 유효이자율
    대) 전환사채
차) 파생상품평가손익            공정가치 변동분
    대) 파생상품부채
※ 전환권대가는 자본이므로 후속 재측정이 없습니다."""
    st.code(je, language=None)

with tabs[2]:
    st.write("조기상환청구권과 매도청구권을 **주계약과 분리해야 하는지**를 계약 "
             "조항에 근거해 판단하고, 분리한다면 어떤 방법으로 재는지까지 "
             "정리합니다. 아래 문안을 그대로 조서에 옮기실 수 있습니다.")
    st.caption("판단 순서가 정해져 있습니다 — 문단 B4.3.5 말미가 "
               "\"제1032호에 따라 전환채무상품의 자본요소를 분리하기 전에 "
               "내재된 콜옵션이나 풋옵션이 주채무계약과 밀접하게 관련되어 "
               "있는지를 판단한다\" 고 못박습니다.")

    st.markdown("**계약 조항 확인** — 사이드바에 없는 사실만 여기서 받습니다")
    f1, f2 = st.columns(2)
    t.k_third = 1 if f1.checkbox(
        "매도청구권을 제3자에게 지정할 수 있다", value=bool(t.k_third),
        help="공시에 \"발행회사 및 발행회사가 지정하는 자\" 로 적혀 있으면 "
             "해당합니다. 거래상대방이 달라질 수 있어 내재파생상품이 아니라 "
             "별도의 금융상품입니다 (문단 4.3.1 마지막 문장).") else 0
    t.k_transfer = 1 if f2.checkbox(
        "매도청구권을 사채와 독립적으로 양도할 수 있다", value=bool(t.k_transfer),
        help="같은 문단의 다른 갈래입니다. 둘 중 하나만 해당해도 별도의 "
             "금융상품입니다.") else 0
    f3, f4 = st.columns(2)
    t.p_lost_int = 1 if f3.checkbox(
        "조기상환 행사금액이 상실이자 보상 수준이다", value=bool(t.p_lost_int),
        help="잔여기간에 못 받게 된 이자의 현재가치를 보상하는 수준이면 주계약과 "
             "밀접하게 관련되어 있어 분리하지 않습니다 (문단 B4.3.5(5)(나)). "
             "국내 사모 CB 는 대개 해당하지 않습니다.") else 0
    t.fvpl_whole = 1 if f4.checkbox(
        "복합계약 전체를 당기손익-공정가치로 지정했다", value=bool(t.fvpl_whole),
        help="전체를 공정가치로 재면 내재파생을 따로 뗄 이유가 없습니다 "
             "(문단 4.3.3(3)). 실무에서 드뭅니다.") else 0

    # 상각표는 아래 탭에서 만들어지므로 여기서 따로 부른다. 판단이 쓰는 것은
    # 실제로 인식한 배분액에서 상각한 장부금액이다.
    _sp = split_test(t, full, b0, b1, b2, ca,
                     eir_table(t, acc_host(t, full, b0, b1, b2, ca))[1])
    st.divider()

    for _key, _nm in (("put", "조기상환청구권"), ("call", "매도청구권")):
        _d = _sp[_key]
        st.markdown(f"### {_nm}")
        if not _d["있음"]:
            st.info(_d["이유"][0]); continue
        _box = (st.success if _d["결론"] in ("분리", "별도의 금융상품", "묶어서 분리")
                else st.warning)
        _box(f"**{_d['결론']}**　—　" + " ".join(_d["이유"]))
        if _d["근거"]:
            st.caption("근거 · " + " · ".join(_d["근거"]))
        if _d["지표"]:
            st.dataframe(pd.DataFrame(
                [[k2, (f"{v2*100:.1f}%" if k2 == "차이" else
                       ("예" if v2 is True else "아니오" if v2 is False
                        else f"{v2:,.4f}"))] for k2, v2 in _d["지표"].items()],
                columns=["항목", "값"]), use_container_width=True, hide_index=True)
        st.markdown("**평가방법** — " + _d["평가"])

    if not _sp["call"]["설정일치"]:
        st.error("사이드바의 **매도청구권 → 회계 처리** 설정이 위 판정과 "
                 f"어긋납니다. 판정은 **{_sp['call']['결론']}** 인데 설정은 "
                 + ("별도 금융상품" if t.k_sep else "복합내재파생에 포함")
                 + " 입니다. 배분표와 분개가 판정과 다르게 나오므로 사이드바에서 "
                   "맞추십시오.")

    st.divider()
    st.markdown("## 평가방법 — 어떻게 잴 것인가")

    # ── 조기상환권 : 확정 계산으로 충분한가, 금리모형이 필요한가 ──
    st.markdown("### 조기상환청구권 — 금리모형(BDT)을 켤 것인가")
    st.caption("전환을 끄면 격자가 주가와 무관해져 스텝마다 값이 하나뿐입니다. "
               "즉 지금 조기상환권은 **미리 내다보고 액면이 더 크면 행사한다**는 "
               "확정 계산이고, 옵션의 시간가치가 들어 있지 않습니다. "
               "행사가 뻔하면 그래도 맞는 답이 나오지만, 애매하면 값을 0 에 "
               "가깝게 잡습니다. 그 자리가 금리모형이 필요한 자리입니다.")
    if t.p_s <= t.p_e and t.T > 0:
        _r0 = engine(t, conv=False, put=False, call=False)
        _dtx = t.T/int(t.n)
        _lo2, _hi2 = step_mapper(t, int(t.n), _dtx)
        _mp = int(t.n)/(t.T*12)
        _pr = max(1, int(round(t.p_f*_mp)))
        _s2, _e2 = _lo2(t.p_s), _hi2(t.p_e)
        _at2 = {}
        for _k3, _v3 in _r0["memo"].items():
            _at2.setdefault(_k3[0], _v3)
        _rows2, _rat2 = [], []
        for _i3 in range(max(_s2, 0), _e2+1):
            if (_i3-_s2) % _pr or _i3 not in _at2: continue
            _hold = _at2[_i3]["E"] + _at2[_i3]["B"]
            _amt = (100*(1 + accrue_rate(_i3*_dtx + t.elapsed_m/12, t.p_yield,
                                         t.cpn, t.p_cmp))
                    if t.p_mode == "accrue" else t.p_rate)
            _rat2.append(_amt/max(_hold, 1e-9))
            _rows2.append([_i3, round(t.elapsed_m + _i3*_dtx*12), _amt, _hold,
                           _rat2[-1]])
        if _rows2:
            _atm2 = sum(1 for x in _rat2 if 0.97 <= x <= 1.03)
            _otm2 = sum(1 for x in _rat2 if x < 0.97)
            st.dataframe(pd.DataFrame(
                _rows2, columns=["스텝", "발행 후 개월", "행사금액", "계속보유가치",
                                 "행사금액 ÷ 계속보유"]).style.format(
                {"행사금액": "{:,.2f}", "계속보유가치": "{:,.2f}",
                 "행사금액 ÷ 계속보유": "{:.3f}"}),
                use_container_width=True, hide_index=True, height=240)
            st.caption("마지막 열이 **1 보다 크면** 그 날 상환받는 편이 낫다는 뜻입니다. "
                       "행사가 확정적이라 금리를 흔들어도 판단이 안 바뀝니다. "
                       "**1 근처(0.97~1.03)이거나 1보다 작으면** 금리에 따라 판단이 "
                       "갈리므로 확정 계산이 값을 적게 잡습니다.")
            if _otm2 == len(_rat2):
                st.error(f"모든 행사일이 **외가격**입니다 (비율 최대 {max(_rat2):.3f}). "
                         f"지금 모델은 조기상환권을 {b1-b0:,.2f} 로 계산하는데, "
                         "외가격 옵션에도 시간가치가 있습니다. **금리모형 없이는 값을 "
                         "0 에 가깝게 잡습니다.** 사이드바에서 BDT 를 켜십시오.")
            elif _atm2 + _otm2 > 0:
                st.warning(f"등가격 근처가 {_atm2}회, 외가격이 {_otm2}회 있습니다 "
                           f"(비율 {min(_rat2):.3f} ~ {max(_rat2):.3f}). 행사 여부가 "
                           "금리에 따라 갈릴 수 있으므로 BDT 를 켜서 차이를 "
                           "확인하고 그 판단을 조서에 남기십시오.")
            else:
                st.success(f"모든 행사일에서 행사금액이 계속보유가치보다 큽니다 "
                           f"(비율 {min(_rat2):.3f} ~ {max(_rat2):.3f}). 행사가 "
                           "확정적이라 금리를 확률변수로 두어도 판단이 바뀌지 "
                           "않습니다. **확정 격자로 충분합니다.**")
        st.caption("현재 설정 — 조기상환권을 "
                   + ("**BDT 금리격자**로 잽니다." if put_bdt_on(t) else
                      "**금리 고정 격자**로 잽니다.")
                   + ("" if put_bdt_on(t) else
                      "  BDT 는 전환권이 자본이고 TF 일 때만 켤 수 있습니다."))
    else:
        st.info("조기상환청구권이 없어 판단할 것이 없습니다.")

    # ── 매도청구권 : 세 방법을 나란히 ──
    st.markdown("### 매도청구권 — 세 방법 중 무엇으로 잴 것인가")
    if t.k_w > 0:
        _mv = []
        for _km, _lb in ((0, "유무가치비교법"), (1, "옵션차익 · 혼합할인율"),
                         (2, "옵션차익 · 지분·부채 분리")):
            _tk = Terms(**asdict(t)); _tk.k_method = _km; derive(_tk)
            _mv.append([_lb, decompose(_tk)[4], "◀ 적용" if t.k_method == _km else ""])
        st.dataframe(pd.DataFrame(_mv, columns=["방법", "값", "　"]).style.format(
            {"값": "{:,.4f}"}), use_container_width=True, hide_index=True)
        st.caption("**어느 쪽이 옳다기보다 재는 대상이 다릅니다.** 유무가치비교법은 "
                   "콜을 넣고 뺀 차액이라 **의무보유로 잃는 전환권 가치까지** 값에 "
                   "들어갑니다. 옵션차익혼합할인법은 전환사채를 기초자산으로 하는 "
                   "콜옵션 자체만 잽니다. 보고서를 검토하실 때도 어느 방법을 썼는지 "
                   "먼저 확인하셔야 합니다.")
        st.info("판정에 따른 권고 — " + _sp["call"]["평가"].replace("**", ""))
    else:
        st.info("매도청구권이 없어 판단할 것이 없습니다.")

    # ── 금리 민감도로 본 금리모형 실익 ──
    with st.expander("금리모형이 값을 얼마나 바꾸는가 — 민감도로 본 실익"):
        RFc, CRc = curves(t)

        def _bump(**kw):
            tt = Terms(**asdict(t))
            for k2, v2 in kw.items(): setattr(tt, k2, v2)
            derive(tt); return pick(engine(tt, call=False), t.model)
        # BDT 는 금리 "수준" 을 확률변수로 둔다. 두 곡선을 함께 흔들어야 노출을
        # 제대로 잰다 — 무위험만 흔들면 스프레드와 상쇄되어 크게 과소하게 잡힌다.
        _par2 = lambda d: dict(rf_curve=[(x, y+d) for x, y in t.rf_curve],
                               cr_curve=[(x, y+d) for x, y in t.cr_curve],
                               cr_curve_b=[(x, y+d) for x, y in t.cr_curve_b])
        _dl = (_bump(**_par2(0.01)) - _bump(**_par2(-0.01)))/2
        _ds = (_bump(cr_curve=[(x, y+0.01) for x, y in t.cr_curve],
                     cr_curve_b=[(x, y+0.01) for x, y in t.cr_curve_b])
               - _bump(cr_curve=[(x, y-0.01) for x, y in t.cr_curve],
                       cr_curve_b=[(x, y-0.01) for x, y in t.cr_curve_b]))/2
        _dv = (_bump(sig=t.sig+0.10) - _bump(sig=max(0.01, t.sig-0.10)))/2
        _ratio = abs(_dl)/max(abs(_dv), 1e-9)
        _spr = CRc(t.T) - RFc(t.T)
        _share = _spr/CRc(t.T) if CRc(t.T) > 1e-9 else 0.0
        st.dataframe(pd.DataFrame([
            ["금리 수준 ±1%p (두 곡선 평행)", f"{_dl:+,.4f}",
             f"{abs(_dl)/max(b2,1e-9)*100:.2f}%"],
            ["신용스프레드 ±1%p (위험 곡선만)", f"{_ds:+,.4f}",
             f"{abs(_ds)/max(b2,1e-9)*100:.2f}%"],
            ["변동성 ±10%p", f"{_dv:+,.4f}", f"{abs(_dv)/max(b2,1e-9)*100:.2f}%"],
            ["금리 수준 ÷ 주가 민감도", f"{_ratio:.3f}", ""],
            [f"{t.T:.2f}년 신용스프레드", f"{_spr*100:.2f}%p",
             f"할인율 중 {_share*100:.0f}%"]],
            columns=["항목", "값", "비중"]), use_container_width=True,
            hide_index=True)
        if _ratio < 0.2 or _share > 0.7:
            st.success("금리를 확률변수로 둘 실익이 작습니다. 신용스프레드가 값을 "
                       "지배하므로 금리 고정 격자로 충분합니다.")
        else:
            st.warning("금리 수준 민감도가 무시할 수준이 아닙니다. BDT 적용 여부를 "
                       "검토하고 그 판단을 조서에 남기십시오.")
        st.caption("두 곡선을 함께 흔드는 이유 — 부채 부분은 위험이자율로 할인되므로 "
                   "무위험 곡선만 흔들면 스프레드 변화와 상쇄되어 노출이 최대 수십 배 "
                   "과소하게 잡힙니다.")

    # ── 회계처리 ──
    st.divider()
    st.markdown("## 회계처리 — 판정대로 배분하면")
    _rows_al, _note_al = allocate(t, full, b0, b1, b2, ca)
    st.dataframe(pd.DataFrame(
        [[k2, v2, fv2] for k2, v2, fv2 in allocate_full(t, _rows_al)],
        columns=["항목", "100 기준", "전액 기준 (원)"]).style.format(
        {"100 기준": "{:,.4f}", "전액 기준 (원)": "{:,.0f}"}),
        use_container_width=True, hide_index=True)
    st.caption(_note_al)
    st.caption("분개는 **회계처리** 탭에 있습니다. 여기서는 판정이 배분에 어떻게 "
               "닿는지만 보입니다.")

    st.divider()
    st.markdown("**조서에 옮길 문안**")
    st.code(split_memo(_sp), language=None)
    st.caption("판단 순서·근거 문단·지표가 함께 들어 있습니다. 결론만 적는 것과 "
               "달리 감사인이 다시 물을 여지를 줄입니다.")

with tabs[3]:
    RF, CR = curves(t)
    dt_ = t.T/t.n
    rows = []
    for k in range(9):
        tt = t.T*k/8
        i = min(t.n-1, round(tt/dt_))
        fr = forward_rate(RF, i*dt_, (i+1)*dt_)
        fc = forward_rate(CR, i*dt_, (i+1)*dt_)
        rows.append([tt, RF(tt), fr, CR(tt), fc, fc-fr])
    cdf = pd.DataFrame(rows, columns=["시점(년)", "무위험 현물", "무위험 선도",
                                      "위험 현물", "위험 선도", "스프레드"])
    st.dataframe(cdf.style.format({"시점(년)": "{:.2f}", "무위험 현물": "{:.2%}",
                                   "무위험 선도": "{:.2%}", "위험 현물": "{:.2%}",
                                   "위험 선도": "{:.2%}", "스프레드": "{:.2%}"}),
                 use_container_width=True, hide_index=True)
    st.line_chart(cdf.set_index("시점(년)")[["무위험 현물", "위험 현물", "위험 선도"]])
    if len(t.cr_curve) >= 2 and t.y_type == "par":
        st.markdown("**부트스트래핑 과정**")
        dfs = bootstrap_df(t.cr_curve, t.T, t.cmp_cr)
        bt = pd.DataFrame([[d[0], _lin(t.cr_curve, d[0]), d[1], -math.log(d[1])/d[0]]
                           for d in dfs if d[0] > 0],
                          columns=["만기(년)", "만기수익률", "할인계수", "현물이자율(연속)"])
        st.dataframe(bt.style.format({"만기(년)": "{:.2f}", "만기수익률": "{:.2%}",
                                      "할인계수": "{:.6f}", "현물이자율(연속)": "{:.4%}"}),
                     use_container_width=True, hide_index=True, height=260)
        st.caption("각 이표 시점마다 1 = 이자 × 앞선 할인계수 합 + 그 시점 할인계수 를 풀어 "
                   "할인계수를 앞에서부터 순차로 구합니다. 현물이자율은 −LN(할인계수) ÷ 만기입니다.")
    step_df = math.exp(-sum(forward_rate(CR, i*dt_, (i+1)*dt_)*dt_ for i in range(t.n)))
    ok = abs(step_df - math.exp(-CR(t.T)*t.T)) < 1e-8
    st.caption(f"검산 — 스텝별 선도이자율을 {t.n}번 곱한 값 {step_df:.8f} 과 "
               f"만기 현물 할인계수 {math.exp(-CR(t.T)*t.T):.8f} 가 "
               + ("일치합니다." if ok else "어긋납니다. 곡선 입력을 확인하십시오."))
    st.caption("선도이자율  f(t, t+Δt) = [ r(t+Δt)×(t+Δt) − r(t)×t ] ÷ Δt. "
               "격자의 한 스텝 할인이 이 값을 씁니다.")

with tabs[4]:
    if not st.session_state.prices:
        st.info("왼쪽 변동성 칸에서 주가를 수집하거나 파일을 넣으십시오. "
                "국내 6자리 종목은 한국거래소를, 실패하면 야후 파이낸스를 씁니다.")
    else:
        px = st.session_state.prices
        pxdf = pd.DataFrame(px, columns=["날짜", "종가"])
        c1, c2, c3 = st.columns(3)
        c1.metric("종가 개수", f"{len(px):,}")
        c2.metric("기간", f"{px[0][0]} ~ {px[-1][0]}")
        c3.metric("마지막 종가", f"{px[-1][1]:,.0f}")
        st.caption("출처 " + st.session_state.get("px_src", ""))
        if pxdf["날짜"].iloc[0]:
            st.line_chart(pxdf.set_index("날짜")["종가"])
        v = vol_from(px, 250, True); v0 = vol_from(px, 250, False)
        st.dataframe(pd.DataFrame([
            ["이상치 제거", v["annual"], v["daily"], v["n"]-v["removed"], v["removed"]],
            ["이상치 포함", v0["annual"], v0["daily"], v0["n"], 0]],
            columns=["구분", "연 변동성", "일 변동성", "관측", "제거"]).style.format(
            {"연 변동성": "{:.2%}", "일 변동성": "{:.2%}"}),
            use_container_width=True, hide_index=True)
        st.caption(f"정상범위 {v['lo']*100:.2f}% ~ {v['hi']*100:.2f}% — "
                   "일별 로그수익률의 중앙값에서 중앙값 절대편차의 3배를 벗어난 값을 뺍니다.")
        st.markdown("**최근 10일**")
        st.dataframe(pxdf.tail(10).iloc[::-1].style.format({"종가": "{:,.0f}"}),
                     use_container_width=True, hide_index=True)

with tabs[5]:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.colors import ListedColormap
    from matplotlib.patches import Patch
    # 한글 글꼴이 없으면 글자가 네모로 나온다. 찾으면 쓰고, 없으면 영문으로 그린다.
    _kf = use_korean_font()
    _L = (dict(x="스텝", y="주가 수준",
               lg=["전환", "조기상환", "매도청구", "보유", "만기상환"]) if _kf else
          dict(x="step", y="stock level",
               lg=["Convert", "Put", "Call", "Hold", "Redeem"]))
    n = t.n
    idx = {}
    for k, v in full["memo"].items():
        kk = (k[0], k[1])
        if kk not in idx: idx[kk] = v
    code_map = {"conv": 1, "put": 2, "call": 3, "hold": 4, "mat": 5}
    grid = np.full((n+1, n+1), np.nan)
    for (i, j), o in idx.items():
        if j <= i: grid[n-j, i] = code_map[o["kind"]]
    cmap = ListedColormap(["#1b6b5a", "#7a4b1e", "#a3312a", "#e4e8ec", "#9aa4ae"])
    fig, ax = plt.subplots(figsize=(11, 3.6))
    ax.imshow(grid, aspect="auto", cmap=cmap, vmin=0.5, vmax=5.5, interpolation="nearest")
    ax.set_xlabel(_L["x"]); ax.set_ylabel(_L["y"])
    ax.set_yticks([]); ax.spines[["top", "right"]].set_visible(False)
    ax.legend(handles=[Patch(facecolor=c, label=l) for c, l in
                       zip(["#1b6b5a", "#7a4b1e", "#a3312a", "#e4e8ec", "#9aa4ae"],
                           _L["lg"])],
              loc="upper center", bbox_to_anchor=(0.5, -0.22), ncol=5, frameon=False)
    st.pyplot(fig, use_container_width=True)
    if not _kf:
        st.caption("한글 글꼴이 없어 그림만 영문으로 그렸습니다. 표와 설명은 그대로입니다. "
                   "한글로 보시려면 저장소에 `packages.txt` 를 만들어 `fonts-nanum` "
                   "한 줄을 넣으십시오. 다만 apt 설치가 실패하면 앱이 아예 뜨지 "
                   "않으므로, 넣으신 뒤 재시작이 되는지 확인하십시오.")
    D = full["dist"]; tot = D["conv"]+D["put"]+D["call"]+D["mat"] or 1
    st.dataframe(pd.DataFrame([
        ["전환", D["conv"]/tot, D["tc"]/D["conv"]/full["mper"] if D["conv"] else None],
        ["조기상환", D["put"]/tot, D["tp"]/D["put"]/full["mper"] if D["put"] else None],
        ["매도청구", D["call"]/tot, D["tk"]/D["call"]/full["mper"] if D["call"] else None],
        ["만기 상환", D["mat"]/tot, t.T*12]],
        columns=["유형", "비중", "평균 시점(개월)"]).style.format(
        {"비중": "{:.1%}", "평균 시점(개월)": "{:,.1f}"}, na_rep="—"),
        use_container_width=True, hide_index=True)
    st.caption("거의 모든 경로가 만기 전에 끝나면 기대만기가 계약만기보다 짧다는 뜻이고, "
               "장기 할인율의 영향이 줄어듭니다.")

with tabs[6]:
    r_eir, rows_eir, red, nper = eir_table(t, acc_host(t, full, b0, b1, b2, ca))
    st.dataframe(pd.DataFrame([
        ["주계약 (옵션 없는 사채)", f"{b0:,.2f}"], ["만기상환금액", f"{red:,.2f}"],
        ["표면이자 (회당)", f"{100*t.cpn*t.ipay/12:,.2f}"], ["상각 횟수", f"{nper}회"],
        ["유효이자율 (연, 이산복리)", f"{r_eir:.2%}"]], columns=["항목", "값"]),
        use_container_width=True, hide_index=True)
    amdf = pd.DataFrame(rows_eir, columns=["회차", "경과연수", "기초 장부금액",
                                           "이자비용", "지급이자", "기말 장부금액"])
    st.dataframe(amdf.style.format({"경과연수": "{:.2f}", "기초 장부금액": "{:,.2f}",
                                    "이자비용": "{:,.2f}", "지급이자": "{:,.2f}",
                                    "기말 장부금액": "{:,.2f}"}),
                 use_container_width=True, hide_index=True, height=320)
    st.caption("기말 장부금액이 만기에 상환금액과 일치해야 합니다.")

with tabs[7]:
    rows = []
    for dvv in (-0.15, -0.075, 0.0, 0.075, 0.15):
        tt = Terms(**asdict(t)); tt.sig = max(0.01, t.sig+dvv)
        rows.append([tt.sig, pick(engine(tt, call=False), t.model)])
    base = rows[2][1]
    st.dataframe(pd.DataFrame([[r[0], r[1], r[1]-base] for r in rows],
                              columns=["변동성", "전체 가치", "변화"]).style.format(
        {"변동성": "{:.1%}", "전체 가치": "{:,.2f}", "변화": "{:+,.2f}"}),
        use_container_width=True, hide_index=True)
    rows2 = []
    for dvv in (-0.05, -0.025, 0.0, 0.025, 0.05):
        tt = Terms(**asdict(t))
        tt.cr_curve = [(x, y+dvv) for x, y in t.cr_curve]
        tt.cr_curve_b = [(x, y+dvv) for x, y in t.cr_curve_b]
        rows2.append([dvv, pick(engine(tt, call=False), t.model)])
    base2 = rows2[2][1]
    st.dataframe(pd.DataFrame([[r[0], r[1], r[1]-base2] for r in rows2],
                              columns=["할인율 조정", "전체 가치", "변화"]).style.format(
        {"할인율 조정": "{:+.1%}", "전체 가치": "{:,.2f}", "변화": "{:+,.2f}"}),
        use_container_width=True, hide_index=True)
    st.caption("구조에 따라 값을 지배하는 인풋이 다릅니다. "
               "변동성 영향이 거의 없다면 민감도 공시 대상은 할인율이어야 합니다.")

with tabs[8]:
    st.write("**계산이 성립하는지**만 봅니다. 무엇을 분리하고 어떻게 잴지는 "
             "「분리 판단」 탭으로 옮겼습니다.")
    imm = 100*t.S0/t.K0
    tot_al = allocate(t, full, b0, b1, b2, ca)[0][-1][1]
    checks = [("위험중립가중치 q", f"{full['q']:.4f}", 0 < full["q"] < 1),
              ("상승계수 u", f"{full['u']:.4f}", full["u"] > 1),
              ("전체 가치 ≥ 순수사채가치", f"{b2:,.2f} ≥ {full['host']:,.2f}",
               b2 >= full["host"]-1e-6),
              ("전체 가치 ≥ 즉시 전환가치", f"{b2:,.2f} ≥ {imm:,.2f}",
               not (t.cv_s <= 0 and b2 < imm-1e-6)),
              ("배분 합계 = 100", f"{tot_al:,.2f}", abs(tot_al-100) < 0.01)]
    st.dataframe(pd.DataFrame([[k, v, "적합" if ok else "확인 필요"] for k, v, ok in checks],
                              columns=["항목", "값", "판정"]),
                 use_container_width=True, hide_index=True)
    st.caption("위험중립가중치가 0과 1을 벗어나면 변동성이나 노드 수 설정이 잘못된 것입니다.")

    st.markdown("**신용스프레드가 발행조건과 맞는가**")
    st.caption("발행일에는 투자자가 100 을 내고 사채 + 조기상환권 + 전환권을 삽니다. "
               "그러니 전체 가치가 100 이어야 합니다. 크게 벗어나면 인풋이 발행조건과 "
               "어긋난 것이고, 대개 위험할인율(신용스프레드) 추정이 원인입니다. "
               "**전체가 100 이 되는 할인율을 역산해** 넣으신 값과 견줍니다.")
    if t.elapsed_m > 0.01:
        st.info(f"평가기준일이 발행일보다 {t.elapsed_m:.1f}개월 뒤입니다. 그 사이 주가와 "
                "신용도가 바뀌었으므로 전체가 100 을 벗어나는 것이 정상입니다. "
                f"현재 전체 {b2:,.2f}. 역산은 발행일 평가에서만 돌립니다.")
    elif len(t.cr_curve) < 2:
        st.info("위험 곡선을 두 점 이상 넣으셔야 역산할 수 있습니다.")
    else:
        _lv = [y for _, y in t.cr_curve]
        _sh = lambda d: [(x, y+d) for x, y in t.cr_curve]

        def _tot(d):
            tt = Terms(**asdict(t)); tt.cr_curve = _sh(d)
            derive(tt); return decompose(tt)[3]
        try:
            _a, _b = -min(_lv)+1e-4, 0.60          # 곡선을 평행이동할 폭
            if _tot(_b) > 100:
                # 스프레드를 아무리 올려도 못 내려간다 — 사채요소가 아니라
                # 전환조건이 값을 떠받치고 있다는 뜻이다.
                _flr = _tot(_b)
                st.warning(
                    f"신용스프레드를 60%p 올려도 전체가 {_flr:,.2f} 아래로 "
                    f"내려가지 않습니다 (현재 {b2:,.2f}). 사채요소를 거의 0 으로 "
                    "만들어도 그만큼이 남는다는 뜻이므로, **원인은 신용이 아니라 "
                    "전환조건**입니다. 주가 ÷ 전환가액 "
                    f"{t.S0/max(t.K0,1e-9):.2f}, 변동성 {t.sig:.1%}, "
                    f"리픽싱 {'있음' if t.rfx_mode else '없음'} 을 먼저 보십시오. "
                    "메자닌은 투자자에게 유리하게 발행되는 경우가 많아 실제로 "
                    "100 을 넘기도 합니다 — 그러면 그 사실을 조서에 적으면 됩니다.")
            elif _tot(_a) < 100:
                st.warning(
                    f"스프레드를 0 까지 낮춰도 전체가 100 에 못 미칩니다 "
                    f"(현재 {b2:,.2f}). 만기보장수익률이나 행사조건이 빠지지 "
                    "않았는지 확인하십시오.")
            else:
                for _ in range(36):
                    _m = (_a+_b)/2
                    if _tot(_m) > 100: _a = _m
                    else: _b = _m
                _d = (_a+_b)/2
                _t2 = Terms(**asdict(t)); _t2.cr_curve = _sh(_d)
                derive(_t2)
                _f2, _z0, _z1, _z2, _zc, _zv = decompose(_t2)
                _RF2, _CR2 = curves(_t2)
                st.dataframe(pd.DataFrame([
                    ["넣으신 위험이자율 (잔존만기)", f"{CRc(t.T)*100:.2f}%",
                     f"전체 {b2:,.2f}"],
                    ["역산된 위험이자율", f"{_CR2(t.T)*100:.2f}%", "전체 100.00"],
                    ["차이 (곡선 평행이동)", f"{_d*100:+.2f}%p", ""],
                    ["역산 상태의 조기상환권", f"{_z1-_z0:,.2f}",
                     f"현재 {b1-b0:,.2f}"],
                    ["역산 상태의 전환권대가", f"{_zv:,.2f}", f"현재 {conv:,.2f}"]],
                    columns=["항목", "값", "참고"]),
                    use_container_width=True, hide_index=True)
                if abs(b2-100) < 1.0:
                    st.success("전체가 발행가액과 거의 맞습니다. 인풋이 발행조건과 "
                               "정합적입니다.")
                elif b2 > 100:
                    st.warning(
                        f"전체가 발행가액보다 {b2-100:,.2f} 큽니다. 그만큼을 "
                        "투자자가 공짜로 받은 셈이라 발행조건과 맞지 않습니다. "
                        f"**신용스프레드를 {_d*100:.2f}%p 낮게 잡았을 소지가 큽니다.** "
                        "위 조기상환권 진단에서 등가격이 잡혔다면, 역산값을 쓰면 "
                        "내가격으로 돌아서 등가격 문제 자체가 사라지는지 먼저 "
                        "확인하십시오.")
                else:
                    st.warning(
                        f"전체가 발행가액보다 {100-b2:,.2f} 작습니다. 스프레드를 "
                        "높게 잡았거나, 리픽싱·조기상환 같은 조건이 빠졌을 수 "
                        "있습니다.")
        except Exception as _ex:
            st.info(f"역산하지 못했습니다 — {_ex}")
        st.caption("역산값을 그대로 쓰라는 뜻은 아닙니다. 시장에서 관측한 등급 "
                   "수익률을 쓰는 것이 원칙이고, 역산은 **인풋이 발행조건과 얼마나 "
                   "떨어져 있는지 재는 자**입니다. 괴리가 크면 등급 추정이나 "
                   "만기보장수익률 입력을 다시 보십시오.")

with tabs[9]:
    st.write("가정 · 트리 시트 · 이자율곡선 · 결과 · 회계처리 · 상각표로 이루어진 조서를 만듭니다. "
             "트리 하나가 시트 하나이고, 모든 시트의 머리 17행이 같은 형태입니다.")
    kind = st.radio("조서 형식", ["값", "수식"], horizontal=True,
                    format_func=lambda x: "값 조서 — 계산 결과 스냅샷"
                    if x == "값" else "수식 조서 — 엑셀에서 다시 계산됨")
    if kind == "값":
        st.caption("앱이 계산한 값을 그대로 담습니다. 셀을 바꿔도 다시 계산되지 않으므로 "
                   "제출용 조서에 적합합니다. 상태확장을 포함한 모든 설정에서 만들 수 있습니다.")
    else:
        st.caption("가정 시트의 노란 셀을 바꾸면 엑셀 안에서 트리가 다시 계산됩니다. "
                   "선도이자율만 값으로 들어갑니다. 노드 수와 리픽싱 주기는 격자 구조라 바꿀 수 없습니다.")
        if t.carry == 0 and t.rfx_mode > 0:
            st.warning("상태확장은 한 노드에 전환가격이 여럿이라 수식으로 펼 수 없습니다. "
                       "경로가중치로 대체해 만듭니다. 값이 조금 달라집니다.")
    c1, c2 = st.columns([1, 2])
    if c1.button("조서 만들기", type="primary", use_container_width=True):
        try:
            with st.spinner("엑셀 작성 중"):
                if kind == "값":
                    data = build_xlsx(t, full, b0, b1, b2, ca, conv,
                                      eir_table(t, acc_host(t, full, b0, b1, b2, ca)))
                    fn = f"CB평가조서_값_{dt.date.today()}.xlsx"
                else:
                    tf = Terms(**asdict(t))
                    if tf.carry == 0 and tf.rfx_mode > 0: tf.carry = 1
                    ff, f0, f1, f2, fca, fconv = decompose(tf)
                    data = build_xlsx_formula(tf, ff, f0, f1, f2, fca, fconv,
                                              eir_table(tf, acc_host(tf, ff, f0, f1, f2, fca)))
                    fn = f"CB평가조서_수식_{dt.date.today()}.xlsx"
            st.session_state.report = (fn, data)
        except ModuleNotFoundError:
            st.error("openpyxl 이 없습니다.  pip install openpyxl  을 실행하고 다시 시도하십시오.")
        except Exception as ex:
            st.error(f"조서를 만들지 못했습니다 — {ex}")

    rep = st.session_state.get("report")
    if rep:
        fn, data = rep
        st.download_button(f"{fn} 내려받기  ({len(data)/1024:,.0f} KB)", data, fn,
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           type="primary", key="dl_report", use_container_width=True)
        st.caption("버튼이 보이지 않거나 눌러도 반응이 없으면 브라우저의 팝업·다운로드 차단을 확인하십시오.")

    st.divider()
    st.subheader("부속 리포트")
    st.write("평가 인풋이 어디서 나왔는지 보여 주는 산출내역입니다. 조서와 함께 철하면 "
             "감사인이 인풋까지 따라올 수 있습니다. 둘 다 계산이 수식으로 들어갑니다.")
    MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    rc1, rc2 = st.columns(2)
    with rc1:
        st.markdown("**변동성 산출내역**")
        _peers = st.session_state.get("peers") or []
        _own = st.session_state.get("prices") or []
        _opt = st.session_state.get("vol_opt") or {}
        if _peers:
            st.caption(f"피어 {len(_peers)}개로 만듭니다 — "
                       + " · ".join(n for n, _ in _peers))
        elif _own:
            st.caption(f"대상회사 종가 {len(_own)}개로 만듭니다 "
                       f"({st.session_state.get('px_src', '')}).")
        else:
            st.caption("먼저 왼쪽 변동성 칸에서 주가를 수집하거나 파일을 넣으십시오.")
        if st.button("변동성 리포트 만들기", use_container_width=True,
                     disabled=not (_peers or _own)):
            try:
                ser = _peers or [(st.session_state.get("px_src") or "대상회사", _own)]
                st.session_state.rep_vol = (
                    f"변동성_산출내역_{dt.date.today()}.xlsx",
                    build_xlsx_vol(ser, tdays=_opt.get("tdays", 250),
                                   drop=_opt.get("drop", True),
                                   pick=_opt.get("pick", "median"),
                                   applied=t.sig,
                                   asof=dt.date.fromisoformat(t.d_base)))
            except Exception as ex:
                st.error(f"만들지 못했습니다 — {ex}")
        rv = st.session_state.get("rep_vol")
        if rv:
            st.download_button(f"{rv[0]}  ({len(rv[1])/1024:,.0f} KB)", rv[1], rv[0],
                               MIME, key="dl_vol", use_container_width=True)

    with rc2:
        st.markdown("**이자율 산출내역**")
        st.caption("입력 곡선 → 부트스트래핑 → 연속복리 현물 → 구간 선도. "
                   "조서 트리 시트 11·12행의 값이 어디서 나왔는지 펼쳐 보여 줍니다.")
        if st.button("이자율 리포트 만들기", use_container_width=True,
                     disabled=not (len(t.rf_curve) >= 2 and len(credit_curve(t)) >= 2)):
            try:
                st.session_state.rep_rate = (
                    f"이자율_산출내역_{dt.date.today()}.xlsx", build_xlsx_rate(t))
            except Exception as ex:
                st.error(f"만들지 못했습니다 — {ex}")
        rr_ = st.session_state.get("rep_rate")
        if rr_:
            st.download_button(f"{rr_[0]}  ({len(rr_[1])/1024:,.0f} KB)", rr_[1], rr_[0],
                               MIME, key="dl_rate", use_container_width=True)
